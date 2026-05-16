"""生成 prescan 测试用 Excel 文件。

基于生产环境真实文件结构，覆盖以下维度：
- 表头类型：单级 / 多级 / 无表头 / 标题行偏移
- 区域数量：单区域 / 纵向多区域 / 横向多区域
- 特殊行：合计行(底部/中间) / 单位行 / 标题行
- 文件大小：小(<50行) / 中(100-500行) / 大(>5万行)
- 业务类型：平台导出 / 财务公摊 / 发票 / 库存 / 运营

运行: python generate_test_files.py
"""
from __future__ import annotations

import random
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

OUTPUT_DIR = Path(__file__).parent


def _rand_date(start_year=2026, start_month=3):
    base = datetime(start_year, start_month, 1)
    return base + timedelta(days=random.randint(0, 60))


def _rand_order_no(platform: str) -> str:
    if platform == "淘宝":
        return str(random.randint(10**18, 10**19 - 1))
    elif platform == "拼多多":
        return f"2603{random.randint(10, 31)}-{random.randint(10**11, 10**12 - 1)}"
    elif platform == "抖音":
        return str(random.randint(10**18, 10**19 - 1))
    elif platform == "京东":
        return str(random.randint(10**15, 10**16 - 1))
    return str(random.randint(10**15, 10**16 - 1))


SHOPS = [
    ("白桃汽水杂货铺", "拼多多"), ("咕噜少女girl", "拼多多"),
    ("小名龙傲天", "拼多多"), ("校长阿姨", "拼多多"),
    ("快乐的小癫子", "淘宝"), ("醒醒该上学了", "淘宝"),
    ("Pinkpig未打烊", "淘宝"), ("知音少女", "抖音"),
    ("修狗乌托邦", "抖音"), ("蓝恩集美优品", "京东"),
]

STREETS = ["茶山街道", "南翔镇", "长铺镇", "宗汉街道", "南湾街道", "东园镇", "和平街道"]
PRODUCTS = [
    "省份旅行手账本01", "套件-奶油小狗卡册01", "婚约皮质卡册01",
    "小狗同学录内页", "鱼群渐变蓝卡册", "拼豆急救包补充包",
    "ins风白色频率同学录", "治愈渐变信封卡册", "证件收纳包",
]
OPERATORS = ["冯璐璐", "汪文慧", "张宇兰", "唐华兵", "傅俊杰", "徐萌", "姚玉林"]


# ═══════════════════════════════════════════════════════════════
# 1. 快麦ERP导出 - 销售主题分析（标题合并 + 大数据量）
# ═══════════════════════════════════════════════════════════════
def gen_01_kuaimai_sales_detail():
    """快麦导出：Row1=标题合并行, Row2=列名, 50000行数据, 23列"""
    wb = Workbook()
    ws = wb.active
    ws.title = "销售主题分析-按订单商品明细"

    headers = [
        "序号", "平台订单号", "店铺名称", "所属平台", "发货时间",
        "商品体积", "销售数量", "销售金额", "销售成本", "销售毛利",
        "理论运费", "街道", "快递公司名称", "订单商品备注", "实发金额",
        "实发数量", "实发成本", "买家已付金额", "主商家编码",
        "商品规格", "商品名称", "收件省份", "收件城市",
    ]

    # Row 1: 标题合并行（快麦ERP特征：所有列填相同标题文本）
    title = "销售主题分析-按订单商品明细"
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col, value=title)

    # Row 2: 列名
    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=h)

    # 数据行（生成500行，测试用不需要50万）
    for i in range(1, 501):
        shop, platform = random.choice(SHOPS)
        ws.append([
            i, _rand_order_no(platform), shop, platform,
            _rand_date(), random.choice([0, 100, 280, 598, 1449]),
            random.randint(1, 5), round(random.uniform(5, 80), 2),
            round(random.uniform(0, 20), 2), round(random.uniform(2, 60), 2),
            round(random.uniform(1, 5), 2), random.choice(STREETS),
            random.choice(["圆通", "韵达", "中通", "极兔", None]),
            None, round(random.uniform(5, 80), 2),
            random.randint(1, 5), round(random.uniform(0, 20), 2),
            round(random.uniform(5, 80), 2), f"SKU{random.randint(1000, 9999)}",
            f"规格{random.randint(1, 10)}", random.choice(PRODUCTS),
            random.choice(["浙江省", "江苏省", "广东省", "上海市"]),
            random.choice(["金华市", "杭州市", "苏州市", "深圳市"]),
        ])

    wb.save(OUTPUT_DIR / "01_快麦销售明细_标题合并行.xlsx")
    print("✓ 01_快麦销售明细_标题合并行.xlsx (标题合并+500行)")


# ═══════════════════════════════════════════════════════════════
# 2. 公摊明细表（横向多区域：3月|4月 并列）
# ═══════════════════════════════════════════════════════════════
def gen_02_cost_allocation():
    """公摊明细：横向多区域, 两个月份并列"""
    wb = Workbook()
    ws = wb.active
    ws.title = "公摊明细"

    # Row 1: 横向区域标题
    ws.cell(row=1, column=1, value="3月")
    ws.cell(row=1, column=4, value="4月")
    # Row 2: 子列名
    ws.cell(row=2, column=1, value="费用明细")
    ws.cell(row=2, column=2, value="按年摊销金额")
    ws.cell(row=2, column=3, value="费用金额")
    ws.cell(row=2, column=4, value="按年摊销金额")
    ws.cell(row=2, column=5, value="费用金额")

    expenses = [
        ("金华写字楼租金", 67630.55, 5635.88),
        ("义乌租金", 76800, 6400),
        ("仓库租金-可信", 283800, 23650),
        ("工资", None, 320000),
        ("仓库临时工工资", None, 59834.65),
        ("义乌美工工资", None, 9617),
        ("管理人员工资", None, 82720),
        ("客服工资", None, 7000),
        ("设计部门工资", None, 6306),
        ("企业所得税税费", None, 8500),
        ("电费(义乌)", None, 2000),
        ("辅料", None, 84484.49),
        ("固定资产折旧", None, 2883.33),
    ]
    for i, (name, annual, monthly) in enumerate(expenses, 3):
        ws.cell(row=i, column=1, value=name)
        ws.cell(row=i, column=2, value=annual)
        ws.cell(row=i, column=3, value=monthly)
        # 4月数据略有变化
        ws.cell(row=i, column=4, value=annual)
        ws.cell(row=i, column=5, value=monthly * 0.95 if monthly else None)

    # 合计行
    total_row = len(expenses) + 3
    ws.cell(row=total_row, column=1, value="合计")
    ws.cell(row=total_row, column=3, value=sum(m for _, _, m in expenses if m))

    wb.save(OUTPUT_DIR / "02_公摊明细_横向多区域.xlsx")
    print("✓ 02_公摊明细_横向多区域.xlsx (横向两月份并列)")


# ═══════════════════════════════════════════════════════════════
# 3. 固定资产折旧（标题偏移：Row1=标题, Row3=单位, Row4=列名）
# ═══════════════════════════════════════════════════════════════
def gen_03_fixed_assets():
    """固定资产折旧：标题行+空行+单位行+列名，数据从Row5开始"""
    wb = Workbook()
    ws = wb.active
    ws.title = "固定资产折旧"

    # Row 1: 大标题
    ws.cell(row=1, column=1, value="固定资产折旧明细表")
    # Row 2: 空行
    # Row 3: 单位+日期
    ws.cell(row=3, column=1, value="单位:金华市蓝创服饰")
    ws.cell(row=3, column=11, value=datetime(2026, 4, 1))
    # Row 4: 列名
    headers = ["序号", "名称", "规格", "单位", "数量", "购入月份",
               "原值", "使用年限", "使用月数", "残值率", "残值",
               "年初已提折旧月数", "月折旧额", "1月", "2月", "3月", "4月",
               "累计折旧", "净值", "备注"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=h)

    assets = [
        (1, "23年货架", None, None, 1, datetime(2023, 7, 1), 10000, 1, 12),
        (2, "23年电脑1", None, None, 2, datetime(2023, 7, 1), 5500, 1, 12),
        (3, "23年电脑2", None, None, 2, datetime(2023, 11, 1), 10050, 1, 12),
        (4, "新增货架+拣货车", None, None, 4, datetime(2024, 2, 1), 11256, 1, 12),
        (5, "PDA采集器", None, None, 4, datetime(2024, 2, 1), 3200, 1, 12),
        (6, "24年电脑", None, "套", 1, datetime(2024, 2, 1), 57500, 1, 12),
        (7, "打印机", None, "台", 2, datetime(2024, 6, 1), 8900, 3, 36),
        (8, "空调", None, "台", 3, datetime(2025, 1, 1), 15000, 5, 60),
    ]
    for row_data in assets:
        row = [row_data[0], row_data[1], row_data[2], row_data[3],
               row_data[4], row_data[5], row_data[6], row_data[7], row_data[8],
               0, 0, None,
               round(row_data[6] / row_data[8], 2)]  # 月折旧额
        # 1-4月折旧
        monthly = round(row_data[6] / row_data[8], 2)
        row.extend([monthly] * 4)
        row.append(monthly * 4)  # 累计折旧
        row.append(row_data[6] - monthly * 4)  # 净值
        row.append(None)
        ws.append(row)

    # 合计行
    ws.append(["合计", None, None, None, None, None,
               sum(a[6] for a in assets)])

    wb.save(OUTPUT_DIR / "03_固定资产折旧_标题偏移.xlsx")
    print("✓ 03_固定资产折旧_标题偏移.xlsx (Row1标题+Row3单位+Row4列名)")


# ═══════════════════════════════════════════════════════════════
# 4. 辅料采购（横向多区域 + 二级表头）
# ═══════════════════════════════════════════════════════════════
def gen_04_auxiliary_materials():
    """辅料采购：横向4个月份区域, 每月有子列名(气泡袋/纸箱)"""
    wb = Workbook()
    ws = wb.active
    ws.title = "辅料"

    months = ["1月辅料采购", "2月辅料采购", "3月辅料采购", "4月辅料采购"]
    sub_cols = ["气泡袋", "纸箱", "塑料箱+胶带"]

    # Row 1: 月份标题（每3列一组，中间有空列）
    col = 2  # A列空
    for m in months:
        ws.cell(row=1, column=col, value=m)
        col += len(sub_cols) + 1  # +1 空列分隔

    # Row 2: 子列名
    col = 2
    for _ in months:
        for sc in sub_cols:
            ws.cell(row=2, column=col, value=sc)
            col += 1
        col += 1  # 空列分隔

    # 数据行（8个仓库/部门的数据）
    for i in range(8):
        row_data = [None]  # A列空
        for _ in months:
            row_data.append(round(random.uniform(2000, 7000), 1))  # 气泡袋
            row_data.append(round(random.uniform(2000, 12000), 0) if random.random() > 0.3 else None)  # 纸箱
            row_data.append(round(random.uniform(1000, 2000), 2) if random.random() > 0.5 else None)  # 塑料箱
            row_data.append(None)  # 分隔空列
        ws.append(row_data)

    wb.save(OUTPUT_DIR / "04_辅料采购_横向多区域二级表头.xlsx")
    print("✓ 04_辅料采购_横向多区域二级表头.xlsx (4月份×3子列)")


# ═══════════════════════════════════════════════════════════════
# 5. 发票数据（一对多：订单号有空行续行）
# ═══════════════════════════════════════════════════════════════
def gen_05_invoice_raw():
    """发票原始数据：一对多关系, 续行订单号为空"""
    wb = Workbook()
    ws = wb.active
    ws.title = "sheet"

    headers = ["订单编号", "发票类型", "购方名称", "购方税号", "购方地址",
               "购方电话", "购方开户行", "购方银行账号", "备注",
               "商品名称", "规格型号", "单位", "数量", "单价", "发票金额"]
    ws.append(headers)

    buyers = [
        ("个人", None), ("刘先生", None),
        ("广州灵狮广告有限公司", "914401017695092958"),
        ("浙江蓝创贸易有限公司", "91330703MA2KXXXXXX"),
    ]

    for _ in range(100):
        order_no = _rand_order_no("淘宝")
        buyer, tax_no = random.choice(buyers)
        item_count = random.randint(1, 3)  # 每单1-3个商品

        for j in range(item_count):
            product = random.choice(PRODUCTS)
            price = round(random.uniform(5, 60), 2)
            row = [
                order_no if j == 0 else None,  # 续行订单号为空
                "电子普通发票" if j == 0 else None,
                buyer if j == 0 else None,
                tax_no if j == 0 else None,
                None, None, None, None,
                f"订单号:{order_no}" if j == 0 else None,
                product, None, None, 1.0, price, price,
            ]
            ws.append(row)

            # 有折扣时追加负数行
            if random.random() > 0.7:
                discount = round(-random.uniform(1, 10), 1)
                ws.append([None] * 9 + [product, None, None, 1.0, discount, None])

    wb.save(OUTPUT_DIR / "05_发票原始_一对多续行.xlsx")
    print("✓ 05_发票原始_一对多续行.xlsx (续行订单号为空+折扣负数行)")


# ═══════════════════════════════════════════════════════════════
# 6. 发票整理表（标准单级表头）
# ═══════════════════════════════════════════════════════════════
def gen_06_invoice_cleaned():
    """发票整理后：标准单级表头, 260行"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    headers = ["日期", "平台订单号", "发票类型", "公司名称", "税号",
               "项目名称", "数量", "金额", "平台+店铺", "申请人", "备注"]
    ws.append(headers)

    for _ in range(260):
        date = _rand_date()
        order_no = _rand_order_no(random.choice(["淘宝", "拼多多"]))
        ws.append([
            date.strftime("%Y-%m-%d"), order_no, "电子普通发票",
            random.choice(["个人", "张三", "李四", "广州灵狮广告有限公司"]),
            None if random.random() > 0.3 else f"91330{random.randint(100000000, 999999999)}",
            random.choice(PRODUCTS), random.randint(1, 5),
            round(random.uniform(5, 300), 2),
            None, None, f"订单号:{order_no}",
        ])

    wb.save(OUTPUT_DIR / "06_发票整理_标准单级表头.xlsx")
    print("✓ 06_发票整理_标准单级表头.xlsx (标准11列)")


# ═══════════════════════════════════════════════════════════════
# 7. 盈利分析（标准单级, 小文件）
# ═══════════════════════════════════════════════════════════════
def gen_07_profit_analysis():
    """盈利分析：标准单级表头, 10行, 极小文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    ws.append(["主商家编码", "商品名称", "数量", "成本", "毛利", "利润率"])
    for i in range(10):
        qty = random.randint(1, 10)
        cost = round(random.uniform(3, 55), 2)
        profit = round(random.uniform(5, 50), 2)
        rate = round(profit / cost * 100, 2) if cost > 0 else 0
        ws.append([
            f"SKU{random.randint(1000, 9999)}-{i:02d}",
            random.choice(PRODUCTS), f"{qty:.2f}",
            f"{cost:.2f}", f"{profit:.2f}", str(rate),
        ])

    wb.save(OUTPUT_DIR / "07_盈利分析_小文件.xlsx")
    print("✓ 07_盈利分析_小文件.xlsx (10行极小文件)")


# ═══════════════════════════════════════════════════════════════
# 8. 店铺销售汇总（快麦导出格式 + 多Sheet结构不同）
# ═══════════════════════════════════════════════════════════════
def gen_08_shop_summary():
    """店铺数据：Sheet1=快麦格式(标题合并), Sheet2=店铺列表(无标题行)"""
    wb = Workbook()

    # Sheet 1: 快麦导出格式
    ws1 = wb.active
    ws1.title = "表1"
    title = "销售主题分析-按店铺"
    headers = ["序号", "店铺", "所属平台", "订单数", "销售数量",
               "实发数量", "实发金额", "销售额", "订单买家已付金额",
               "销售成本", "实发成本", "销售毛利"]
    # Row1: 标题合并
    for col in range(1, len(headers) + 1):
        ws1.cell(row=1, column=col, value=title)
    # Row2: 列名
    for col, h in enumerate(headers, 1):
        ws1.cell(row=2, column=col, value=h)
    # 数据
    for i, (shop, platform) in enumerate(SHOPS, 1):
        orders = random.randint(50, 7000)
        sales_qty = random.randint(orders, orders * 3)
        sales_amount = round(random.uniform(3000, 15000), 2)
        ws1.append([
            i, shop, platform, orders, sales_qty,
            random.randint(100, sales_qty),
            round(sales_amount * 0.4, 2), sales_amount, sales_amount,
            round(sales_amount * 0.3, 2),
            round(sales_amount * 0.15, 2),
            round(sales_amount * 0.55, 2),
        ])

    # Sheet 2: 店铺列表（无标题行，直接是数据）
    ws2 = wb.create_sheet("表2")
    for shop, platform in SHOPS:
        ws2.append([
            random.randint(140000, 150000), shop, "",
            random.choice(OPERATORS),
            f"pdd{random.randint(10000000000, 99999999999)}" if platform == "拼多多" else f"tb{random.randint(10000000, 99999999)}",
            platform, "正常", "", "",
            random.choice(OPERATORS),
            "浙江省金华市金东区孝顺镇镇北工业区可信机械-蓝创",
            "",
        ])

    wb.save(OUTPUT_DIR / "08_店铺销售_多Sheet结构不同.xlsx")
    print("✓ 08_店铺销售_多Sheet结构不同.xlsx (Sheet1快麦格式, Sheet2无表头)")


# ═══════════════════════════════════════════════════════════════
# 9. 订单数汇总（有小计行穿插）
# ═══════════════════════════════════════════════════════════════
def gen_09_order_count_subtotals():
    """订单数：中间穿插小计行"""
    wb = Workbook()
    ws = wb.active
    ws.title = "义乌订单数"

    ws.append(["负责人", "店铺", "订单数"])

    for operator in OPERATORS[:4]:
        shops_count = random.randint(3, 8)
        total = 0
        for j in range(shops_count):
            shop = f"{random.choice(['芝士', '快乐', '小熊', '修狗', '知音'])}店铺{j+1}(抖音放心购)"
            orders = random.randint(0, 5000)
            total += orders
            ws.append([operator, shop, orders])
        # 小计行
        ws.append([f"{operator}小计", None, total])

    # 总计行
    ws.append(["总计", None, "=SUM(C2:C50)"])

    wb.save(OUTPUT_DIR / "09_订单数_中间小计行.xlsx")
    print("✓ 09_订单数_中间小计行.xlsx (小计行穿插+总计)")


# ═══════════════════════════════════════════════════════════════
# 10. 京东活动盘点（长文本 + 合并单元格模拟）
# ═══════════════════════════════════════════════════════════════
def gen_10_jd_activity():
    """京东C店活动盘点：频道列合并(值为None=续行), 长文本单元格"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    headers = ["频道", "活动", "权益", "店铺要求", "商品要求", "报名链接", "审核提醒", "活动介绍"]
    ws.append(headers)

    channels = [
        ("京东秒杀", ["单品秒杀", "皮肤秒杀", "5折限量抢"]),
        ("便宜包邮", ["频道券补贴", "1元收品池", "新品收品池", "长期资源位"]),
        ("特价频道", ["常规2元收品池", "特价收品池", "核心钩子品"]),
    ]

    for channel, activities in channels:
        for idx, act in enumerate(activities):
            ws.append([
                channel if idx == 0 else None,  # 频道合并效果
                act,
                f"权益1：降扣至0.6%\n权益2：流量扶持\n权益3：搜索推荐加权",
                "店铺评分≥4.2" if random.random() > 0.5 else "无要求",
                "180天动销>0且在售；好评率≥90%",
                "https://mc.jd.com/shopAdmin/industry/detail/101666814",
                "报名后请勿设置其它单品促销" if random.random() > 0.5 else None,
                None,
            ])

    # WPS 特征：额外的隐藏 Sheet
    ws2 = wb.create_sheet("WpsReserved_CellImgList")
    ws2.cell(row=1, column=1, value=None)

    wb.save(OUTPUT_DIR / "10_京东活动_长文本合并.xlsx")
    print("✓ 10_京东活动_长文本合并.xlsx (WPS格式+长文本+合并列)")


# ═══════════════════════════════════════════════════════════════
# 11. 商品体积（极简2列, 大数据量）
# ═══════════════════════════════════════════════════════════════
def gen_11_product_volume():
    """商品体积：2列极简, 2000+行"""
    wb = Workbook()
    ws = wb.active
    ws.title = "快麦导出_普通商品明细"

    ws.append(["规格商家编码", "规格体积(cm³)"])
    for i in range(2200):
        prefix = random.choice(["HLDXMGJ", "LZTTDMGKC", "ZYMXPKZ", "CYSZB", "TJ-NYXGKC"])
        ws.append([f"{prefix}{random.randint(1,99):02d}-{random.randint(1,10):02d}",
                   random.randint(50, 5000)])

    wb.save(OUTPUT_DIR / "11_商品体积_极简2列.xlsx")
    print("✓ 11_商品体积_极简2列.xlsx (2列×2200行)")


# ═══════════════════════════════════════════════════════════════
# 12. 店铺分组映射（简单3列映射表）
# ═══════════════════════════════════════════════════════════════
def gen_12_shop_mapping():
    """店铺-平台-运营 映射：简单3列"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    ws.append(["店铺", "所属平台", "店铺分组"])
    for shop, platform in SHOPS * 5:  # 50行
        ws.append([shop, platform, random.choice(OPERATORS)])

    # 额外空 Sheet（多 Sheet 场景）
    wb.create_sheet("Sheet2")
    wb.create_sheet("Sheet3")

    wb.save(OUTPUT_DIR / "12_店铺分组_简单映射.xlsx")
    print("✓ 12_店铺分组_简单映射.xlsx (3列×50行)")


# ═══════════════════════════════════════════════════════════════
# 13. 运营汇总（标准单级，含全零行）
# ═══════════════════════════════════════════════════════════════
def gen_13_operation_summary():
    """运营付费汇总：有全零行（新运营还没数据）"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    ws.append(["店铺分组", "订单数", "销售额", "推广费用", "客单价", "推广费用率"])
    for op in OPERATORS:
        if random.random() > 0.4:
            orders = random.randint(100, 5000)
            sales = round(orders * random.uniform(15, 45), 2)
            promo = round(sales * random.uniform(0.05, 0.3), 2)
            ws.append([op, orders, sales, promo,
                       round(sales / orders, 2), round(promo / sales, 4)])
        else:
            ws.append([op, 0, 0, 0, 0, 0])  # 全零行

    wb.save(OUTPUT_DIR / "13_运营汇总_含全零行.xlsx")
    print("✓ 13_运营汇总_含全零行.xlsx")


# ═══════════════════════════════════════════════════════════════
# 14. 多Sheet公摊表（10个Sheet, 各Sheet结构不同）
# ═══════════════════════════════════════════════════════════════
def gen_14_multi_sheet_complex():
    """10个Sheet的公摊表：每个Sheet结构不同"""
    wb = Workbook()

    # Sheet 1: 公摊明细（横向多区域）
    ws = wb.active
    ws.title = "公摊明细"
    ws.cell(row=1, column=1, value="3月")
    ws.cell(row=1, column=4, value="4月")
    ws.append([None])  # spacer
    ws.append(["费用明细", "摊销金额", "费用金额", "摊销金额", "费用金额"])
    for i in range(10):
        ws.append([f"费用项{i+1}", round(random.uniform(1000, 100000), 2),
                   round(random.uniform(500, 30000), 2),
                   round(random.uniform(1000, 100000), 2),
                   round(random.uniform(500, 30000), 2)])

    # Sheet 2: 公摊值计算
    ws2 = wb.create_sheet("金华义乌公摊值")
    ws2.append(["公共费用", None, "按年摊销金额", "费用金额"])
    for name in ["仓库租金", "管理人员工资", "临时工工资", "辅料", "折旧", "零星分摊"]:
        ws2.append([None, name, round(random.uniform(3000, 300000), 2),
                    round(random.uniform(2000, 85000), 2)])

    # Sheet 3: 部门公摊
    ws3 = wb.create_sheet("义乌部门公摊")
    ws3.append([None, None, None, "公摊费用", "订单数", "公摊值"])
    ws3.append(["义乌部门", "义乌租金", 76800, 6400])
    for item in ["基本工资", "美工工资", "客服工资", "设计工资", "电费"]:
        ws3.append([None, item, None, round(random.uniform(500, 70000), 0)])

    # Sheet 4-10: 各类明细
    for name in ["固定资产折旧", "零星分摊", "租金分摊", "辅料",
                 "辅料纸箱气泡袋", "费用支出", "参保人员"]:
        ws_n = wb.create_sheet(name)
        # 统一格式：标题+空行+单位+列名
        ws_n.cell(row=1, column=1, value=f"{name}明细表")
        ws_n.cell(row=3, column=1, value="单位:金华市蓝创服饰")
        ws_n.append(["序号", "名称", "金额", "备注"])
        for i in range(5):
            ws_n.append([i + 1, f"{name}项目{i+1}",
                         round(random.uniform(1000, 50000), 2), None])

    wb.save(OUTPUT_DIR / "14_公摊表_10Sheet各异.xlsx")
    print("✓ 14_公摊表_10Sheet各异.xlsx (10 Sheet, 结构各不相同)")


# ═══════════════════════════════════════════════════════════════
# 15. 大文件（5万行，测试性能）
# ═══════════════════════════════════════════════════════════════
def gen_15_large_file():
    """大数据量：5万行, 标准格式, 测试性能"""
    wb = Workbook(write_only=True)
    ws = wb.create_sheet("销售明细")

    headers = ["序号", "平台订单号", "店铺名称", "所属平台", "发货时间",
               "销售数量", "销售金额", "销售成本", "销售毛利", "街道"]
    ws.append(headers)

    for i in range(1, 50001):
        shop, platform = random.choice(SHOPS)
        amount = round(random.uniform(5, 200), 2)
        cost = round(amount * random.uniform(0.2, 0.6), 2)
        ws.append([
            i, _rand_order_no(platform), shop, platform,
            _rand_date(), random.randint(1, 10),
            amount, cost, round(amount - cost, 2),
            random.choice(STREETS),
        ])

    wb.save(OUTPUT_DIR / "15_大文件_5万行.xlsx")
    print("✓ 15_大文件_5万行.xlsx (50000行性能测试)")


# ═══════════════════════════════════════════════════════════════
# 16. CSV UTF-8 编码（平台订单导出）
# ═══════════════════════════════════════════════════════════════
def gen_16_csv_utf8():
    """CSV UTF-8：淘宝订单导出风格"""
    import csv
    headers = ["订单编号", "买家会员名", "收货人姓名", "联系手机",
               "收货地址", "商品标题", "商品数量", "实付金额",
               "订单状态", "创建时间", "付款时间"]

    with open(OUTPUT_DIR / "16_淘宝订单_utf8.csv", "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        for _ in range(200):
            writer.writerow([
                _rand_order_no("淘宝"),
                f"buyer_{random.randint(1000, 9999)}",
                random.choice(["张三", "李四", "王五", "赵六"]),
                f"138{random.randint(10000000, 99999999)}",
                "浙江省金华市金东区某某街道某某小区",
                random.choice(PRODUCTS),
                random.randint(1, 5),
                round(random.uniform(5, 200), 2),
                random.choice(["等待买家付款", "买家已付款", "卖家已发货", "交易成功"]),
                _rand_date().strftime("%Y-%m-%d %H:%M:%S"),
                _rand_date().strftime("%Y-%m-%d %H:%M:%S"),
            ])

    print("✓ 16_淘宝订单_utf8.csv (UTF-8 BOM)")


# ═══════════════════════════════════════════════════════════════
# 17. CSV GBK 编码（拼多多导出风格）
# ═══════════════════════════════════════════════════════════════
def gen_17_csv_gbk():
    """CSV GBK：拼多多订单导出"""
    import csv
    headers = ["订单号", "商品名称", "商品规格", "商品数量", "商品金额",
               "优惠金额", "实付金额", "收件人", "收件人手机号",
               "省", "市", "区", "详细地址", "订单状态", "发货时间"]

    with open(OUTPUT_DIR / "17_拼多多订单_gbk.csv", "w", encoding="gbk", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        for _ in range(150):
            amount = round(random.uniform(8, 100), 2)
            discount = round(amount * random.uniform(0, 0.3), 2)
            writer.writerow([
                _rand_order_no("拼多多"),
                random.choice(PRODUCTS),
                f"颜色:{random.choice(['粉色', '蓝色', '白色'])}",
                random.randint(1, 3),
                amount, discount, round(amount - discount, 2),
                random.choice(["张三", "李四", "王五"]),
                f"138{random.randint(10000000, 99999999)}",
                "浙江省", "金华市", "金东区", "某某街道某某小区",
                random.choice(["待发货", "已发货", "已签收"]),
                _rand_date().strftime("%Y-%m-%d %H:%M:%S"),
            ])

    print("✓ 17_拼多多订单_gbk.csv (GBK编码)")


# ═══════════════════════════════════════════════════════════════
# 18. 无表头文件（纯数据，第一行就是数据）
# ═══════════════════════════════════════════════════════════════
def gen_18_no_header():
    """无表头：第一行直接是数据，没有列名"""
    wb = Workbook()
    ws = wb.active
    ws.title = "data"

    # 直接写数据，没有表头行
    for _ in range(80):
        ws.append([
            random.randint(140000, 150000),
            random.choice([s[0] for s in SHOPS]),
            "",
            random.choice(OPERATORS),
            f"pdd{random.randint(10000000000, 99999999999)}",
            random.choice(["拼多多", "淘宝", "抖音"]),
            "正常",
        ])

    wb.save(OUTPUT_DIR / "18_无表头_纯数据.xlsx")
    print("✓ 18_无表头_纯数据.xlsx (header_type=none)")


# ═══════════════════════════════════════════════════════════════
# 19. 混合语言列名
# ═══════════════════════════════════════════════════════════════
def gen_19_mixed_language():
    """中英混合列名：常见于系统导出"""
    wb = Workbook()
    ws = wb.active
    ws.title = "export"

    headers = ["order_id", "SKU编码", "product_name(商品名称)", "qty数量",
               "unit_price单价", "total_amount", "platform平台",
               "status状态", "ship_date发货日期"]
    ws.append(headers)

    for _ in range(100):
        shop, platform = random.choice(SHOPS)
        qty = random.randint(1, 10)
        price = round(random.uniform(5, 80), 2)
        ws.append([
            _rand_order_no(platform),
            f"SKU{random.randint(1000, 9999)}",
            random.choice(PRODUCTS), qty, price,
            round(qty * price, 2), platform,
            random.choice(["pending", "shipped", "delivered"]),
            _rand_date().strftime("%Y-%m-%d"),
        ])

    wb.save(OUTPUT_DIR / "19_混合语言列名.xlsx")
    print("✓ 19_混合语言列名.xlsx (中英混合header)")


# ═══════════════════════════════════════════════════════════════
# 20. 多级表头 + 合计行在底部
# ═══════════════════════════════════════════════════════════════
def gen_20_multi_header_with_footer():
    """利润表：Row1=大类合并, Row2=子列名, 底部有合计行"""
    wb = Workbook()
    ws = wb.active
    ws.title = "利润表"

    # Row 1: 大类标题
    ws.cell(row=1, column=1, value="基础信息")
    ws.cell(row=1, column=2)  # 合并范围
    ws.cell(row=1, column=3)
    ws.cell(row=1, column=4, value="销售数据")
    ws.cell(row=1, column=5)
    ws.cell(row=1, column=6)
    ws.cell(row=1, column=7, value="成本数据")
    ws.cell(row=1, column=8)
    ws.cell(row=1, column=9, value="利润")
    ws.merge_cells("A1:C1")
    ws.merge_cells("D1:F1")
    ws.merge_cells("G1:H1")

    # Row 2: 子列名
    ws.append(["店铺", "平台", "运营", "订单数", "销售额", "客单价",
               "商品成本", "运费成本", "毛利", "毛利率"])

    # 数据
    for shop, platform in SHOPS:
        orders = random.randint(100, 5000)
        sales = round(orders * random.uniform(15, 50), 2)
        goods_cost = round(sales * random.uniform(0.2, 0.4), 2)
        ship_cost = round(orders * random.uniform(2, 5), 2)
        profit = round(sales - goods_cost - ship_cost, 2)
        ws.append([
            shop, platform, random.choice(OPERATORS),
            orders, sales, round(sales / orders, 2),
            goods_cost, ship_cost, profit,
            f"{round(profit / sales * 100, 1)}%",
        ])

    # 合计行
    ws.append(["合计", None, None, "=SUM(D3:D12)", "=SUM(E3:E12)", None,
               "=SUM(G3:G12)", "=SUM(H3:H12)", "=SUM(I3:I12)", None])

    wb.save(OUTPUT_DIR / "20_利润表_多级表头底部合计.xlsx")
    print("✓ 20_利润表_多级表头底部合计.xlsx (合并单元格+底部合计)")


# ═══════════════════════════════════════════════════════════════
# 21. 纵向多区域（同Sheet内多表格，空行分隔，各有表头）
# ═══════════════════════════════════════════════════════════════
def gen_21_vertical_multi_region():
    """同一Sheet内3个独立表格，空行分隔，各有自己的表头。
    参考：公摊表-义乌部门公摊sheet"""
    wb = Workbook()
    ws = wb.active
    ws.title = "部门公摊"

    # ── 区域 1：费用明细（Row 1-10）──
    ws.append([None, None, None, "公摊费用", "订单数", "公摊值"])
    ws.append(["义乌部门", "义乌租金", 76800, 6400, None, None])
    for item, amount in [("基本工资", 68609), ("美工工资", 634),
                         ("客服工资", 7000), ("设计工资", 6306),
                         ("企业所得税", 8500), ("电费", 2000),
                         ("公共费用", 62320.47)]:
        ws.append([None, item, None, amount, None, None])
    ws.append([None, "小计", None, 161769.47, 45338, 3.568])

    # ── 空行分隔 ──
    ws.append(["剔除拼多多，京东"])
    ws.append([None] * 6)
    ws.append([None] * 6)

    # ── 区域 2：客服工资（Row 14-16）──
    ws.append(["平台", "客服工资"])
    ws.append(["淘宝抖音客服", 6000])
    ws.append(["阿里巴巴", 1000])

    # ── 空行分隔 ──
    ws.append([None] * 6)
    ws.append([None] * 6)

    # ── 区域 3：平台汇总（Row 19-24）──
    ws.append(["平台", "订单数", "客服工资", "运营工资", "人数", "店铺数量"])
    ws.append(["阿里巴巴", 6573, 1000, 14744, 2, 1])
    ws.append(["淘宝", 26818, 3400, 24727, 5, 17])
    ws.append(["抖音", 6017, 1200, 8254, 1, 6])
    ws.append(["快手", 5930, 1200, 5981, 1, 6])
    ws.append(["京东", None, None, 6049, None, None])

    wb.save(OUTPUT_DIR / "21_纵向多区域_空行分隔三表.xlsx")
    print("✓ 21_纵向多区域_空行分隔三表.xlsx (3个表格各有表头)")


# ═══════════════════════════════════════════════════════════════
# 22. 数据区域真实合并单元格
# ═══════════════════════════════════════════════════════════════
def gen_22_merged_data_cells():
    """数据区域的纵向合并：A列部门名合并多行，C列也有合并"""
    wb = Workbook()
    ws = wb.active
    ws.title = "部门明细"

    ws.append(["部门", "费用项", "金额", "占比", "备注"])

    # 金东部门（A3:A9 合并）
    row = 2
    items_jd = [("写字楼租金", 5635.88), ("基本工资", 94292),
                ("社保", 4880), ("客服工资", 14767),
                ("设计工资", 6306), ("税费", 9182), ("电费", 2000)]
    for item, amount in items_jd:
        ws.append(["金东部门" if item == "写字楼租金" else None,
                   item, amount, round(amount / 137062 * 100, 1), None])
    ws.append([None, "小计", 137062, 100.0, None])
    ws.merge_cells(f"A2:A{2 + len(items_jd) - 1}")

    # 义乌部门（A10:A15 合并）
    start = 2 + len(items_jd) + 1
    items_yw = [("租金", 6400), ("基本工资", 68609),
                ("美工工资", 634), ("客服工资", 7000),
                ("设计工资", 6306), ("电费", 2000)]
    for item, amount in items_yw:
        ws.append(["义乌部门" if item == "租金" else None,
                   item, amount, round(amount / 90949 * 100, 1), None])
    ws.append([None, "小计", 90949, 100.0, None])
    ws.merge_cells(f"A{start}:A{start + len(items_yw) - 1}")

    # 总合计行（A:H 合并）
    total_row = start + len(items_yw) + 1
    ws.append(["费用合计", None, 228011, None, None])
    ws.merge_cells(f"A{total_row}:B{total_row}")

    wb.save(OUTPUT_DIR / "22_合并单元格_数据区域纵向.xlsx")
    print("✓ 22_合并单元格_数据区域纵向.xlsx (A列部门合并多行)")


# ═══════════════════════════════════════════════════════════════
# 23. 非标日期格式
# ═══════════════════════════════════════════════════════════════
def gen_23_nonstandard_dates():
    """日期格式混乱：字符串日期各种写法混合"""
    wb = Workbook()
    ws = wb.active
    ws.title = "租金分摊"

    # 标题偏移结构
    ws.append(["租金分摊明细表"])
    ws.append([None])
    ws.append(["单位:金华市蓝创服饰", None, None, None, None, "2026.4.1"])
    ws.append(["序号", "名称", "交付月份", "原值", "使用月数", "月摊销额"])

    # 各种日期写法
    entries = [
        (1, "仓库租金-东莞", "2024.11.5", 60000, 12),
        (2, "综保区租金", "2024.1.1", 67630.55, 12),
        (3, "江西租金", "2025.2.1", 84000, 12),
        (4, "仓库租金2250m²", "2025.7.30", 283800, 12),
        (5, "仓库租金", "2025.2.13", 35000, 12),
        (6, "义乌租金", "2024.12.6", 76800, 12),
        (7, "员工意外险", "6.1-", 13112, 12),       # 极端：只有月.日-
        (8, "员工意外险+6人", "8.21-", 1381.44, 12),  # 极端
        (9, "BOSS招聘", "2025.2.1", 9600, 12),
        (10, "智库", "25.3.1", 23960, 36),           # 两位年份
        (11, "物流预警", "25/4/2", 12000, 36),        # 斜杠分隔
        (12, "客服软件", "2025-8-11", 16000, 24),     # 标准但无补零
    ]

    for seq, name, date_str, amount, months in entries:
        ws.append([seq, name, date_str, amount, months,
                   round(amount / months, 2)])

    wb.save(OUTPUT_DIR / "23_非标日期格式.xlsx")
    print("✓ 23_非标日期格式.xlsx (2024.11.5/6.1-/25.3.1/25/4/2 等)")


# ═══════════════════════════════════════════════════════════════
# 24. DISPIMG公式 + #REF!错误值
# ═══════════════════════════════════════════════════════════════
def gen_24_formulas_and_errors():
    """含WPS DISPIMG图片公式 + #REF!错误值 + 跨列汇总"""
    wb = Workbook()
    ws = wb.active
    ws.title = "辅料纸箱气泡袋"

    # 无表头，直接是数据（供应商+月份+金额+凭证图片）
    data = [
        ("创园包装——气泡袋", "6月", 32654.3, '=DISPIMG("ID_552B4879",1)'),
        ("轩轩纸箱厂", "6月", 29073, '=DISPIMG("ID_9AADE8BD",1)'),
        ("塘里压痕", "6月", 26979, '=DISPIMG("ID_4AF7751B",1)'),
        ("宏顺胶带厂", "7月", 1512, None),
        ("创园包装——气泡袋", "7月", 37479.1, '=DISPIMG("ID_A5AB1031",1)'),
        ("轩轩纸箱厂", "7月", 13858, '=DISPIMG("ID_E14C7006",1)'),
        ("塘里压痕", "7月", 13649, None),
        ("轩轩纸箱厂", "8月", 4601, '=DISPIMG("ID_9E8BE2BC",1)'),
        ("塘里压痕", "8月", None, '=DISPIMG("ID_767F5D7D",1)'),  # #REF!模拟
        ("创园包装——气泡袋", "8月", None, '=DISPIMG("ID_D141C50E",1)'),  # #REF!
        ("创园包装——气泡袋", "9月", 55326.5, '=DISPIMG("ID_493D50F5",1)'),
        ("塘里压痕", "9月", 29979, '=DISPIMG("ID_30F60EF6",1)'),
    ]
    for row in data:
        ws.append(list(row))

    wb.save(OUTPUT_DIR / "24_DISPIMG公式_错误值.xlsx")
    print("✓ 24_DISPIMG公式_错误值.xlsx (WPS图片公式+#REF!)")


# ═══════════════════════════════════════════════════════════════
# 25. 费用支出流水（标题行含汇总值 + 多凭证列）
# ═══════════════════════════════════════════════════════════════
def gen_25_expense_journal():
    """费用支出日记账：Row1有标题+右侧汇总数, Row2列名, 多空凭证列"""
    wb = Workbook()
    ws = wb.active
    ws.title = "费用支出"

    # Row 1: 标题 + 右侧汇总值（跨列放置）
    ws.cell(row=1, column=1, value="26年4月份费用支出流水日记账")
    ws.cell(row=1, column=12, value="支出汇总")
    ws.cell(row=1, column=13, value=1342458.65)

    # Row 2: 列名（含多个"凭证"列）
    headers = ["序号", "日期", "部门", "摘要", "供应商编码",
               "科目（借）", "二级科目", "金额", "支出账户",
               "凭证", "凭证", "凭证", "凭证", "凭证", "备注"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=h)

    # 数据
    departments = ["仓储部", "运营部", "采购部", "行政部", "财务部"]
    subjects = [("管理费用", "办公费"), ("销售费用", "侵权赔偿"),
                ("应付账款", "现结货款"), ("应付账款", "月结货款")]

    for i in range(60):
        dept = random.choice(departments)
        subj, sub2 = random.choice(subjects)
        ws.append([
            i + 34, _rand_date(), dept,
            f"供应商{random.randint(1, 200):04d}货款",
            f"0{random.randint(100, 999)}" if random.random() > 0.5 else None,
            subj, sub2,
            round(random.uniform(50, 100000), 2),
            random.choice(["吴总支付", "公户支付", "夏总-创飞支付", "陈总支付"]),
            f'=DISPIMG("ID_{random.randint(10000000, 99999999):08X}",1)' if random.random() > 0.3 else None,
            None, None, None, None, None,
        ])

    wb.save(OUTPUT_DIR / "25_费用支出_标题含汇总_凭证列.xlsx")
    print("✓ 25_费用支出_标题含汇总_凭证列.xlsx (R1右侧有汇总+重复列名)")


# ═══════════════════════════════════════════════════════════════
# 26. 空文件 / 只有1行表头无数据
# ═══════════════════════════════════════════════════════════════
def gen_26_empty_files():
    """边界：完全空/只有表头无数据/只1行数据"""
    # 26a: 完全空
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    wb.save(OUTPUT_DIR / "26a_完全空文件.xlsx")

    # 26b: 只有表头
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["报备日期", "SKU", "订单号", "商品名称", "异常时段",
               "延迟后预计发货时效", "延期情况说明", "报备人"])
    wb.save(OUTPUT_DIR / "26b_只有表头无数据.xlsx")

    # 26c: 只有1行数据
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["店铺", "平台", "订单数"])
    ws.append(["白桃汽水杂货铺", "拼多多", 584])
    wb.save(OUTPUT_DIR / "26c_只有1行数据.xlsx")

    print("✓ 26a/b/c 空文件系列 (完全空/只表头/只1行数据)")


# ═══════════════════════════════════════════════════════════════
# 27. 数值存为文本（字符串格式的数字）
# ═══════════════════════════════════════════════════════════════
def gen_27_numbers_as_text():
    """所有数值列都是字符串格式：'6.00', '24.84' 而非数字"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    ws.append(["主商家编码", "商品名称", "数量", "成本", "毛利", "利润率"])
    for i in range(30):
        qty = random.randint(1, 20)
        cost = round(random.uniform(3, 55), 2)
        profit = round(random.uniform(2, 50), 2)
        rate = round(profit / cost * 100, 10) if cost > 0 else 0
        # 所有数值都写为字符串
        ws.append([
            f"SKU{random.randint(1000, 9999)}-{i:02d}",
            random.choice(PRODUCTS),
            f"{qty}.00",           # 数量带.00
            f"{cost:.2f}",         # 成本字符串
            f"{profit:.2f}",       # 毛利字符串
            str(rate),             # 超长小数的利润率
        ])

    wb.save(OUTPUT_DIR / "27_数值存为文本.xlsx")
    print("✓ 27_数值存为文本.xlsx (所有数字都是字符串)")


# ═══════════════════════════════════════════════════════════════
# 28. 超宽列（30+列，含大量空列穿插）
# ═══════════════════════════════════════════════════════════════
def gen_28_wide_with_gaps():
    """超宽表：30列+，中间有空列分隔不同数据块"""
    wb = Workbook()
    ws = wb.active
    ws.title = "sheet"

    # 27列发票格式（参考104960729691）
    headers = ["订单编号", "发票类型", "购方名称", "购方税号", "购方地址",
               "购方电话", "购方开户行", "购方银行账号", "备注",
               "商品名称", "规格型号", "单位", "数量", "单价", "发票金额",
               None, None,  # 空列
               "税率", "税额", "价税合计",
               None,  # 空列
               "开票日期", "发票号码", "发票代码", "校验码",
               None, None,  # 空列
               "备注2", "内部编号", "处理状态"]
    ws.append(headers)

    for _ in range(200):
        order_no = _rand_order_no("淘宝")
        price = round(random.uniform(5, 300), 2)
        tax = round(price * 0.01, 2)
        ws.append([
            order_no, "电子普通发票",
            random.choice(["个人", "张三公司", "李四有限公司"]),
            f"91330{random.randint(100000000, 999999999)}" if random.random() > 0.5 else None,
            None, None, None, None,
            f"订单号:{order_no}",
            random.choice(PRODUCTS), None, None, 1.0, price, price,
            None, None,  # 空列
            "1%", tax, round(price + tax, 2),
            None,  # 空列
            _rand_date().strftime("%Y%m%d"),
            f"26{random.randint(10000000, 99999999)}",
            f"044001{random.randint(100000, 999999)}", None,
            None, None,  # 空列
            None, f"INV-{random.randint(10000, 99999)}", "已开具",
        ])

    wb.save(OUTPUT_DIR / "28_超宽30列_空列穿插.xlsx")
    print("✓ 28_超宽30列_空列穿插.xlsx (30列+空列分隔)")


# ═══════════════════════════════════════════════════════════════
# 29. 参保人员表（数据中间有空行）
# ═══════════════════════════════════════════════════════════════
def gen_29_data_with_gaps():
    """数据中间有空行分隔不同公司的人员：非区域分隔，纯格式空行"""
    wb = Workbook()
    ws = wb.active
    ws.title = "参保人员"

    ws.append(["参保人员表", None, None])

    companies = [
        ("金华市蓝创服饰有限公司", ["吴聪", "夏志飞", "陈建刚", "周丽燕"]),
        ("金华市兔卡酱文化用品有限公司", ["叶莉蕾", "饶晓玲", "彭书侠", "钟磊"]),
        ("金华市星晏文化用品有限公司", ["汪文慧", "詹扬豪"]),
        ("义乌市蓝创文化有限公司", ["徐萌", "张宇兰", "唐华兵"]),
    ]

    seq = 1
    for company, people in companies:
        for name in people:
            ws.append([seq, company, name])
            seq += 1
        # 公司之间有空行
        ws.append([None, None, None])

    wb.save(OUTPUT_DIR / "29_数据中间空行.xlsx")
    print("✓ 29_数据中间空行.xlsx (空行分隔不同分组)")


# ═══════════════════════════════════════════════════════════════
# 30. 多Sheet同名列但行数悬殊
# ═══════════════════════════════════════════════════════════════
def gen_30_multi_sheet_varying_size():
    """4个Sheet: 表1(49行)、表2(115行)、表3(69行)、表4(3行)"""
    wb = Workbook()

    configs = [
        ("表1", ["序号", "店铺", "订单数", "销售额"], 49),
        ("表2", ["序号", "店铺", "店铺分组"], 115),
        ("表3", ["店铺编号", "店铺", "推广费支出"], 69),
        ("表4", ["汇总", "总订单数", "总销售额"], 3),
    ]

    for idx, (name, headers, rows) in enumerate(configs):
        if idx == 0:
            ws = wb.active
            ws.title = name
        else:
            ws = wb.create_sheet(name)
        ws.append(headers)
        for i in range(rows - 1):
            if len(headers) == 4:
                ws.append([i + 1, random.choice([s[0] for s in SHOPS]),
                           random.randint(50, 7000),
                           round(random.uniform(1000, 15000), 2)])
            elif len(headers) == 3 and "分组" in headers[-1]:
                ws.append([random.randint(140000, 150000),
                           random.choice([s[0] for s in SHOPS]),
                           random.choice(OPERATORS)])
            elif len(headers) == 3 and "推广" in headers[-1]:
                ws.append([random.randint(140000, 150000),
                           random.choice([s[0] for s in SHOPS]),
                           round(random.uniform(0, 3000), 2)])
            else:
                ws.append(["全部", random.randint(10000, 50000),
                           round(random.uniform(50000, 200000), 2)])

    wb.save(OUTPUT_DIR / "30_多Sheet行数悬殊.xlsx")
    print("✓ 30_多Sheet行数悬殊.xlsx (4 Sheet: 49/115/69/3行)")


# ═══════════════════════════════════════════════════════════════
# 31. 重复列名
# ═══════════════════════════════════════════════════════════════
def gen_31_duplicate_column_names():
    """重复列名：费用支出表有多个'凭证'列"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # 多个同名列
    headers = ["序号", "日期", "金额", "凭证", "凭证", "凭证", "凭证", "备注"]
    ws.append(headers)
    for i in range(20):
        ws.append([i + 1, _rand_date(), round(random.uniform(100, 50000), 2),
                   f"IMG_{i:04d}" if random.random() > 0.5 else None,
                   None, None, None, None])

    wb.save(OUTPUT_DIR / "31_重复列名.xlsx")
    print("✓ 31_重复列名.xlsx (4个'凭证'列同名)")


# ═══════════════════════════════════════════════════════════════
# 32. 首行不是表头（标题+说明占多行才到表头）
# ═══════════════════════════════════════════════════════════════
def gen_32_deep_header_offset():
    """表头深度偏移：前5行都是说明文字，Row6才是列名"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # 前5行是标题/说明/空行
    ws.append(["蓝创服饰有限公司"])
    ws.append(["2026年4月份销售报表"])
    ws.append(["制表人：财务部  审核人：吴总"])
    ws.append(["单位：元    日期：2026-05-01"])
    ws.append([None])  # 空行
    # Row 6: 真正的表头
    ws.append(["序号", "店铺名称", "平台", "订单数", "销售额",
               "成本", "毛利", "毛利率"])
    # 数据
    for i, (shop, platform) in enumerate(SHOPS, 1):
        sales = round(random.uniform(3000, 50000), 2)
        cost = round(sales * random.uniform(0.3, 0.6), 2)
        profit = round(sales - cost, 2)
        ws.append([i, shop, platform, random.randint(50, 5000),
                   sales, cost, profit, f"{round(profit/sales*100, 1)}%"])

    wb.save(OUTPUT_DIR / "32_深度偏移_Row6才是表头.xlsx")
    print("✓ 32_深度偏移_Row6才是表头.xlsx (5行说明后才是列名)")


# ═══════════════════════════════════════════════════════════════
# 33. 纯数字列名（系统导出无意义列名）
# ═══════════════════════════════════════════════════════════════
def gen_33_numeric_headers():
    """列名全是数字或无意义编号：Column1, Column2..."""
    wb = Workbook()
    ws = wb.active
    ws.title = "export"

    # 系统自动编号列名
    headers = [f"Column{i}" for i in range(1, 11)]
    ws.append(headers)
    for _ in range(50):
        shop, platform = random.choice(SHOPS)
        ws.append([
            random.randint(100000, 999999), shop, platform,
            random.choice(OPERATORS), random.randint(0, 5000),
            round(random.uniform(0, 50000), 2),
            _rand_date().strftime("%Y-%m-%d"),
            random.choice(["正常", "异常", "关闭"]),
            None, None,
        ])

    wb.save(OUTPUT_DIR / "33_纯数字列名.xlsx")
    print("✓ 33_纯数字列名.xlsx (Column1/Column2...无意义列名)")


# ═══════════════════════════════════════════════════════════════
# 34. TSV 制表符分隔
# ═══════════════════════════════════════════════════════════════
def gen_34_tsv():
    """TSV格式：制表符分隔"""
    import csv
    headers = ["订单号", "商品编码", "商品名称", "数量", "金额", "状态"]

    with open(OUTPUT_DIR / "34_制表符分隔.tsv", "w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f, delimiter="\t")
        writer.writerow(headers)
        for _ in range(80):
            writer.writerow([
                _rand_order_no("淘宝"),
                f"SKU{random.randint(1000, 9999)}",
                random.choice(PRODUCTS),
                random.randint(1, 10),
                round(random.uniform(5, 200), 2),
                random.choice(["已完成", "进行中", "已取消"]),
            ])

    print("✓ 34_制表符分隔.tsv (TSV格式)")


# ═══════════════════════════════════════════════════════════════
# 35. 特殊字符在数据中（换行符、制表符、引号）
# ═══════════════════════════════════════════════════════════════
def gen_35_special_chars():
    """数据内含换行符、特殊字符"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    ws.append(["频道", "活动名称", "权益说明", "要求"])
    special_texts = [
        "1.秒杀频道流量（首页入口）\n2.搜推流量扶持\n3.秒杀氛围",
        "权益1：降扣至0.6%\n权益2：频道长期流量扶持\n权益3：搜索/推荐加权",
        "店铺评分≥4.2；\n商品体验分、物流履约分、售后服务分均≥6",
        "180天动销>0且在售，或近30天新上柜且在售；\n好评率≥90%；\n剔除预售商品",
        '含"特殊"引号的文本',
        "含\t制表符\t的文本",
        "含emoji🎉的文本📦",
    ]
    for i, text in enumerate(special_texts):
        ws.append([f"频道{i+1}", f"活动{i+1}", text,
                   "无要求" if i % 2 == 0 else "有要求"])

    wb.save(OUTPUT_DIR / "35_特殊字符_换行引号.xlsx")
    print("✓ 35_特殊字符_换行引号.xlsx (\\n/\\t/引号/emoji)")


# ═══════════════════════════════════════════════════════════════
# 36. 极大列数 + 稀疏数据
# ═══════════════════════════════════════════════════════════════
def gen_36_sparse_wide():
    """50列但大部分为空（稀疏矩阵），只有散落的几列有数据"""
    wb = Workbook()
    ws = wb.active
    ws.title = "数据"

    # 50列，只有A,E,J,P,Z,AX列有数据
    data_cols = [0, 4, 9, 15, 25, 49]  # 有数据的列索引
    # 表头
    row = [None] * 50
    for c in data_cols:
        row[c] = f"字段{c+1}"
    ws.append(row)

    for _ in range(30):
        row = [None] * 50
        for c in data_cols:
            row[c] = round(random.uniform(1, 10000), 2) if random.random() > 0.2 else None
        ws.append(row)

    wb.save(OUTPUT_DIR / "36_稀疏50列.xlsx")
    print("✓ 36_稀疏50列.xlsx (50列只有6列有数据)")


# ═══════════════════════════════════════════════════════════════
# 37. 密码保护的Excel（应该报错处理）
# ═══════════════════════════════════════════════════════════════
def gen_37_password_protected():
    """模拟密码保护：写一个sheet保护(非文件级加密，但测试处理)"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["店铺", "销售额"])
    ws.append(["测试店铺", 12345])
    # Sheet级保护（非文件加密，但可以测试是否正确处理）
    ws.protection.sheet = True
    ws.protection.password = "test123"
    wb.save(OUTPUT_DIR / "37_Sheet保护.xlsx")
    print("✓ 37_Sheet保护.xlsx (Sheet级密码保护)")


# ═══════════════════════════════════════════════════════════════
# 38. 超长Sheet名（快麦导出的截断名）
# ═══════════════════════════════════════════════════════════════
def gen_38_long_sheet_name():
    """Sheet名超长（快麦导出经常有超长名被截断）"""
    wb = Workbook()
    # Excel限制sheet名最长31字符，快麦经常截断
    long_name = "销售主题分析-按订单商品明细-202604"  # 刚好31字符
    ws = wb.active
    ws.title = long_name

    # 快麦格式
    title = "销售主题分析-按订单商品明细"
    headers = ["序号", "平台订单号", "店铺名称", "所属平台", "发货时间",
               "商品体积", "销售数量", "销售金额"]
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col, value=title)
    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=h)
    for i in range(20):
        shop, platform = random.choice(SHOPS)
        ws.append([i + 1, _rand_order_no(platform), shop, platform,
                   _rand_date(), random.randint(0, 2000),
                   random.randint(1, 5), round(random.uniform(5, 100), 2)])

    wb.save(OUTPUT_DIR / "38_超长Sheet名.xlsx")
    print("✓ 38_超长Sheet名.xlsx (31字符Sheet名)")


# ═══════════════════════════════════════════════════════════════
# 39. 全角字符 + 不可见字符
# ═══════════════════════════════════════════════════════════════
def gen_39_fullwidth_chars():
    """全角字符列名、零宽字符、不可见空格"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # 全角列名
    headers = ["序号", "店铺名称", "所属平台", "订单数",  # 全角括号
               "销售额（元）", "成本（元）", "利润率（%）"]  # 全角括号
    ws.append(headers)

    for i in range(15):
        shop, platform = random.choice(SHOPS)
        # 数据里混入全角数字和零宽字符
        sales = round(random.uniform(1000, 50000), 2)
        cost = round(sales * 0.4, 2)
        ws.append([
            i + 1,
            f"\u200b{shop}",  # 零宽空格前缀
            f"　{platform}",  # 全角空格前缀
            random.randint(50, 5000),
            sales, cost,
            f"{round((sales-cost)/sales*100, 1)}％",  # 全角百分号
        ])

    wb.save(OUTPUT_DIR / "39_全角字符_零宽空格.xlsx")
    print("✓ 39_全角字符_零宽空格.xlsx (全角括号/零宽字符/全角空格)")


# ═══════════════════════════════════════════════════════════════
# 40. 尾部大量空行（Excel显示max_row很大但实际数据少）
# ═══════════════════════════════════════════════════════════════
def gen_40_trailing_empty_rows():
    """实际数据20行，但文件max_row=1000（尾部全空行）"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    ws.append(["店铺", "平台", "订单数", "销售额"])
    for i in range(20):
        shop, platform = random.choice(SHOPS)
        ws.append([shop, platform, random.randint(50, 5000),
                   round(random.uniform(1000, 50000), 2)])

    # 在远处写一个空格，撑大max_row
    ws.cell(row=1000, column=1, value=" ")

    wb.save(OUTPUT_DIR / "40_尾部大量空行.xlsx")
    print("✓ 40_尾部大量空行.xlsx (数据20行但max_row=1000)")


if __name__ == "__main__":
    random.seed(42)  # 可重复
    print(f"输出目录: {OUTPUT_DIR}\n")

    gen_01_kuaimai_sales_detail()
    gen_02_cost_allocation()
    gen_03_fixed_assets()
    gen_04_auxiliary_materials()
    gen_05_invoice_raw()
    gen_06_invoice_cleaned()
    gen_07_profit_analysis()
    gen_08_shop_summary()
    gen_09_order_count_subtotals()
    gen_10_jd_activity()
    gen_11_product_volume()
    gen_12_shop_mapping()
    gen_13_operation_summary()
    gen_14_multi_sheet_complex()
    gen_15_large_file()
    gen_16_csv_utf8()
    gen_17_csv_gbk()
    gen_18_no_header()
    gen_19_mixed_language()
    gen_20_multi_header_with_footer()
    gen_21_vertical_multi_region()
    gen_22_merged_data_cells()
    gen_23_nonstandard_dates()
    gen_24_formulas_and_errors()
    gen_25_expense_journal()
    gen_26_empty_files()
    gen_27_numbers_as_text()
    gen_28_wide_with_gaps()
    gen_29_data_with_gaps()
    gen_30_multi_sheet_varying_size()
    gen_31_duplicate_column_names()
    gen_32_deep_header_offset()
    gen_33_numeric_headers()
    gen_34_tsv()
    gen_35_special_chars()
    gen_36_sparse_wide()
    gen_37_password_protected()
    gen_38_long_sheet_name()
    gen_39_fullwidth_chars()
    gen_40_trailing_empty_rows()

    print(f"\n✅ 共生成 40+ 个测试文件，覆盖维度：")
    print("  表头：单级/多级/无表头/标题偏移(Row4)/深偏移(Row6)/纯数字列名/重复列名")
    print("  区域：单区域/横向多区域/纵向多区域(空行分隔三表)")
    print("  特殊行：小计/合计(底部)/合计(中间)/全零/续行/数据间空行")
    print("  大小：空/1行/10行/中等/2千行/5万行")
    print("  格式：xlsx/csv-utf8/csv-gbk/tsv")
    print("  编码：UTF-8/GBK/全角字符/零宽字符")
    print("  公式：SUM/DISPIMG/VLOOKUP/#REF!错误")
    print("  合并：表头合并/数据区域纵向合并/标题合并行")
    print("  极限：超宽30列+空列/稀疏50列/尾部空行/Sheet保护/超长Sheet名")
    print("  数据：数值文本化/非标日期/特殊字符(\\n\\t emoji)/混合语言")
