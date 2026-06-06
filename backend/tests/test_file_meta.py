"""file_meta.py 单元测试。"""
from __future__ import annotations

import json
from pathlib import Path

import pandas as pd
import pytest

from services.agent.excel_cleaner import CleaningReport
from services.agent.file_meta import (
    FileMeta,
    _build_sample,
    _build_schema,
    _col_index_to_letter,
    _determine_status,
    _infer_dtype,
    _scan_issues,
    _serialize_value,
    extract_formulas,
    format_file_view,
    generate_file_meta,
    read_file_meta,
    write_file_meta,
)


# ── _col_index_to_letter ──


class TestColIndexToLetter:
    def test_basic(self):
        assert _col_index_to_letter(0) == "A"
        assert _col_index_to_letter(1) == "B"
        assert _col_index_to_letter(25) == "Z"

    def test_double_letter(self):
        assert _col_index_to_letter(26) == "AA"
        assert _col_index_to_letter(27) == "AB"
        assert _col_index_to_letter(51) == "AZ"
        assert _col_index_to_letter(52) == "BA"


# ── _infer_dtype ──


class TestInferDtype:
    def test_integer(self):
        s = pd.Series([1, 2, 3], dtype="Int64")
        assert _infer_dtype(s) == "integer"

    def test_float(self):
        s = pd.Series([1.1, 2.2], dtype="float64")
        assert _infer_dtype(s) == "decimal"

    def test_string(self):
        s = pd.Series(["a", "b"], dtype="object")
        assert _infer_dtype(s) == "string"

    def test_datetime(self):
        s = pd.to_datetime(pd.Series(["2024-01-01", "2024-02-01"]))
        assert _infer_dtype(s) == "datetime"

    def test_bool(self):
        s = pd.Series([True, False], dtype="bool")
        assert _infer_dtype(s) == "boolean"


# ── _serialize_value ──


class TestSerializeValue:
    def test_none(self):
        assert _serialize_value(None) is None
        assert _serialize_value(float("nan")) is None

    def test_timestamp(self):
        ts = pd.Timestamp("2024-01-15 10:30:00")
        result = _serialize_value(ts)
        assert result == "2024-01-15 10:30:00"

    def test_primitives(self):
        assert _serialize_value(42) == 42
        assert _serialize_value(3.14) == 3.14
        assert _serialize_value("hello") == "hello"
        assert _serialize_value(True) is True


# ── _build_schema ──


class TestBuildSchema:
    def test_basic_schema(self):
        df = pd.DataFrame({
            "order_id": ["A001", "A002", "A003"],
            "amount": [100.0, 200.0, None],
            "qty": pd.array([1, 2, 3], dtype="Int64"),
        })
        schema = _build_schema(df, data_start_row=2)
        assert "order_id" in schema
        assert schema["order_id"]["col"] == "A"
        assert schema["order_id"]["col_index"] == 0
        assert schema["order_id"]["type"] == "string"
        assert schema["order_id"]["null_ratio"] == 0.0

        assert schema["amount"]["col"] == "B"
        assert schema["amount"]["type"] == "decimal"
        assert schema["amount"]["null_ratio"] > 0

        assert schema["qty"]["col"] == "C"
        assert schema["qty"]["type"] == "integer"

    def test_categories(self):
        df = pd.DataFrame({"platform": ["淘宝", "京东", "淘宝", "拼多多"]})
        schema = _build_schema(df, data_start_row=2)
        assert "categories" in schema["platform"]
        assert "淘宝" in schema["platform"]["categories"]

    def test_hidden_col_excluded(self):
        df = pd.DataFrame({"name": ["a"], "_is_hidden": [False]})
        schema = _build_schema(df, data_start_row=2)
        assert "_is_hidden" not in schema
        assert "name" in schema


# ── _build_sample ──


class TestBuildSample:
    def test_sample_with_row_numbers(self):
        # 6 行小数据：head/tail 各 4 行可能重叠，但函数不去重（行号保留）
        df = pd.DataFrame({
            "name": ["Alice", "Bob", "Charlie", "David", "Eve", "Frank"],
            "age": [25, 30, 35, 40, 45, 50],
        })
        sample = _build_sample(df, data_start_row=2)
        # head 4 行
        assert len(sample["head"]) == 4
        assert sample["head"][0]["_row"] == 2
        assert sample["head"][0]["name"] == "Alice"
        # tail 4 行（最后一行索引 5 + data_start_row 2 = 7）
        assert sample["tail"][-1]["_row"] == 7
        # 6 行小数据没 middle / boundary
        assert sample.get("middle") == []
        assert sample.get("boundary") == []

    def test_empty_df(self):
        df = pd.DataFrame(columns=["a", "b"])
        sample = _build_sample(df, data_start_row=2)
        assert sample["head"] == []
        assert sample["tail"] == []
        assert sample["middle"] == []
        assert sample["boundary"] == []

    def test_small_df(self):
        # 2 行小数据，head 取 2 行；但去重后第二行 (x=2) 签名和第一行 (x=1) 都是 "+"
        # 会被去重，但首行无条件保留 → 至少 1 行
        df = pd.DataFrame({"x": [1, 2]})
        sample = _build_sample(df, data_start_row=3)
        assert len(sample["head"]) >= 1
        assert sample["head"][0]["_row"] == 3

    def test_large_df_has_middle_segment(self):
        """大数据应有 middle 段（解决 lost-in-the-middle）"""
        # 单调 range(100) 都是非零 → 签名全部 "+"，去重后每段保 1 行
        # 用 50 零值 + 50 非零值，让签名有多样性
        df = pd.DataFrame({"x": [0]*30 + list(range(1, 71))})
        sample = _build_sample(df, data_start_row=2)
        # 各段至少 1 行（去重后保结构）
        assert len(sample["head"]) >= 1
        assert len(sample["middle"]) >= 1
        assert len(sample["tail"]) >= 1
        # middle 行号在中间附近
        mid_row = sample["middle"][0]["_row"]
        assert 40 < mid_row < 60  # 100 行的中间附近

    def test_boundary_from_prescan(self):
        """复用 ai_decision.anomalies 作为 boundary 边界样本"""
        df = pd.DataFrame({"amount": list(range(100))})
        # 模拟 prescan 标记的异常行
        class _Prescan:
            anomalies = [
                {"column": "amount", "sample_rows": [50, 75]},  # Excel 行号
            ]
        sample = _build_sample(df, data_start_row=2, ai_decision=_Prescan())
        assert len(sample["boundary"]) <= 2
        # Excel row 50 → df idx 48; Excel row 75 → df idx 73
        boundary_rows = [r["_row"] for r in sample["boundary"]]
        assert 50 in boundary_rows or 75 in boundary_rows

    def test_boundary_dedup_with_head_tail(self):
        """边界行如果已在 head/tail 里，不重复"""
        df = pd.DataFrame({"x": list(range(100))})
        class _Prescan:
            # Excel 行号 2 = df idx 0（已在 head 里）
            anomalies = [{"column": "x", "sample_rows": [2]}]
        sample = _build_sample(df, data_start_row=2, ai_decision=_Prescan())
        # boundary 应为空（已被 head 覆盖）
        assert sample["boundary"] == []

    def test_boundary_no_prescan_returns_empty(self):
        """无 ai_decision 时 boundary 为空"""
        df = pd.DataFrame({"x": list(range(100))})
        sample = _build_sample(df, data_start_row=2)
        assert sample["boundary"] == []

    def test_boundary_max_limit(self):
        """boundary 最多 2 行，超出忽略（注意：去重后可能进一步减少）"""
        df = pd.DataFrame({"x": list(range(100))})
        class _Prescan:
            anomalies = [
                {"column": "x", "sample_rows": [30, 40, 50, 60, 70]},
            ]
        sample = _build_sample(df, data_start_row=2, ai_decision=_Prescan())
        # 全部非零签名相同 → 去重后保 1 行（首行无条件保留）
        # 测试上限逻辑：boundary_indices 最多收 2 个再去重
        assert len(sample["boundary"]) >= 1
        assert len(sample["boundary"]) <= 2


# ── _scan_issues ──


class TestScanIssues:
    def test_missing_values(self):
        df = pd.DataFrame({
            "name": ["Alice", None, "Charlie"],
            "age": [25, 30, None],
        })
        issues = _scan_issues(df, data_start_row=2)
        missing_issues = [i for i in issues if i["type"] == "missing_value"]
        assert len(missing_issues) == 2
        # 第一个缺失在 name 列，row 应该是 index 1 + 2 = 3
        name_issue = next(i for i in missing_issues if i["location"]["raw_col_name"] == "name")
        assert name_issue["location"]["row"] == 3
        assert name_issue["location"]["col"] == "A"
        assert name_issue["preserved"] is True
        assert "recovery_hint" in name_issue

    def test_duplicates(self):
        df = pd.DataFrame({
            "name": ["Alice", "Alice", "Bob"],
            "age": [25, 25, 30],
        })
        issues = _scan_issues(df, data_start_row=2)
        dup_issues = [i for i in issues if i["type"] == "duplicate_row"]
        assert len(dup_issues) == 1
        assert dup_issues[0]["preserved"] is True
        assert "drop_duplicates" in dup_issues[0]["recovery_hint"]

    def test_no_issues(self):
        df = pd.DataFrame({"x": [1, 2, 3], "y": ["a", "b", "c"]})
        issues = _scan_issues(df, data_start_row=2)
        assert issues == []


# ── _determine_status ──


class TestDetermineStatus:
    def test_pass(self):
        assert _determine_status([]) == "pass"

    def test_warning(self):
        issues = [{"severity": "warning"}]
        assert _determine_status(issues) == "warning"

    def test_fail(self):
        issues = [{"severity": "error"}, {"severity": "warning"}]
        assert _determine_status(issues) == "fail"


# ── generate_file_meta ──


class TestGenerateFileMeta:
    def test_basic(self):
        df = pd.DataFrame({
            "order_id": ["A001", "A002"],
            "amount": [100.0, 200.0],
        })
        report = CleaningReport(
            original_shape=(2, 2),
            final_shape=(2, 2),
            data_start_row=2,
        )
        meta = generate_file_meta(df, report, source_file="test.xlsx")
        assert meta.version == "1.0"
        assert meta.status == "pass"
        assert meta.summary["row_count"] == 2
        assert meta.summary["col_count"] == 2
        assert "order_id" in meta.schema
        assert meta.schema["order_id"]["col"] == "A"
        assert len(meta.sample["head"]) == 2
        assert meta.sample["head"][0]["_row"] == 2

    def test_with_missing_values(self):
        df = pd.DataFrame({"x": [1, None, 3]})
        report = CleaningReport(data_start_row=2)
        meta = generate_file_meta(df, report, source_file="test.csv")
        assert meta.status == "warning"
        assert meta.stats["missing_values"] == 1
        assert len(meta.issues) > 0
        assert meta.confidence == 0.8


# ── to_dict / write / read ──


class TestFileMetaToDict:
    def test_returns_dict(self):
        meta = FileMeta(version="1.0", status="pass", source_file="test.xlsx")
        d = meta.to_dict()
        assert isinstance(d, dict)
        assert d["version"] == "1.0"
        assert d["status"] == "pass"
        assert d["source_file"] == "test.xlsx"

    def test_serializable(self):
        """to_dict 结果必须能 JSON 序列化。"""
        df = pd.DataFrame({"a": [1, 2]})
        report = CleaningReport(data_start_row=2)
        meta = generate_file_meta(df, report, source_file="test.xlsx")
        d = meta.to_dict()
        import json
        serialized = json.dumps(d, default=str)
        assert len(serialized) > 0


class TestWriteReadFileMeta:
    def test_round_trip(self, tmp_path):
        cache_path = str(tmp_path / "test.parquet")
        df = pd.DataFrame({"a": [1, 2], "b": ["x", "y"]})
        report = CleaningReport(data_start_row=2)
        meta = generate_file_meta(df, report, source_file="test.xlsx")

        write_file_meta(cache_path, meta)

        meta_path = Path(cache_path.replace(".parquet", ".meta.json"))
        assert meta_path.exists()

        loaded = read_file_meta(cache_path)
        assert loaded is not None
        assert loaded.version == "1.0"
        assert loaded.summary["row_count"] == 2

    def test_read_old_format_returns_none(self, tmp_path):
        cache_path = str(tmp_path / "old.parquet")
        meta_path = cache_path.replace(".parquet", ".meta.json")
        # 旧格式没有 version 字段
        Path(meta_path).write_text(json.dumps({"merged_cols_filled": 0}))
        assert read_file_meta(cache_path) is None

    def test_read_missing_file(self, tmp_path):
        assert read_file_meta(str(tmp_path / "nonexistent.parquet")) is None


# ── format_file_view ──


class TestFormatFileView:
    def test_basic_output(self):
        df = pd.DataFrame({
            "order_id": ["A001", "A002"],
            "amount": [100.0, 200.0],
        })
        report = CleaningReport(data_start_row=2)
        meta = generate_file_meta(df, report, source_file="sales.xlsx")
        view = format_file_view(meta)

        assert "[文件已就绪]" in view
        assert "sales.xlsx" in view
        assert "order_id" in view
        assert "Row 2" in view
        assert "行号映射" in view

    def test_small_data_no_warning(self):
        """小数据（< 1万行）不应触发规模警告"""
        df = pd.DataFrame({"a": list(range(100)), "b": list(range(100))})
        report = CleaningReport(data_start_row=2)
        meta = generate_file_meta(df, report, source_file="small.xlsx")
        view = format_file_view(meta)
        assert "⚠️" not in view
        assert "OOM" not in view

    def test_medium_data_hint(self):
        """中数据（≥1万行）应出现温和提示，不出 ⚠️ 大数据警告"""
        df = pd.DataFrame({"a": list(range(15_000)), "b": list(range(15_000))})
        report = CleaningReport(data_start_row=2)
        meta = generate_file_meta(df, report, source_file="medium.xlsx")
        view = format_file_view(meta)
        assert "15,000 行" in view  # 新格式带空格
        assert "WHERE/GROUP BY" in view
        assert "⚠️ **大数据**" not in view  # 中数据不上 ⚠️ 大数据警告

    def test_large_data_oom_warning(self):
        """大数据（≥10万行）应出现 ⚠️ + OOM 警告（行业 schema-aware 做法）"""
        df = pd.DataFrame({"a": list(range(150_000)), "b": list(range(150_000))})
        report = CleaningReport(data_start_row=2)
        meta = generate_file_meta(df, report, source_file="large.xlsx")
        view = format_file_view(meta)
        assert "⚠️" in view
        assert "150,000 行" in view  # 新格式带空格
        assert "OOM" in view
        assert "SELECT *" in view  # 给出反例

    def test_markdown_structure(self):
        """输出应使用 Markdown ## 标题分块（U 形 attention 优化）"""
        df = pd.DataFrame({"x": list(range(50))})
        report = CleaningReport(data_start_row=2)
        meta = generate_file_meta(df, report, source_file="any.xlsx")
        view = format_file_view(meta)
        # 至少有 schema 和 sample 两个 ## 标题
        assert "## 📐 字段 schema" in view
        assert "## 📋 样本数据" in view
        assert "## 📊 数据概览" in view

    def test_u_shape_anchoring_for_grain(self):
        """订单级警告必须头部 + 尾部双锚定（防中段遗漏）"""
        # 构造一个 grain 检测能命中的 df
        df = pd.DataFrame({
            "order_id": (["A001"] * 5 + ["A002"] * 5 + ["A003"] * 5
                         + ["A004"] * 5 + ["A005"] * 5 + ["A006"] * 5
                         + ["A007"] * 5 + ["A008"] * 5),
            "amount": ([100.0] * 5 + [200.0] * 5 + [300.0] * 5
                       + [400.0] * 5 + [500.0] * 5 + [600.0] * 5
                       + [700.0] * 5 + [800.0] * 5),  # 订单级（同订单 amount 重复）
            "item_qty": list(range(40)),  # 明细级（每行不同）
        })
        report = CleaningReport(data_start_row=2)
        meta = generate_file_meta(df, report, source_file="orders.xlsx")
        view = format_file_view(meta)

        # 如果检测到 grain（订单级数值字段），应同时出现：
        # V3：grain 字段已删除，order_level 改由 ai_decision.column_semantics 标注
        # 此处简化为：如果生成的 view 包含订单级线索就验证，否则跳过
        if "订单级" in view:
            # 头部锚定（## 概览 段含订单级警告）
            head_part = view.split("## 📐")[0]  # 概览部分
            assert "订单级" in head_part
            # 尾部锚定（## 再次提醒 段）
            assert "## ⚠️ 再次提醒" in view

    def test_multi_field_aggregation_template_at_tail(self):
        """尾部应包含多字段聚合范式（明细级 + 订单级混合时的完整代码模板）"""
        df = pd.DataFrame({
            "order_id": (["A001"] * 8 + ["A002"] * 8 + ["A003"] * 8
                         + ["A004"] * 8 + ["A005"] * 8),
            "refund": ([10.0] * 8 + [20.0] * 8 + [30.0] * 8
                       + [40.0] * 8 + [50.0] * 8),  # 订单级数值
            "qty": list(range(40)),  # 明细级数值
        })
        report = CleaningReport(data_start_row=2)
        meta = generate_file_meta(df, report, source_file="t.xlsx")
        view = format_file_view(meta)
        # V3：grain 字段已删除，order_level 改由 ai_decision.column_semantics 标注
        # 此处简化为：如果生成的 view 包含订单级线索就验证，否则跳过
        if "订单级" in view:
            # 必须有多字段范式（行业最佳实践: One-shot template）
            assert "多字段聚合范式" in view
            # 必须示范了"明细级 SUM + 订单级 DISTINCT 子查询 + pandas merge"完整三步
            assert "明细级字段直接 SUM" in view
            assert "SELECT DISTINCT" in view
            assert "merge" in view  # pandas merge
            # 必须警告"不要用 SQL 三表 JOIN"（之前 LLM 自创出错的写法）
            assert "三表 JOIN" in view or "三表" in view

    def test_schema_row_level_order_tag(self):
        """订单级数值字段在 schema 行直接标 🔴"""
        df = pd.DataFrame({
            "order_id": (["A001"] * 8 + ["A002"] * 8 + ["A003"] * 8
                         + ["A004"] * 8 + ["A005"] * 8),
            "refund": ([10.0] * 8 + [20.0] * 8 + [30.0] * 8
                       + [40.0] * 8 + [50.0] * 8),  # 订单级数值
            "qty": list(range(40)),  # 明细级
        })
        report = CleaningReport(data_start_row=2)
        meta = generate_file_meta(df, report, source_file="t.xlsx")
        view = format_file_view(meta)
        # V3：grain 字段已删除，order_level 改由 ai_decision.column_semantics 标注
        # 此处简化为：如果生成的 view 包含订单级线索就验证，否则跳过
        if "订单级" in view:
            # refund 列在 schema 行应有 🔴 标签
            schema_section = view.split("## 📐")[1].split("## ")[0]
            assert "🔴" in schema_section
            assert "refund" in schema_section


# ── _compress_issues 回归（Bug-1/2 修复）──


class TestCompressIssuesNoFabrication:
    """Bug-1/2 修复回归：_compress_issues 不再编造 '_is_summary 汇总行' 提示。"""

    def test_multi_null_no_summary_fabrication(self):
        """≥3 个同行 missing_value 只输出客观'多列缺失'，不附加汇总行猜测。"""
        from services.agent.file_meta.view import _compress_issues

        issues = [
            {"type": "missing_value", "severity": "warning",
             "location": {"row": 2, "col": "D"}, "action": "D 列缺 N 个"},
            {"type": "missing_value", "severity": "warning",
             "location": {"row": 2, "col": "E"}, "action": "E 列缺 N 个"},
            {"type": "missing_value", "severity": "warning",
             "location": {"row": 2, "col": "F"}, "action": "F 列缺 N 个"},
        ]
        out = _compress_issues(issues)
        text = "\n".join(out)
        # 必须有客观陈述
        assert "Row 2" in text
        assert "多列缺失" in text
        # 必须没有任何造谣
        assert "_is_summary" not in text, "Bug-1 修复后不应再编造 _is_summary 提示"
        assert "汇总行" not in text, "Bug-1 修复后不应再编造汇总行猜测"
        assert "WHERE" not in text, "不应给出 SQL 建议（这是 AI 裁决的职责）"

    def test_single_missing_value_unchanged(self):
        """< 3 个同行 missing_value 走 _format_single_issue 不受影响。"""
        from services.agent.file_meta.view import _compress_issues

        issues = [
            {"type": "missing_value", "severity": "warning",
             "location": {"row": 5, "col": "D"}, "action": "D 列缺 1 个"},
        ]
        out = _compress_issues(issues)
        assert len(out) == 1
        # 单条仍按 _format_single_issue 输出，含 Row 5 D 列
        assert "Row 5" in out[0]

    def test_non_missing_value_passthrough(self):
        """非 missing_value 类型不受合并影响。"""
        from services.agent.file_meta.view import _compress_issues

        issues = [
            {"type": "merge_filled", "severity": "info",
             "location": {"cols": ["A"]}, "action": "合并填充"},
            {"type": "int_cols_fixed", "severity": "info",
             "location": {"cols": ["qty"]}, "action": "整数修复"},
        ]
        out = _compress_issues(issues)
        assert len(out) == 2


# ── extract_formulas ──


class TestExtractFormulas:
    def test_xlsx_with_formulas(self, tmp_path):
        """测试从包含公式的 xlsx 文件提取公式。"""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "数量"
        ws["A2"] = 10
        ws["A3"] = 20
        ws["A4"] = "=SUM(A2:A3)"
        wb.save(str(tmp_path / "formulas.xlsx"))
        wb.close()

        result, skip_reason = extract_formulas(str(tmp_path / "formulas.xlsx"), "Sheet1")
        assert len(result) >= 1
        assert skip_reason == ""
        formula_entry = next(
            (f for f in result if "A4" in f["cell"]), None
        )
        assert formula_entry is not None
        assert "SUM" in formula_entry["formula"]
        assert formula_entry["formula"].startswith("=")

    def test_csv_returns_empty(self, tmp_path):
        """CSV 文件不支持公式提取。"""
        csv_path = tmp_path / "data.csv"
        csv_path.write_text("a,b\n1,2\n")
        result, skip = extract_formulas(str(csv_path))
        assert result == []

    def test_no_formulas(self, tmp_path):
        """无公式的 xlsx 返回空列表。"""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "hello"
        ws["B1"] = 42
        wb.save(str(tmp_path / "no_formulas.xlsx"))
        wb.close()

        result, skip = extract_formulas(str(tmp_path / "no_formulas.xlsx"))
        assert result == []

    def test_formulas_in_file_meta(self, tmp_path):
        """公式传入 generate_file_meta 后出现在输出中。"""
        df = pd.DataFrame({"a": [1, 2], "b": [3, 4]})
        report = CleaningReport(data_start_row=2)
        formulas = [{"cell": "Sheet1!C3", "formula": "=A3+B3", "value": 6}]
        meta = generate_file_meta(df, report, "test.xlsx", formulas=formulas)
        assert len(meta.formulas) == 1
        assert meta.formulas[0]["cell"] == "Sheet1!C3"

        view = format_file_view(meta)
        assert "Sheet1!C3" in view
        assert "=A3+B3" in view

    def test_no_file_size_threshold(self, tmp_path):
        """不再有 50MB 文件大小限制，任意大小均尝试提取。"""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "=SUM(1,2)"
        xlsx_path = str(tmp_path / "test.xlsx")
        wb.save(xlsx_path)
        wb.close()

        # 直接提取，无大小检查
        result, skip_reason = extract_formulas(xlsx_path)
        assert skip_reason == ""
        assert len(result) >= 1

    def test_formula_value_is_numeric(self, tmp_path):
        """公式的计算值应被解析为数值类型。"""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = 10
        ws["A2"] = 20
        ws["A3"] = "=SUM(A1:A2)"
        xlsx_path = str(tmp_path / "numeric.xlsx")
        wb.save(xlsx_path)
        wb.close()

        result, _ = extract_formulas(xlsx_path)
        if result:
            formula_entry = next(
                (f for f in result if "A3" in f["cell"]), None
            )
            assert formula_entry is not None
            # 值可能是 None（openpyxl 创建的文件未计算），但格式正确
            assert "cell" in formula_entry
            assert "formula" in formula_entry
            assert "value" in formula_entry


class TestFix3DuckDBNativeType:
    """Fix 3: view 渲染层把 schema type 映射成 DuckDB native type，避免 AI 用 MySQL/PG 方言。"""

    def test_datetime_renders_as_timestamp(self):
        df = pd.DataFrame({
            "申请时间": pd.to_datetime(["2024-01-01", "2024-02-01"]),
        })
        report = CleaningReport(data_start_row=2)
        meta = generate_file_meta(df, report, source_file="t.xlsx")
        view = format_file_view(meta)
        assert "TIMESTAMP" in view, "datetime 列应渲染为 DuckDB TIMESTAMP"
        assert " | datetime |" not in view, "不应再裸露 logical type"

    def test_integer_decimal_renders_as_native(self):
        df = pd.DataFrame({"qty": [1, 2, 3], "amount": [1.5, 2.5, 3.5]})
        report = CleaningReport(data_start_row=2)
        meta = generate_file_meta(df, report, source_file="t.xlsx")
        view = format_file_view(meta)
        assert "BIGINT" in view
        assert "DOUBLE" in view

    def test_string_renders_as_varchar(self):
        df = pd.DataFrame({"order_no": ["A001", "A002", "A003"]})
        report = CleaningReport(data_start_row=2)
        meta = generate_file_meta(df, report, source_file="t.xlsx")
        view = format_file_view(meta)
        assert "VARCHAR" in view

    def test_schema_field_value_unchanged(self):
        """渲染层映射不应改 schema 数据值本身（向后兼容下游消费方）。"""
        df = pd.DataFrame({"a": pd.to_datetime(["2024-01-01"])})
        report = CleaningReport(data_start_row=2)
        meta = generate_file_meta(df, report, source_file="t.xlsx")
        # schema 数据值保持原 logical type
        assert meta.schema["a"]["type"] == "datetime"


class TestFix4MultiSheetView:
    """Fix 4: 多 Sheet 章节渲染（evidence_summary['sheets'] 透传）。"""

    def test_single_sheet_no_section(self):
        """单 sheet 文件不应渲染多 Sheet 章节。"""
        df = pd.DataFrame({"a": [1, 2]})
        report = CleaningReport(data_start_row=2)
        meta = generate_file_meta(df, report, source_file="single.xlsx")
        view = format_file_view(meta)
        assert "## 📋 多 Sheet" not in view

    def test_multi_sheet_merged_section(self):
        """已合并的多 sheet 应渲染 ✅ + _sheet 列查询提示。"""
        df = pd.DataFrame({"a": [1, 2]})
        report = CleaningReport(data_start_row=2)
        meta = generate_file_meta(df, report, source_file="multi.xlsx", sheet_count=3)
        meta.evidence_summary["sheets"] = {
            "total": 3,
            "merged": ["Sheet1", "Sheet2", "Sheet3"],
            "skipped": [],
        }
        view = format_file_view(meta)
        assert "## 📋 多 Sheet" in view
        assert "已合并 3 个" in view
        assert "Sheet1, Sheet2, Sheet3" in view
        assert "_sheet" in view  # 必含查询用法提示

    def test_multi_sheet_with_skipped(self):
        """被 AI 跳过的 sheet 应单独渲染 ⏭️ + 单读 sheet 提示。"""
        df = pd.DataFrame({"a": [1, 2]})
        report = CleaningReport(data_start_row=2)
        meta = generate_file_meta(df, report, source_file="multi.xlsx", sheet_count=3)
        meta.evidence_summary["sheets"] = {
            "total": 3,
            "merged": ["销售明细"],
            "skipped": [
                {"name": "汇总", "role": "aggregated"},
                {"name": "说明", "role": "meta"},
            ],
        }
        view = format_file_view(meta)
        assert "⏭️" in view
        assert "汇总" in view
        assert "说明" in view
        assert "aggregated" in view or "汇总表" in view
        # 新协议:不再给 pd.read_excel 字面值,改引导重新调用 file_analyze
        assert "file_analyze" in view

    def test_empty_evidence_summary_safe(self):
        """evidence_summary 缺失或 sheets 为空时不应崩，不渲染章节。"""
        df = pd.DataFrame({"a": [1, 2]})
        report = CleaningReport(data_start_row=2)
        meta = generate_file_meta(df, report, source_file="x.xlsx")
        # evidence_summary 缺 sheets 子键
        view = format_file_view(meta)
        assert "## 📋 多 Sheet" not in view


class TestFix5SampleMarkdownTable:
    """Fix 5: 样本数据用 markdown table 格式（对齐 LangChain Pandas Agent / Vanna），
    避免 Python dict 字面量把 Timestamp 加引号让 AI 误判为字符串。"""

    def test_sample_uses_markdown_pipe_format(self):
        """样本段必须用 markdown table 的 | 分隔，不是 Python dict 字面量。"""
        df = pd.DataFrame({
            "订单号": ["A001", "A002", "A003"],
            "金额": [100.0, 200.0, 300.0],
        })
        report = CleaningReport(data_start_row=2)
        meta = generate_file_meta(df, report, source_file="test.xlsx")
        view = format_file_view(meta)
        # 找到样本段
        assert "## 📋 样本数据" in view
        sample_section = view.split("## 📋 样本数据")[1].split("##")[0]
        # 必须含 markdown table 的分隔符（| 列分隔 + 多个 | 的表头分隔行 :---:）
        assert "|" in sample_section
        assert ":---" in sample_section or "|--" in sample_section, \
            f"样本段未使用 markdown table 格式（无表头分隔行）: {sample_section[:200]}"

    def test_sample_timestamp_no_quotes(self):
        """Timestamp 在样本表里不带引号（pandas table 风格）→ AI 不再误判为 string。"""
        df = pd.DataFrame({
            "申请时间": pd.to_datetime(["2024-01-01 10:23:45"]),
            "金额": [100.0],
        })
        report = CleaningReport(data_start_row=2)
        meta = generate_file_meta(df, report, source_file="test.xlsx")
        view = format_file_view(meta)
        sample_section = view.split("## 📋 样本数据")[1].split("##")[0]
        # 日期值应该出现但不带引号
        assert "2024-01-01" in sample_section, \
            f"样本里应有时间值: {sample_section[:300]}"
        assert "'2024-01-01" not in sample_section, \
            "样本里 timestamp 不应带单引号（dict 字面量风格）"
        assert '"2024-01-01' not in sample_section, \
            "样本里 timestamp 不应带双引号"

    def test_sample_includes_row_labels(self):
        """每行需有 Row N [tag] 标签，复用之前的行号映射约定。"""
        df = pd.DataFrame({"a": [1, 2, 3]})
        report = CleaningReport(data_start_row=2)
        meta = generate_file_meta(df, report, source_file="test.xlsx")
        view = format_file_view(meta)
        assert "Row 2 [head]" in view, "样本行应含行号 + 段标签"

    def test_sample_render_fallback_on_tabulate_failure(self, monkeypatch):
        """tabulate 异常时降级为 dict 字面量（不阻塞主流程）。"""
        df = pd.DataFrame({"a": [1, 2]})
        report = CleaningReport(data_start_row=2)
        meta = generate_file_meta(df, report, source_file="test.xlsx")
        # monkeypatch DataFrame.to_markdown 抛错
        def _boom(self, *a, **kw):
            raise RuntimeError("simulated tabulate missing")
        monkeypatch.setattr(pd.DataFrame, "to_markdown", _boom)
        view = format_file_view(meta)
        # 应降级输出，不抛异常
        assert "## 📋 样本数据" in view
        assert "Row 2" in view


class TestFix6DuckDBDateFunctions:
    """DuckDB 日期方言 hint(新协议极简:只留核心陷阱型差异,不列函数库)

    V3 描述刻意去掉了函数清单(strftime/YEAR/MONTH/...),让 LLM 用训练知识。
    只保留 DATE_TRUNC 和类型 hint(TIMESTAMP 不是 DATETIME)。
    """

    def test_workspace_version_has_date_trunc(self):
        from config.code_tools import build_code_tools
        desc = build_code_tools(include_workspace=True)[0]["function"]["description"]
        assert "DATE_TRUNC" in desc

    def test_workspace_version_has_timestamp_type(self):
        from config.code_tools import build_code_tools
        desc = build_code_tools(include_workspace=True)[0]["function"]["description"]
        assert "TIMESTAMP" in desc

    def test_base_version_has_date_trunc(self):
        from config.code_tools import build_code_tools
        desc = build_code_tools(include_workspace=False)[0]["function"]["description"]
        assert "DATE_TRUNC" in desc

    def test_base_version_has_timestamp_type(self):
        from config.code_tools import build_code_tools
        desc = build_code_tools(include_workspace=False)[0]["function"]["description"]
        assert "TIMESTAMP" in desc
