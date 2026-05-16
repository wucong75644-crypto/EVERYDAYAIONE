"""测试 excel_cleaner — 三层清洗防线"""
from __future__ import annotations

import json
import os
import sys
import tempfile
from pathlib import Path

backend_dir = Path(__file__).parent.parent
if str(backend_dir) not in sys.path:
    sys.path.insert(0, str(backend_dir))

import openpyxl
import pandas as pd
import pytest

from services.agent.data_query_cache import detect_header_depth, detect_header_row
from services.agent.excel_cleaner import (
    CleaningReport,
    ExcelStructure,
    _col_letter_to_index,
    _detect_structure,
    _flatten_multi_header,
    _resolve_sheet_xml_path,
    clean_excel,
    read_cleaning_report,
    write_cleaning_report,
)


# ============================================================
# Fixtures
# ============================================================


@pytest.fixture
def simple_xlsx(tmp_path):
    """普通 Excel（无合并/隐藏/汇总）"""
    p = tmp_path / "simple.xlsx"
    pd.DataFrame({"name": ["Alice", "Bob", "Carol"], "age": [20, 30, 40]}).to_excel(
        str(p), index=False
    )
    return str(p)


@pytest.fixture
def merged_xlsx(tmp_path):
    """带垂直合并单元格的 Excel"""
    p = tmp_path / "merged.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"], ws["B1"], ws["C1"] = "订单号", "商品", "金额"
    ws["A2"], ws["B2"], ws["C2"] = "ORD001", "苹果", 100
    ws["A3"], ws["B3"], ws["C3"] = None, "香蕉", 200
    ws["A4"], ws["B4"], ws["C4"] = None, "橙子", 300
    ws["A5"], ws["B5"], ws["C5"] = "ORD002", "葡萄", 150
    ws["A6"], ws["B6"], ws["C6"] = None, "西瓜", 250
    ws.merge_cells("A2:A4")  # 垂直合并 ORD001
    ws.merge_cells("A5:A6")  # 垂直合并 ORD002
    wb.save(str(p))
    wb.close()
    return str(p)


@pytest.fixture
def hidden_xlsx(tmp_path):
    """带隐藏行和隐藏列的 Excel"""
    p = tmp_path / "hidden.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"], ws["B1"], ws["C1"] = "name", "hidden_col", "value"
    for i in range(2, 12):
        ws[f"A{i}"] = f"item{i - 1}"
        ws[f"B{i}"] = f"secret{i}"
        ws[f"C{i}"] = i * 10
    ws.row_dimensions[5].hidden = True
    ws.row_dimensions[6].hidden = True
    ws.column_dimensions["B"].hidden = True
    wb.save(str(p))
    wb.close()
    return str(p)


@pytest.fixture
def autofilter_xlsx(tmp_path):
    """带自动筛选的 Excel"""
    p = tmp_path / "filter.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"], ws["B1"] = "category", "value"
    for i in range(2, 22):
        ws[f"A{i}"] = f"cat{i % 3}"
        ws[f"B{i}"] = i * 10
    ws.auto_filter.ref = "A1:B21"
    wb.save(str(p))
    wb.close()
    return str(p)


@pytest.fixture
def multi_sheet_xlsx(tmp_path):
    """多 Sheet Excel"""
    p = tmp_path / "multi.xlsx"
    with pd.ExcelWriter(str(p), engine="openpyxl") as w:
        pd.DataFrame({"a": [1, 2, 3]}).to_excel(w, sheet_name="数据", index=False)
        pd.DataFrame({"b": [4, 5, 6]}).to_excel(w, sheet_name="汇总", index=False)
    return str(p)


# ============================================================
# Layer 1: 结构检测
# ============================================================


class TestColLetterToIndex:
    def test_single_letter(self):
        assert _col_letter_to_index("A") == 1
        assert _col_letter_to_index("Z") == 26

    def test_double_letter(self):
        assert _col_letter_to_index("AA") == 27
        assert _col_letter_to_index("AZ") == 52

    def test_lowercase(self):
        assert _col_letter_to_index("a") == 1


class TestDetectStructure:
    def test_simple_no_structure(self, simple_xlsx):
        s = _detect_structure(simple_xlsx, 0)
        assert s is not None
        assert len(s.merged_ranges) == 0
        assert len(s.hidden_rows) == 0
        assert len(s.hidden_cols) == 0
        assert s.has_auto_filter is False

    def test_merged_cells(self, merged_xlsx):
        s = _detect_structure(merged_xlsx, 0)
        assert s is not None
        assert len(s.merged_ranges) == 2
        # A2:A4 → (2, 4, 1, 1)
        assert (2, 4, 1, 1) in s.merged_ranges
        # A5:A6 → (5, 6, 1, 1)
        assert (5, 6, 1, 1) in s.merged_ranges

    def test_hidden_rows_and_cols(self, hidden_xlsx):
        s = _detect_structure(hidden_xlsx, 0)
        assert s is not None
        assert 5 in s.hidden_rows
        assert 6 in s.hidden_rows
        assert 2 in s.hidden_cols  # B = col 2

    def test_auto_filter(self, autofilter_xlsx):
        s = _detect_structure(autofilter_xlsx, 0)
        assert s is not None
        assert s.has_auto_filter is True

    def test_xls_fallback(self, tmp_path):
        """旧格式 .xls 应返回 None（降级）"""
        p = tmp_path / "old.xls"
        p.write_bytes(b"fake")
        assert _detect_structure(str(p), 0) is None

    def test_nonexistent_file(self):
        assert _detect_structure("/tmp/nonexistent.xlsx", 0) is None

    def test_sheet_by_name(self, multi_sheet_xlsx):
        s = _detect_structure(multi_sheet_xlsx, "数据")
        assert s is not None

    def test_sheet_by_index(self, multi_sheet_xlsx):
        s = _detect_structure(multi_sheet_xlsx, 1)
        assert s is not None


# ============================================================
# Layer 2: 智能清洗
# ============================================================


class TestMergeFill:
    def test_merge_not_auto_filled(self, merged_xlsx):
        """合并单元格不再自动 ffill——由 AI 在 code_execute 中按需处理。"""
        df = pd.read_excel(merged_xlsx, engine="calamine")
        assert pd.isna(df.iloc[1, 0])  # 合并后为 NaN

        df_clean, report = clean_excel(df, merged_xlsx, 0)
        assert df_clean.iloc[0, 0] == "ORD001"
        assert pd.isna(df_clean.iloc[1, 0])  # 不再自动填充
        assert pd.isna(df_clean.iloc[2, 0])

    def test_other_cols_untouched(self, merged_xlsx):
        df = pd.read_excel(merged_xlsx, engine="calamine")
        df_clean, _ = clean_excel(df, merged_xlsx, 0)
        assert df_clean.iloc[0, 1] == "苹果"
        assert df_clean.iloc[1, 1] == "香蕉"


class TestHiddenRowMark:
    def test_no_auto_hidden_mark(self, hidden_xlsx):
        """隐藏行/列不再自动标记——结构检测已移除，由 AI 按需处理。"""
        df = pd.read_excel(hidden_xlsx, engine="calamine")
        df_clean, _ = clean_excel(df, hidden_xlsx, 0)
        assert "_is_hidden" not in df_clean.columns

    def test_no_hidden_no_column(self, simple_xlsx):
        df = pd.read_excel(simple_xlsx, engine="calamine")
        df_clean, _ = clean_excel(df, simple_xlsx, 0)
        assert "_is_hidden" not in df_clean.columns


class TestHiddenColReport:
    def test_data_preserved(self, hidden_xlsx):
        """隐藏列的数据仍保留（不删除）。"""
        df = pd.read_excel(hidden_xlsx, engine="calamine")
        df_clean, _ = clean_excel(df, hidden_xlsx, 0)
        assert "hidden_col" in df_clean.columns


class TestAutoFilter:
    def test_basic_clean(self, autofilter_xlsx):
        """自动筛选不再检测——结构检测已移除。"""
        df = pd.read_excel(autofilter_xlsx, engine="calamine")
        df_clean, _ = clean_excel(df, autofilter_xlsx, 0)
        assert len(df_clean) > 0


# ============================================================
# Layer 3: 质量校验
# ============================================================


class TestEmptyRowsCols:
    def test_empty_cols_preserved_and_annotated(self):
        """空列不再删除，只在 warnings 里标注。"""
        df = pd.DataFrame({"a": [1, 2], "Unnamed: 0": [None, None], "b": [3, 4]})
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            p = f.name
        df.to_excel(p, index=False)
        df_clean, report = clean_excel(df.copy(), p, 0)
        assert "Unnamed: 0" in df_clean.columns  # 不再删除
        assert report.empty_cols_removed == 0
        assert any("全空列" in w for w in report.warnings)
        os.unlink(p)

    def test_empty_rows_preserved_and_annotated(self):
        """空行不再删除，只在 warnings 里标注。"""
        df = pd.DataFrame({"a": [1, None, 3], "b": [4, None, 6]})
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            p = f.name
        df.to_excel(p, index=False)
        df_clean, report = clean_excel(df.copy(), p, 0)
        assert len(df_clean) == 3  # 不再删除
        assert report.empty_rows_removed == 0
        assert any("全空行" in w for w in report.warnings)
        os.unlink(p)


class TestIntFix:
    def test_float_to_int64(self):
        df = pd.DataFrame({"order_id": [1.0, 2.0, 3.0, None], "name": ["a", "b", "c", "d"]})
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            p = f.name
        df.to_excel(p, index=False)
        df_clean, report = clean_excel(df.copy(), p, 0)
        assert df_clean["order_id"].dtype.name == "Int64"
        assert report.int_cols_fixed == 1
        os.unlink(p)

    def test_real_float_untouched(self):
        df = pd.DataFrame({"price": [1.5, 2.7, 3.3]})
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            p = f.name
        df.to_excel(p, index=False)
        df_clean, _ = clean_excel(df.copy(), p, 0)
        assert df_clean["price"].dtype == "float64"
        os.unlink(p)


class TestDuplicateColumns:
    def test_dedup(self):
        df = pd.DataFrame([[1, 2, 3]], columns=["a", "a", "b"])
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            p = f.name
        df.to_excel(p, index=False)
        df_clean, _ = clean_excel(df.copy(), p, 0)
        assert len(set(df_clean.columns)) == len(df_clean.columns)
        os.unlink(p)


# ============================================================
# 清洗报告
# ============================================================


class TestCleaningReport:
    def test_write_and_read(self, tmp_path):
        cache = str(tmp_path / "test.parquet")
        report = CleaningReport(
            merged_cols_filled=2,
            hidden_rows_marked=3,
            original_shape=(100, 10),
            final_shape=(97, 9),
        )
        write_cleaning_report(cache, report)
        loaded = read_cleaning_report(cache)
        assert loaded is not None
        assert loaded.merged_cols_filled == 2
        assert loaded.original_shape == (100, 10)

    def test_no_changes_no_file(self, tmp_path):
        cache = str(tmp_path / "test2.parquet")
        report = CleaningReport()
        write_cleaning_report(cache, report)
        meta = Path(cache.replace(".parquet", ".meta.json"))
        assert not meta.exists()

    def test_read_nonexistent(self, tmp_path):
        assert read_cleaning_report(str(tmp_path / "nope.parquet")) is None

    def test_llm_text_with_markers(self):
        report = CleaningReport(
            merged_cols_filled=3,
            hidden_rows_marked=5,
            int_cols_fixed=1,
            has_auto_filter=True,
            hidden_cols_names=["辅助列"],
            original_shape=(100, 10),
            final_shape=(100, 12),
        )
        text = report.to_llm_text()
        assert "合并单元格已填充（3列）" in text
        assert "标记隐藏行（5行）" in text
        assert "_is_hidden = false" in text
        assert "自动筛选" in text
        assert "辅助列" in text

    def test_llm_text_empty(self):
        report = CleaningReport()
        assert report.to_llm_text() == ""

    def test_merge_reports(self):
        """多 Sheet 报告合并累加"""
        r1 = CleaningReport(
            merged_cols_filled=2,
            has_auto_filter=True, original_shape=(50, 5), final_shape=(48, 6),
        )
        r2 = CleaningReport(
            merged_cols_filled=1, hidden_rows_marked=3,
            original_shape=(30, 5), final_shape=(30, 6),
        )
        r1.merge(r2)
        assert r1.merged_cols_filled == 3
        assert r1.hidden_rows_marked == 3
        assert r1.has_auto_filter is True
        assert r1.original_shape == (80, 5)
        assert r1.final_shape == (78, 6)


# ============================================================
# 集成场景
# ============================================================


class TestMixedScenario:
    def test_merged_cells_cleaned(self, tmp_path):
        """合并单元格正确 ffill"""
        p = tmp_path / "mixed.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"], ws["B1"] = "分类", "金额"
        ws["A2"], ws["B2"] = "水果", 100
        ws["A3"], ws["B3"] = None, 200
        ws["A4"], ws["B4"] = "蔬菜", 300
        ws.merge_cells("A2:A3")
        wb.save(str(p))
        wb.close()

        df = pd.read_excel(str(p), engine="calamine")
        df_clean, report = clean_excel(df, str(p), 0)

        assert pd.isna(df_clean.iloc[1, 0])  # 不再自动 ffill
        assert "_is_summary" not in df_clean.columns

    def test_no_cleaning_needed(self, simple_xlsx):
        """普通文件完全不受影响"""
        df = pd.read_excel(simple_xlsx, engine="calamine")
        original_cols = list(df.columns)
        original_len = len(df)
        df_clean, report = clean_excel(df, simple_xlsx, 0)
        assert list(df_clean.columns) == original_cols
        assert len(df_clean) == original_len
        assert not report.has_changes()


class TestMultiLevelHeader:
    def test_two_level_header(self, tmp_path):
        """二级表头展平为 大类_小类（基于合并元数据检测）"""
        from services.agent.excel_cleaner import _detect_structure

        p = tmp_path / "multi_header.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"], ws["B1"], ws["C1"], ws["D1"] = "销售", "销售", "库存", "库存"
        ws["A2"], ws["B2"], ws["C2"], ws["D2"] = "数量", "金额", "入库", "出库"
        for i in range(3, 8):
            ws[f"A{i}"], ws[f"B{i}"] = (i - 2) * 10, (i - 2) * 100
            ws[f"C{i}"], ws[f"D{i}"] = (i - 2) * 5, (i - 2) * 3
        ws.merge_cells("A1:B1")
        ws.merge_cells("C1:D1")
        wb.save(str(p))
        wb.close()

        # 检测多级表头：合并元数据驱动
        structure = _detect_structure(str(p), 0)
        merged = structure.merged_ranges if structure else None

        df_raw = pd.read_excel(str(p), engine="calamine", header=None, nrows=20)
        header_row = detect_header_row(df_raw.values.tolist())
        actual_start, depth = detect_header_depth(header_row, merged)
        assert depth == 2
        assert actual_start == 0

        # 读取并展平
        header_param = list(range(actual_start, actual_start + depth))
        df = pd.read_excel(str(p), engine="calamine", header=header_param)
        assert isinstance(df.columns, pd.MultiIndex)

        df_clean, _ = clean_excel(df, str(p), 0, actual_start, structure)
        assert "销售_数量" in df_clean.columns
        assert "销售_金额" in df_clean.columns
        assert "库存_入库" in df_clean.columns
        assert "库存_出库" in df_clean.columns
        assert len(df_clean) == 5

    def test_single_header_not_affected(self, simple_xlsx):
        """无合并 → depth=1，不做展平"""
        _, depth = detect_header_depth(header_row=0, merged_ranges=None)
        assert depth == 1

    def test_flatten_multi_index(self):
        """直接测试 _flatten_multi_header"""
        arrays = [["A", "A", "B"], ["x", "y", "z"]]
        mi = pd.MultiIndex.from_arrays(arrays)
        df = pd.DataFrame([[1, 2, 3]], columns=mi)
        _flatten_multi_header(df)
        assert list(df.columns) == ["A_x", "A_y", "B_z"]

    def test_flatten_skips_normal_columns(self):
        """普通 Index 不做展平"""
        df = pd.DataFrame({"a": [1], "b": [2]})
        _flatten_multi_header(df)
        assert list(df.columns) == ["a", "b"]

    def test_long_name_preserved(self):
        """展平后不截断，保留完整列名"""
        arrays = [["A" * 30, "A" * 30], ["B" * 30, "C" * 30]]
        mi = pd.MultiIndex.from_arrays(arrays)
        df = pd.DataFrame([[1, 2]], columns=mi)
        _flatten_multi_header(df)
        assert df.columns[0] == "A" * 30 + "_" + "B" * 30  # 61 字符，不截断


class TestXlsxSafetyValve:
    def test_corrupted_file_degrades(self, tmp_path):
        """损坏文件降级到只做 Layer 3"""
        p = tmp_path / "corrupt.xlsx"
        p.write_bytes(b"PK not really a zip")
        df = pd.DataFrame({"a": [1, 2]})
        df_clean, report = clean_excel(df, str(p), 0)
        assert len(df_clean) == 2  # 不崩溃
        assert report.merged_cols_filled == 0
