"""
企微文件内容提取

将下载的文件（PDF/Word/Excel/TXT/CSV 等）转换为纯文本，
供 AI 分析。超长内容自动截断并提示。
"""

import csv
import io
import json
from typing import Optional, Tuple

from loguru import logger

# 提取文本的最大字符数（超过截断）
MAX_TEXT_LENGTH = 5000
# 支持的文件扩展名 → 解析器名
SUPPORTED_EXTENSIONS = {
    # 文本类
    "txt", "md", "log", "csv", "json", "xml", "yaml", "yml",
    "py", "js", "ts", "html", "css", "sql", "sh",
    # 文档类
    "pdf", "docx", "xlsx",
}


def is_supported(filename: str) -> bool:
    """判断文件类型是否支持解析"""
    ext = _get_ext(filename)
    return ext in SUPPORTED_EXTENSIONS


def parse_file(data: bytes, filename: str) -> Tuple[str, bool]:
    """提取文件文本内容

    Args:
        data: 文件二进制内容
        filename: 文件名（用于判断类型）

    Returns:
        (提取的文本, 是否被截断)
    """
    ext = _get_ext(filename)

    try:
        if ext == "pdf":
            text = _parse_pdf(data)
        elif ext == "docx":
            text = _parse_docx(data)
        elif ext == "xlsx":
            text = _parse_xlsx(data)
        elif ext == "csv":
            text = _parse_csv(data)
        elif ext == "json":
            text = _parse_json(data)
        else:
            text = _parse_text(data)
    except Exception as e:
        logger.warning(f"File parse failed | filename={filename} | error={e}")
        return f"[文件解析失败: {filename}]", False

    if not text or not text.strip():
        return f"[文件内容为空: {filename}]", False

    text = text.strip()
    truncated = len(text) > MAX_TEXT_LENGTH
    if truncated:
        text = text[:MAX_TEXT_LENGTH]

    return text, truncated


def _get_ext(filename: str) -> str:
    """提取小写扩展名"""
    if not filename or "." not in filename:
        return ""
    return filename.rsplit(".", 1)[-1].lower()


def _parse_pdf(data: bytes) -> str:
    """解析 PDF 文件"""
    from PyPDF2 import PdfReader

    reader = PdfReader(io.BytesIO(data))
    pages = []
    for i, page in enumerate(reader.pages):
        page_text = page.extract_text() or ""
        if page_text.strip():
            pages.append(f"[第{i + 1}页]\n{page_text.strip()}")
        # 提前退出：已有足够文本
        if sum(len(p) for p in pages) > MAX_TEXT_LENGTH:
            break
    return "\n\n".join(pages)


def _parse_docx(data: bytes) -> str:
    """解析 Word 文档"""
    from docx import Document

    doc = Document(io.BytesIO(data))
    paragraphs = []
    total_len = 0
    for para in doc.paragraphs:
        text = para.text.strip()
        if text:
            paragraphs.append(text)
            total_len += len(text)
            if total_len > MAX_TEXT_LENGTH:
                break
    return "\n".join(paragraphs)


def _parse_xlsx(data: bytes) -> str:
    """解析 Excel 文件（所有 sheet，表格格式输出）"""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        parts = []
        total_len = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                if any(cells):
                    rows.append(" | ".join(cells))
                    total_len += len(rows[-1])
                if total_len > MAX_TEXT_LENGTH:
                    break

            if rows:
                header = f"[Sheet: {sheet_name}]"
                parts.append(f"{header}\n" + "\n".join(rows))

            if total_len > MAX_TEXT_LENGTH:
                break

        return "\n\n".join(parts)
    finally:
        wb.close()


def _parse_csv(data: bytes) -> str:
    """解析 CSV 文件"""
    text = _decode_bytes(data)
    reader = csv.reader(io.StringIO(text))
    rows = []
    total_len = 0
    for row in reader:
        line = " | ".join(row)
        rows.append(line)
        total_len += len(line)
        if total_len > MAX_TEXT_LENGTH:
            break
    return "\n".join(rows)


def _parse_json(data: bytes) -> str:
    """解析 JSON 文件（格式化输出）"""
    text = _decode_bytes(data)
    try:
        obj = json.loads(text)
        return json.dumps(obj, ensure_ascii=False, indent=2)
    except json.JSONDecodeError:
        return text


def _parse_text(data: bytes) -> str:
    """通用文本文件解析"""
    return _decode_bytes(data)


def _decode_bytes(data: bytes) -> str:
    """尝试 UTF-8 → GBK 解码"""
    for encoding in ("utf-8", "gbk", "latin-1"):
        try:
            return data.decode(encoding)
        except (UnicodeDecodeError, LookupError):
            continue
    return data.decode("utf-8", errors="replace")
