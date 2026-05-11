"""excel_reader 单元测试

覆盖：
- _col_letter 列号转换
- _read_sheet_structured 公式提取 / 空值跳过 / 合并单元格去重 / 区域分隔
- _format_structured_output Row标号 / 小文件全量 / 大文件截断 / 公式对照独立段落
- _write_staging_parquet Parquet 输出 / 公式计算值
- read_excel_structured 完整流程 / 多 sheet / 空文件
"""
import os
import shutil
import tempfile
from pathlib import Path

import pytest

from services.agent.excel_reader import (
    _FORMULA_PREFIX,
    _col_letter,
    _format_structured_output,
    _read_sheet_structured,
    _write_staging_parquet,
    read_excel_structured,
)


# ── fixtures ──

@pytest.fixture
def tmp_dir():
    d = tempfile.mkdtemp()
    yield d
    shutil.rmtree(d, ignore_errors=True)


def _make_formula_excel(path: str) -> None:
    """创建带公式的测试 Excel"""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "公摊"
    ws["A1"] = "部门"
    ws["B1"] = "年费"
    ws["C1"] = "月费"
    ws["A2"] = "义乌"
    ws["B2"] = 76800
    ws["C2"] = "=B2/12"
    ws["A3"] = "金华"
    ws["B3"] = 50000
    ws["C3"] = "=B3/12"
    ws["C4"] = "=SUM(C2:C3)"

    ws2 = wb.create_sheet("明细")
    ws2["A1"] = "项目"
    ws2["B1"] = "金额"
    ws2["A2"] = "水电"
    ws2["B2"] = 3000
    wb.save(path)


def _make_empty_excel(path: str) -> None:
    from openpyxl import Workbook
    wb = Workbook()
    wb.save(path)


# ============================================================
# _col_letter
# ============================================================

class TestColLetter:
    def test_single_letter(self):
        assert _col_letter(1) == "A"
        assert _col_letter(26) == "Z"

    def test_double_letter(self):
        assert _col_letter(27) == "AA"
        assert _col_letter(28) == "AB"
        assert _col_letter(52) == "AZ"

    def test_triple_letter(self):
        assert _col_letter(703) == "AAA"


# ============================================================
# _read_sheet_structured
# ============================================================

class TestReadSheetStructured:
    def test_formula_extracted(self, tmp_dir):
        """公式单元格正确提取"""
        path = os.path.join(tmp_dir, "test.xlsx")
        _make_formula_excel(path)
        rows, formula_values, cross_refs, total_rows, total_cols, formula_count = (
            _read_sheet_structured(path)
        )
        assert formula_count == 3  # C2, C3, C4
        assert len(formula_values) == 3

    def test_empty_cells_skipped(self, tmp_dir):
        """空单元格不出现在输出中"""
        path = os.path.join(tmp_dir, "test.xlsx")
        _make_formula_excel(path)
        rows, *_ = _read_sheet_structured(path)
        all_cells = [cell for _rn, cells in rows for cell in cells if cells != ["---"]]
        for cell in all_cells:
            assert "None" not in cell

    def test_specific_sheet(self, tmp_dir):
        """指定 sheet 名读取"""
        path = os.path.join(tmp_dir, "test.xlsx")
        _make_formula_excel(path)
        rows, formula_values, _, total_rows, _, formula_count = (
            _read_sheet_structured(path, sheet_name="明细")
        )
        assert formula_count == 0
        assert total_rows == 2

    def test_cell_coordinates(self, tmp_dir):
        """每个值都带单元格编号"""
        path = os.path.join(tmp_dir, "test.xlsx")
        _make_formula_excel(path)
        rows, *_ = _read_sheet_structured(path)
        first_rn, first_cells = rows[0]
        assert any("A1:" in c for c in first_cells)
        assert any("B1:" in c for c in first_cells)

    def test_row_numbers(self, tmp_dir):
        """返回正确的行号"""
        path = os.path.join(tmp_dir, "test.xlsx")
        _make_formula_excel(path)
        rows, *_ = _read_sheet_structured(path)
        row_numbers = [rn for rn, cells in rows if cells != ["---"]]
        assert row_numbers[0] == 1
        assert row_numbers[1] == 2

    def test_area_separation(self, tmp_dir):
        """空行后出现新数据时插入分隔"""
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "区域1"
        ws["A3"] = "区域2"  # Row 2 空
        path = os.path.join(tmp_dir, "test.xlsx")
        wb.save(path)

        rows, *_ = _read_sheet_structured(path)
        separators = [1 for _rn, cells in rows if cells == ["---"]]
        assert len(separators) == 1


# ============================================================
# _format_structured_output
# ============================================================

class TestFormatStructuredOutput:
    def test_row_labels(self):
        """输出带 Row 行标号"""
        rows = [(1, ["A1:部门", "B1:年费"]), (2, ["A2:义乌"])]
        text = _format_structured_output(
            rows, [], [], 2, 2, 0, "Sheet1", "", "test.xlsx",
        )
        assert "Row1:" in text
        assert "Row2:" in text

    def test_formula_separate_section(self):
        """公式对照表独立段落"""
        rows = [(1, ["A1:test"]), (2, [f"C2:{_FORMULA_PREFIX}=B2/12"])]
        fv = [("C2", "=B2/12", "6400")]
        text = _format_structured_output(
            rows, fv, [], 2, 2, 1, "Sheet1", "", "test.xlsx",
        )
        assert "关键单元格公式 vs 值" in text
        assert "C2: 公式==B2/12  |  计算值=6400" in text

    def test_no_formula(self):
        """无公式时显示0"""
        rows = [(1, ["A1:test"])]
        text = _format_structured_output(
            rows, [], [], 1, 1, 0, "Sheet1", "", "test.xlsx",
        )
        assert "0个" in text

    def test_cross_refs(self):
        """跨 sheet 引用输出"""
        rows = [(1, ["A1:test"])]
        fv = [("D4", "=明细!E7", "123")]
        text = _format_structured_output(
            rows, fv, ["D4 → =明细!E7"], 1, 1, 1, "Sheet1", "", "test.xlsx",
        )
        assert "跨Sheet引用" in text

    def test_follow_up_hint(self):
        """后续查询提示用 file_read"""
        rows = [(1, ["A1:test"])]
        text = _format_structured_output(
            rows, [], [], 1, 1, 0, "Sheet1", "", "test.xlsx",
        )
        assert 'file_read(path="test.xlsx"' in text

    def test_large_file_truncated(self):
        """大文件截断 + 底部总行列数"""
        rows = [(i, [f"A{i}:val{i}"]) for i in range(1, 20001)]
        text = _format_structured_output(
            rows, [], [], 20000, 1, 0, "Sheet1", "", "test.xlsx",
        )
        assert "省略" in text
        assert "总行数: 20000" in text


# ============================================================
# _write_staging_parquet
# ============================================================

class TestWriteStagingParquet:
    def test_parquet_created(self, tmp_dir):
        """staging Parquet 文件创建"""
        rows = [(1, ["A1:部门", "B1:年费"]), (2, [f"C2:{_FORMULA_PREFIX}=B2/12"])]
        fv = [("C2", "=B2/12", "6400")]
        staging = os.path.join(tmp_dir, "staging")
        _write_staging_parquet(rows, fv, staging, "test.xlsx")
        parquets = list(Path(staging).glob("*.parquet"))
        assert len(parquets) == 1

    def test_parquet_columns(self, tmp_dir):
        """Parquet 包含 cell/row/col/value/formula 列"""
        import pandas as pd
        rows = [(1, ["A1:部门"]), (2, [f"C2:{_FORMULA_PREFIX}=B2/12"])]
        fv = [("C2", "=B2/12", "6400")]
        staging = os.path.join(tmp_dir, "staging")
        _write_staging_parquet(rows, fv, staging, "test.xlsx")
        df = pd.read_parquet(str(list(Path(staging).glob("*.parquet"))[0]))
        assert set(df.columns) == {"cell", "row", "col", "value", "formula"}

    def test_formula_has_calc_value(self, tmp_dir):
        """公式单元格有计算值"""
        import pandas as pd
        rows = [(2, [f"C2:{_FORMULA_PREFIX}=B2/12"])]
        fv = [("C2", "=B2/12", "6400")]
        staging = os.path.join(tmp_dir, "staging")
        _write_staging_parquet(rows, fv, staging, "test.xlsx")
        df = pd.read_parquet(str(list(Path(staging).glob("*.parquet"))[0]))
        row = df.iloc[0]
        assert row["formula"] == "=B2/12"
        assert row["value"] == "6400"

    def test_value_cell_no_formula(self, tmp_dir):
        """普通值单元格 formula 为空"""
        import pandas as pd
        rows = [(1, ["A1:部门"])]
        staging = os.path.join(tmp_dir, "staging")
        _write_staging_parquet(rows, [], staging, "test.xlsx")
        df = pd.read_parquet(str(list(Path(staging).glob("*.parquet"))[0]))
        row = df.iloc[0]
        assert row["value"] == "部门"
        assert row["formula"] is None or pd.isna(row["formula"])


# ============================================================
# read_excel_structured (完整流程)
# ============================================================

class TestReadExcelStructured:
    @pytest.mark.asyncio
    async def test_full_flow(self, tmp_dir):
        """完整流程：公式 Excel → 结构化输出 + staging"""
        path = os.path.join(tmp_dir, "formula.xlsx")
        _make_formula_excel(path)
        staging = os.path.join(tmp_dir, "staging")

        result = await read_excel_structured(path, None, staging)
        assert result.status == "success"
        assert "公摊" in result.summary
        assert "Row" in result.summary
        assert "关键单元格公式 vs 值" in result.summary
        assert 'file_read(path=' in result.summary
        assert list(Path(staging).glob("*.parquet"))

    @pytest.mark.asyncio
    async def test_specific_sheet(self, tmp_dir):
        path = os.path.join(tmp_dir, "test.xlsx")
        _make_formula_excel(path)
        staging = os.path.join(tmp_dir, "staging")
        result = await read_excel_structured(path, "明细", staging)
        assert result.status == "success"

    @pytest.mark.asyncio
    async def test_empty_excel(self, tmp_dir):
        path = os.path.join(tmp_dir, "empty.xlsx")
        _make_empty_excel(path)
        staging = os.path.join(tmp_dir, "staging")
        result = await read_excel_structured(path, None, staging)
        assert result.status == "empty"

    @pytest.mark.asyncio
    async def test_multi_sheet_overview(self, tmp_dir):
        path = os.path.join(tmp_dir, "multi.xlsx")
        _make_formula_excel(path)
        staging = os.path.join(tmp_dir, "staging")
        result = await read_excel_structured(path, None, staging)
        assert "Sheet 概览" in result.summary
