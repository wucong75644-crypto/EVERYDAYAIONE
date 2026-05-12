"""
Excel 结构化读取（openpyxl 两次读取，保留公式+编号）

输出格式对标 Claude：
  Row1: ['A1:公共费用', 'C1:按年摊销金额', 'D1:费用金额']
  Row2: ['B2:仓库租金', 'C2:283800', 'D2:[公式]=C2/12']

  === 关键单元格公式 vs 值 ===
  D2: 公式==C2/12  |  计算值=6400

空单元格跳过，合并单元格去重，不做 ffill。

设计文档：docs/document/TECH_file_read统一工具.md
"""
from __future__ import annotations

import asyncio
import time
from pathlib import Path
from typing import Optional

from loguru import logger

from services.agent.agent_result import AgentResult

# ── 常量 ──

_MAX_ROWS_FULL = 10000       # 全量输出行数上限，超过则截断
_MAX_FORMULA_CELLS = 500     # 公式单元格输出上限（Pass 1 扫描提前终止）
_PREVIEW_HEAD = 5            # 大文件预览：前 N 行
_PREVIEW_TAIL = 10           # 大文件预览：后 N 行（对标 Claude）
_FORMULA_PREFIX = "[公式]"   # 公式标记前缀


def _col_letter(col_idx: int) -> str:
    """1-indexed 列号 → Excel 列字母（1=A, 27=AA）。"""
    result = ""
    while col_idx > 0:
        col_idx, remainder = divmod(col_idx - 1, 26)
        result = chr(65 + remainder) + result
    return result


def _read_sheet_structured(
    excel_path: str,
    sheet_name: str | None = None,
) -> tuple[
    list[tuple[int, list[str]]],       # (行号, 单元格列表)
    list[tuple[str, str, str]],        # (坐标, 公式, 计算值) — 公式对照表
    list[str],                          # 跨 sheet 引用
    int, int, int,                      # total_rows, total_cols, formula_count
]:
    """读取单个 sheet，返回结构化行数据 + 公式对照表。"""
    import openpyxl

    # Pass 1: 读公式字符串
    wb_formula = openpyxl.load_workbook(
        excel_path, read_only=True, data_only=False,
    )
    ws_formula = wb_formula[sheet_name] if sheet_name else wb_formula.active
    if ws_formula is None:
        wb_formula.close()
        return [], [], [], 0, 0, 0

    formulas: dict[tuple[int, int], str] = {}
    _scan_done = False
    for row in ws_formula.iter_rows():
        if _scan_done:
            break
        for cell in row:
            if not hasattr(cell, "data_type"):
                continue
            if cell.data_type == "f" and cell.value:
                formulas[(cell.row, cell.column)] = str(cell.value)
                if len(formulas) >= _MAX_FORMULA_CELLS:
                    _scan_done = True
                    break
    wb_formula.close()

    # Pass 2: 读计算值
    wb_value = openpyxl.load_workbook(
        excel_path, read_only=True, data_only=True,
    )
    ws_value = wb_value[sheet_name] if sheet_name else wb_value.active
    if ws_value is None:
        wb_value.close()
        return [], [], [], 0, 0, 0

    rows: list[tuple[int, list[str]]] = []
    formula_values: list[tuple[str, str, str]] = []  # 公式对照表
    cross_refs: list[str] = []
    total_rows = 0
    max_col = 0
    empty_streak = 0

    for row in ws_value.iter_rows():
        total_rows += 1
        row_cells: list[str] = []

        for cell in row:
            if not hasattr(cell, "column") or cell.column is None:
                continue
            if cell.column > max_col:
                max_col = cell.column

            col_letter = _col_letter(cell.column)
            coord = f"{col_letter}{cell.row}"
            formula = formulas.get((cell.row, cell.column))

            if formula:
                row_cells.append(f"{coord}:{_FORMULA_PREFIX}{formula}")
                val_str = str(cell.value) if cell.value is not None else "未缓存"
                formula_values.append((coord, formula, val_str))
                if "!" in formula:
                    cross_refs.append(f"{coord} → {formula}")
            elif cell.value is not None:
                row_cells.append(f"{coord}:{cell.value}")

        if row_cells:
            if empty_streak >= 1 and rows:
                rows.append((0, ["---"]))
            empty_streak = 0
            rows.append((total_rows, row_cells))
        else:
            empty_streak += 1

    wb_value.close()
    return rows, formula_values, cross_refs, total_rows, max_col, len(formulas)


def _format_structured_output(
    rows: list[tuple[int, list[str]]],
    formula_values: list[tuple[str, str, str]],
    cross_refs: list[str],
    total_rows: int,
    total_cols: int,
    formula_count: int,
    sheet_name: str,
    sheet_overview: str,
    rel_path: str,
) -> str:
    """格式化输出，对标 Claude 格式。"""
    lines: list[str] = []
    lines.append(f"=== Sheet: {sheet_name} ===")
    lines.append("")

    is_large = total_rows > _MAX_ROWS_FULL

    if is_large:
        # 大文件：列名单独列 + 前5行 + 后10行 + 底部总行列数
        # 找列名行（通常是第一或第二个非分隔行）
        data_rows = [(rn, cells) for rn, cells in rows if cells != ["---"]]
        head = data_rows[:_PREVIEW_HEAD]
        tail = data_rows[-_PREVIEW_TAIL:] if len(data_rows) > _PREVIEW_TAIL else []

        # 列名提取（从前两行中找包含最多单元格的行）
        if len(data_rows) >= 2:
            r1_len = len(data_rows[0][1])
            r2_len = len(data_rows[1][1])
            header_row = data_rows[1] if r2_len > r1_len else data_rows[0]
            lines.append("完整列名:")
            for cell_str in header_row[1]:
                lines.append(f"  {cell_str}")
            lines.append("")

        for rn, cells in head:
            lines.append(f"  Row{rn}: {cells}")
        lines.append(f"  ... (省略中间数据) ...")
        for rn, cells in tail:
            lines.append(f"  Row{rn}: {cells}")
        lines.append(f"  总行数: {total_rows}, 总列数: {total_cols}")
    else:
        # 小文件：全量输出，每行带 Row 标号
        for rn, cells in rows:
            if cells == ["---"]:
                lines.append("")
            else:
                lines.append(f"  Row{rn}: {cells}")

    # 公式 vs 值对照表（独立段落）
    if formula_values:
        lines.append("")
        lines.append(f"=== {sheet_name} - 关键单元格公式 vs 值 ===")
        for coord, formula, val in formula_values:
            lines.append(f"{coord}: 公式={formula}  |  计算值={val}")
        if cross_refs:
            lines.append("")
            for ref in cross_refs[:10]:
                lines.append(f"跨Sheet引用: {ref}")
            if len(cross_refs) > 10:
                lines.append(f"... 等{len(cross_refs)}个跨Sheet引用")
    elif formula_count == 0:
        lines.append("")
        lines.append("公式统计: 0个（纯数据表）")

    # Sheet 概览
    if sheet_overview:
        lines.append("")
        lines.append(sheet_overview)

    # 后续操作提示
    lines.append("")
    lines.append(f"后续查询: file_read(path=\"{rel_path}\", sql=\"SELECT ... FROM data\")")

    return "\n".join(lines)


def _build_all_sheets_preview(
    excel_path: str, preview_rows: int = 3,
) -> tuple[str, list[str]]:
    """所有 sheet 预览（对标 Claude：Sheet 列表 + 每个 sheet 行列数 + 前 3 行含公式）。"""
    import openpyxl

    # 用 data_only=False 拿公式，read_only=True 保证低内存
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=False)
    sheet_names = wb.sheetnames

    if not sheet_names:
        wb.close()
        return "", []

    # 同时打开 data_only=True 拿计算值
    wb_val = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

    lines = [f"Sheet列表: {sheet_names}"]

    for sn in sheet_names[:10]:
        ws_f = wb[sn]
        ws_v = wb_val[sn]
        lines.append(f"\n=== Sheet: {sn} ===")
        lines.append(f"  行数: {ws_f.max_row}, 列数: {ws_f.max_column}")

        # 只读前 preview_rows 行（不全量遍历）
        row_idx = 0
        val_iter = ws_v.iter_rows()
        for row_f in ws_f.iter_rows():
            row_idx += 1
            if row_idx > preview_rows:
                break
            # 同步读值行
            try:
                row_v = next(val_iter)
            except StopIteration:
                row_v = row_f

            row_cells: list[str] = []
            for cell_f, cell_v in zip(row_f, row_v):
                if not hasattr(cell_f, "column") or cell_f.column is None:
                    continue
                col_letter = _col_letter(cell_f.column)
                coord = f"{col_letter}{cell_f.row}"
                if hasattr(cell_f, "data_type") and cell_f.data_type == "f" and cell_f.value:
                    row_cells.append(f"{coord}:{_FORMULA_PREFIX}{cell_f.value}")
                elif cell_f.value is not None:
                    row_cells.append(f"{coord}:{cell_v.value if cell_v.value is not None else cell_f.value}")
            if row_cells:
                lines.append(f"  Row{row_idx}: {row_cells}")

    wb.close()
    wb_val.close()

    if len(sheet_names) > 10:
        lines.append(f"\n... 等{len(sheet_names)}个 Sheet")

    return "\n".join(lines), sheet_names


async def read_excel_structured(
    abs_path: str,
    sheet: str | None,
    staging_dir: str,
) -> AgentResult:
    """Excel 结构化读取入口（异步包装）。"""
    start = time.monotonic()
    filename = Path(abs_path).name

    loop = asyncio.get_running_loop()
    try:
        result = await loop.run_in_executor(
            None, _read_excel_structured_sync,
            abs_path, sheet, staging_dir,
        )
        elapsed = time.monotonic() - start
        logger.info(
            f"Excel structured read | file={filename} | "
            f"elapsed={elapsed:.2f}s | status={result.status}"
        )
        return result
    except Exception as e:
        logger.error(f"Excel structured read failed | file={filename} | error={e}")
        return AgentResult(
            summary=f"Excel 读取失败: {e}",
            status="error",
            error_message=str(e),
        )


def _read_excel_structured_sync(
    abs_path: str,
    sheet: str | None,
    staging_dir: str,
) -> AgentResult:
    """Excel 结构化读取（同步，线程池执行）。

    不指定 sheet：返回所有 sheet 预览（对标 Claude 第一次调用）
    指定 sheet：返回该 sheet 完整内容 + 公式对照表
    """
    filename = Path(abs_path).name

    # 不指定 sheet → 全 sheet 预览
    if not sheet:
        overview_text, sheet_names = _build_all_sheets_preview(abs_path)
        if not overview_text:
            return AgentResult(
                summary=f"Excel 文件为空或无数据: {filename}",
                status="empty",
            )
        overview_text += f"\n\n读取指定 Sheet: file_read(path=\"{filename}\", sheet=\"Sheet名\")"
        return AgentResult(summary=overview_text, status="success")

    # 指定 sheet → 完整内容
    import openpyxl
    wb = openpyxl.load_workbook(abs_path, read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    wb.close()

    target_sheet = sheet
    if sheet_names:
        from services.agent.data_query_cache import fuzzy_match_sheet
        target_sheet = fuzzy_match_sheet(sheet, sheet_names)

    rows, formula_values, cross_refs, total_rows, total_cols, formula_count = (
        _read_sheet_structured(abs_path, target_sheet)
    )

    if not rows:
        return AgentResult(
            summary=f"Sheet \"{target_sheet}\" 为空或无数据",
            status="empty",
        )

    text = _format_structured_output(
        rows, formula_values, cross_refs,
        total_rows, total_cols, formula_count,
        target_sheet, "", filename,
    )

    _write_staging_parquet(rows, formula_values, staging_dir, filename)

    return AgentResult(summary=text, status="success")


def _write_staging_parquet(
    rows: list[tuple[int, list[str]]],
    formula_values: list[tuple[str, str, str]],
    staging_dir: str,
    filename: str,
) -> None:
    """将结构化行数据写入 staging Parquet（cell/row/col/value/formula）。"""
    import re

    staging = Path(staging_dir)
    staging.mkdir(parents=True, exist_ok=True)

    # 公式对照表 → {坐标: (公式, 计算值)}
    fv_map = {coord: (formula, val) for coord, formula, val in formula_values}

    records: list[dict] = []
    pattern = re.compile(r"^([A-Z]+)(\d+):(.*)$")

    for _rn, row_cells in rows:
        if row_cells == ["---"]:
            continue
        for cell_str in row_cells:
            m = pattern.match(cell_str)
            if not m:
                continue
            col_letter, row_num, content = m.group(1), int(m.group(2)), m.group(3)
            coord = f"{col_letter}{row_num}"

            if content.startswith(_FORMULA_PREFIX):
                formula_str = content[len(_FORMULA_PREFIX):]
                # 计算值从公式对照表取
                calc_value = fv_map.get(coord, (None, None))[1]
                records.append({
                    "cell": coord, "row": row_num, "col": col_letter,
                    "value": calc_value, "formula": formula_str,
                })
            else:
                records.append({
                    "cell": coord, "row": row_num, "col": col_letter,
                    "value": str(content), "formula": None,
                })

    if not records:
        return

    try:
        import pandas as pd
        df = pd.DataFrame(records)
        stem = Path(filename).stem
        parquet_name = f"_structured_{stem}.parquet"
        df.to_parquet(str(staging / parquet_name), index=False, engine="pyarrow")
        logger.info(
            f"Structured Parquet written | file={parquet_name} | cells={len(records)}"
        )
    except Exception as e:
        logger.warning(f"Structured Parquet write failed: {e}")
