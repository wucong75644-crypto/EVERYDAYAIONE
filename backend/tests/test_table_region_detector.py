"""table_region_detector.py 单元测试。"""
from __future__ import annotations

from pathlib import Path

from services.agent.table_region_detector import (
    TableRegion,
    _extract_region_name,
    _find_empty_rows,
    _is_header_row,
    _is_single_value_row,
    _split_by_empty_rows,
    detect_table_regions,
)


class TestFindEmptyRows:
    def test_basic(self):
        rows = [["a", "b"], [None, None], ["c", "d"]]
        assert _find_empty_rows(rows) == [1]

    def test_empty_string(self):
        rows = [["a"], ["", "  "], ["b"]]
        assert _find_empty_rows(rows) == [1]

    def test_no_empty(self):
        rows = [["a", "b"], ["c", "d"]]
        assert _find_empty_rows(rows) == []

    def test_multiple_empty(self):
        rows = [["a"], [None], ["b"], [None, None], ["c"]]
        assert _find_empty_rows(rows) == [1, 3]


class TestSplitByEmptyRows:
    def test_single_split(self):
        assert _split_by_empty_rows(5, [2]) == [(0, 2), (3, 5)]

    def test_consecutive_empty(self):
        assert _split_by_empty_rows(5, [1, 2]) == [(0, 1), (3, 5)]

    def test_empty_at_start(self):
        assert _split_by_empty_rows(4, [0]) == [(1, 4)]

    def test_empty_at_end(self):
        assert _split_by_empty_rows(4, [3]) == [(0, 3)]


class TestIsHeaderRow:
    def test_header(self):
        assert _is_header_row(["订单号", "金额", "日期"], threshold=2) is True

    def test_data_row(self):
        assert _is_header_row(["001", 100, "2024-01-01"], threshold=2) is False

    def test_sparse(self):
        assert _is_header_row(["标题", None, None, None], threshold=2) is False


class TestIsSingleValueRow:
    def test_title(self):
        assert _is_single_value_row(["退货表", None, None]) is True

    def test_normal(self):
        assert _is_single_value_row(["a", "b"]) is False

    def test_empty(self):
        assert _is_single_value_row([None, None]) is False


class TestExtractRegionName:
    def test_found(self):
        rows = [
            ["退货表", None, None],
            ["编号", "金额", "原因"],
            ["R001", 50, "质量问题"],
        ]
        assert _extract_region_name(rows, 0, 1) == "退货表"

    def test_not_found(self):
        rows = [
            ["编号", "金额", "原因"],
            ["R001", 50, "质量问题"],
        ]
        assert _extract_region_name(rows, 0, 0) is None


class TestDetectTableRegions:
    def test_two_tables(self):
        """经典场景：空行分隔的两张表。"""
        rows = [
            ["订单表", None, None],
            ["订单号", "金额", "日期"],
            ["001", 100, "2024-01-01"],
            ["002", 200, "2024-01-02"],
            [None, None, None],           # 空行分隔
            ["退货表", None, None],
            ["编号", "金额", "原因"],
            ["R001", 50, "质量问题"],
            ["R002", 80, "尺寸不合"],
        ]
        regions = detect_table_regions(rows)
        assert len(regions) == 2

        assert regions[0].name == "订单表"
        assert regions[0].columns == ["订单号", "金额", "日期"]
        assert regions[0].row_count == 2

        assert regions[1].name == "退货表"
        assert regions[1].columns == ["编号", "金额", "原因"]
        assert regions[1].row_count == 2

    def test_single_table(self):
        """单表格 → 返回空列表，走现有逻辑。"""
        rows = [
            ["订单号", "金额", "日期"],
            ["001", 100, "2024-01-01"],
            ["002", 200, "2024-01-02"],
        ]
        assert detect_table_regions(rows) == []

    def test_no_name_row(self):
        """无表格名称行 → name=None。"""
        rows = [
            ["订单号", "金额", "日期"],
            ["001", 100, "2024-01-01"],
            [None, None, None],
            ["编号", "金额", "原因"],
            ["R001", 50, "质量问题"],
        ]
        regions = detect_table_regions(rows)
        assert len(regions) == 2
        assert regions[0].name is None
        assert regions[1].name is None

    def test_empty_between_is_data_gap(self):
        """空行后面不是表头 → 不分割（空行是数据缺失）。"""
        rows = [
            ["订单号", "金额", "日期"],
            ["001", 100, "2024-01-01"],
            [None, None, None],           # 空行
            ["002", 200, "2024-01-02"],   # 不是表头（数字行）
            ["003", 300, "2024-01-03"],
        ]
        assert detect_table_regions(rows) == []

    def test_three_tables(self):
        """三张表（列数 ≥ 3 才能让单值名称行与表头行区分开）。"""
        rows = [
            ["表1", None, None],
            ["A", "B", "C"],
            [1, 2, 3],
            [None, None, None],
            ["表2", None, None],
            ["D", "E", "F"],
            [4, 5, 6],
            [None, None, None],
            ["表3", None, None],
            ["G", "H", "I"],
            [7, 8, 9],
        ]
        regions = detect_table_regions(rows)
        assert len(regions) == 3
        assert regions[0].name == "表1"
        assert regions[1].name == "表2"
        assert regions[2].name == "表3"

    def test_empty_input(self):
        assert detect_table_regions([]) == []


# ── convert_multi_region ──


class TestConvertMultiRegion:
    def _make_multi_region_xlsx(self, tmp_path):
        """创建包含两个表格区域的 Excel 文件。"""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        # 区域1：订单表
        ws["A1"] = "订单表"
        ws["A2"] = "订单号"
        ws["B2"] = "金额"
        ws["C2"] = "日期"
        ws["A3"] = "TB001"
        ws["B3"] = 100
        ws["C3"] = "2024-01-01"
        ws["A4"] = "TB002"
        ws["B4"] = 200
        ws["C4"] = "2024-01-02"
        # Row 5 空（分隔）
        # 区域2：退货表
        ws["A6"] = "退货表"
        ws["A7"] = "编号"
        ws["B7"] = "金额"
        ws["C7"] = "原因"
        ws["A8"] = "R001"
        ws["B8"] = 50
        ws["C8"] = "质量问题"
        xlsx_path = str(tmp_path / "multi.xlsx")
        wb.save(xlsx_path)
        wb.close()
        return xlsx_path

    def test_merged_parquet_with_region_col(self, tmp_path):
        """多区域合并为一个 Parquet + _region 列。"""
        from services.agent.table_region_detector import convert_multi_region, detect_table_regions
        import fastexcel, duckdb

        xlsx_path = self._make_multi_region_xlsx(tmp_path)
        staging = tmp_path / "staging"
        staging.mkdir()
        cache_path = str(staging / "cache.parquet")
        snapshot_path = str(staging / "cache.snapshot")

        reader = fastexcel.read_excel(xlsx_path)
        scan_raw = reader.load_sheet(0, header_row=None, n_rows=5000)
        regions = detect_table_regions(scan_raw.to_pandas().values.tolist())
        assert len(regions) == 2

        convert_multi_region(
            xlsx_path, cache_path, regions, reader.sheet_names,
            "Sheet1", 0.0, 0, snapshot_path,
        )
        # cache_path 存在（路径对齐）
        assert Path(cache_path).exists()
        assert Path(cache_path.replace(".parquet", ".meta.json")).exists()
        # _region 列标识来源
        df = duckdb.sql("SELECT * FROM read_parquet(?)", params=[cache_path]).to_df()
        assert "_region" in df.columns
        region_values = set(df["_region"].unique())
        assert len(region_values) == 2

    def test_session_files_single_entry(self, tmp_path):
        """多区域合并后 session_files 只注册一个文件。"""
        from services.agent.table_region_detector import convert_multi_region, detect_table_regions
        from services.agent.session_files import read_session_files
        import fastexcel

        xlsx_path = self._make_multi_region_xlsx(tmp_path)
        staging = tmp_path / "staging"
        staging.mkdir()
        cache_path = str(staging / "cache.parquet")
        snapshot_path = str(staging / "cache.snapshot")

        reader = fastexcel.read_excel(xlsx_path)
        scan_raw = reader.load_sheet(0, header_row=None, n_rows=5000)
        regions = detect_table_regions(scan_raw.to_pandas().values.tolist())

        convert_multi_region(
            xlsx_path, cache_path, regions, reader.sheet_names,
            "Sheet1", 0.0, 0, snapshot_path,
        )

        sf = read_session_files(str(staging))
        assert len(sf["files"]) == 1  # 合并为一个文件

    def test_meta_has_region_info(self, tmp_path):
        """meta 包含区域信息。"""
        from services.agent.table_region_detector import convert_multi_region, detect_table_regions
        from services.agent.file_meta import read_file_meta
        import fastexcel

        xlsx_path = self._make_multi_region_xlsx(tmp_path)
        staging = tmp_path / "staging"
        staging.mkdir()
        cache_path = str(staging / "cache.parquet")
        snapshot_path = str(staging / "cache.snapshot")

        reader = fastexcel.read_excel(xlsx_path)
        scan_raw = reader.load_sheet(0, header_row=None, n_rows=5000)
        regions = detect_table_regions(scan_raw.to_pandas().values.tolist())

        convert_multi_region(
            xlsx_path, cache_path, regions, reader.sheet_names,
            "Sheet1", 0.0, 0, snapshot_path,
        )

        meta = read_file_meta(cache_path)
        assert meta is not None
        assert "区域" in meta.summary.get("description", "")
