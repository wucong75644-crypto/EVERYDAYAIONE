"""
沙盒子进程隔离 E2E 测试

完整链路测试：SandboxExecutor.execute → _run_in_subprocess → sandbox_worker_entry
覆盖：执行链路、文件操作、安全拦截、超时杀死、并发隔离、边缘情况。
"""

import asyncio
import os
import time
from pathlib import Path

import pytest

from services.sandbox.executor import SandboxExecutor


# ============================================================
# Fixtures
# ============================================================

@pytest.fixture
def ws(tmp_path):
    """用户 workspace（含 output/staging 子目录）"""
    output = tmp_path / "下载"
    staging = tmp_path / "staging" / "conv_001"
    output.mkdir(parents=True)
    staging.mkdir(parents=True)
    return {
        "workspace": str(tmp_path),
        "output": str(output),
        "staging": str(staging),
    }


@pytest.fixture
def executor(ws):
    """标准执行器（完整路径配置）"""
    uploaded = []

    async def mock_upload(filename, size):
        uploaded.append(filename)
        return f"✅ 文件已生成: {filename}\n[FILE]https://cdn.test/{filename}|{filename}|application/octet-stream|{size}[/FILE]"

    ex = SandboxExecutor(
        timeout=30.0,
        max_result_chars=8000,
        workspace_dir=ws["workspace"],
        staging_dir=ws["staging"],
        output_dir=ws["output"],
        upload_fn=mock_upload,
    )
    ex._uploaded = uploaded  # 测试用：追踪上传的文件
    return ex


# ============================================================
# 1. 完整执行链路（子进程 spawn → exec → 返回）
# ============================================================

class TestExecutionEndToEnd:
    """完整执行链路 E2E"""

    @pytest.mark.asyncio
    async def test_simple_calculation(self, executor):
        result = await executor.execute("2 ** 10", "幂运算")
        assert "1024" in result.summary

    @pytest.mark.asyncio
    async def test_print_and_expression(self, executor):
        result = await executor.execute("print('hello')\n42", "混合输出")
        assert "hello" in result.summary
        assert "42" in result.summary

    @pytest.mark.asyncio
    async def test_multi_line_data_processing(self, executor):
        code = (
            "data = [1, 2, 3, 4, 5]\n"
            "total = sum(data)\n"
            "avg = total / len(data)\n"
            "print(f'total={total}, avg={avg}')"
        )
        result = await executor.execute(code, "数据处理")
        assert "total=15" in result.summary
        assert "avg=3.0" in result.summary

    @pytest.mark.asyncio
    async def test_json_processing(self, executor):
        code = (
            "data = json.dumps({'name': '测试', 'value': 42}, ensure_ascii=False)\n"
            "print(data)"
        )
        result = await executor.execute(code, "JSON处理")
        assert "测试" in result.summary
        assert "42" in result.summary

    @pytest.mark.asyncio
    async def test_datetime_usage(self, executor):
        result = await executor.execute("str(datetime.now().year)", "日期")
        assert "202" in result.summary

    @pytest.mark.asyncio
    async def test_decimal_precision(self, executor):
        code = "str(Decimal('0.1') + Decimal('0.2'))"
        result = await executor.execute(code, "精度")
        assert "0.3" in result.summary

    @pytest.mark.asyncio
    async def test_collections_usage(self, executor):
        code = "str(dict(Counter('aabbbc')))"
        result = await executor.execute(code, "Counter")
        assert "'b': 3" in result.summary


# ============================================================
# 2. 文件操作链路
# ============================================================

class TestFileOperations:
    """文件读写操作 E2E"""

    @pytest.mark.asyncio
    async def test_read_workspace_file_by_name(self, executor, ws):
        """chdir 后直接用文件名读取 workspace 文件"""
        Path(ws["workspace"], "input.txt").write_text("workspace data here")

        result = await executor.execute(
            "print(open('input.txt').read())", "读workspace文件",
        )
        assert "workspace data here" in result.summary

    @pytest.mark.asyncio
    async def test_read_workspace_csv(self, executor, ws):
        """读取 workspace 中的 CSV 文件"""
        Path(ws["workspace"], "data.csv").write_text("name,score\nAlice,90\nBob,85")

        code = (
            "lines = open('data.csv').readlines()\n"
            "print(f'rows={len(lines)-1}')\n"
            "print(lines[1].strip())"
        )
        result = await executor.execute(code, "读CSV")
        assert "rows=2" in result.summary
        assert "Alice,90" in result.summary

    @pytest.mark.asyncio
    async def test_read_subdirectory_file(self, executor, ws):
        """读取子目录中的文件"""
        sub = Path(ws["workspace"], "reports")
        sub.mkdir()
        (sub / "q1.txt").write_text("Q1 revenue: 1000")

        result = await executor.execute(
            "print(open('reports/q1.txt').read())", "读子目录文件",
        )
        assert "Q1 revenue: 1000" in result.summary

    @pytest.mark.asyncio
    async def test_read_chinese_filename(self, executor, ws):
        """中文文件名正常读取"""
        Path(ws["workspace"], "利润表.txt").write_text("利润: 50万")

        result = await executor.execute(
            "print(open('利润表.txt').read())", "中文文件名",
        )
        assert "利润: 50万" in result.summary

    @pytest.mark.asyncio
    async def test_read_staging_file(self, executor, ws):
        """读取 staging 目录的文件（用 STAGING_DIR 绝对路径）"""
        Path(ws["staging"], "data.json").write_text('{"total": 100}')

        code = (
            "import json\n"
            "with open(STAGING_DIR + '/data.json') as f:\n"
            "    data = json.load(f)\n"
            "print(data['total'])"
        )
        result = await executor.execute(code, "读staging")
        assert "100" in result.summary

    @pytest.mark.asyncio
    async def test_write_file_to_output_dir(self, executor, ws):
        """写文件到 OUTPUT_DIR → 主进程检测并上传"""
        code = (
            "with open(OUTPUT_DIR + '/result.json', 'w') as f:\n"
            "    json.dump({'answer': 42}, f)\n"
            "print('done')"
        )
        result = await executor.execute(code, "写文件")
        assert "done" in result.summary
        # 文件应被 auto_upload 检测到
        assert "result.json" in executor._uploaded

    @pytest.mark.asyncio
    async def test_write_xlsx_to_output_dir(self, executor, ws):
        """写 Excel 到 OUTPUT_DIR（需要 openpyxl 或 xlsxwriter）"""
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not installed")

        code = (
            "import openpyxl\n"
            "wb = openpyxl.Workbook()\n"
            "ws = wb.active\n"
            "ws.append(['Name', 'Score'])\n"
            "ws.append(['Alice', 90])\n"
            "wb.save(OUTPUT_DIR + '/scores.xlsx')\n"
            "print('saved')"
        )
        result = await executor.execute(code, "写Excel")
        assert "saved" in result.summary
        assert Path(ws["output"], "scores.xlsx").exists()

    @pytest.mark.asyncio
    async def test_staging_not_uploaded(self, executor, ws):
        """staging 目录的文件不被 auto_upload"""
        Path(ws["staging"], "intermediate.xlsx").write_bytes(b"staging data")

        result = await executor.execute("print('ok')", "触发upload扫描")
        assert "intermediate.xlsx" not in executor._uploaded


# ============================================================
# 3. 同名文件保护（dedup）
# ============================================================

class TestFileDedupEndToEnd:
    """Google Drive 风格同名文件保护 E2E"""

    @pytest.mark.asyncio
    async def test_overwrite_file_dedup(self, executor, ws):
        """已有文件被覆盖 → 旧文件保留，新文件重命名"""
        # 预先创建文件
        old_content = b"old report data"
        Path(ws["output"], "report.txt").write_bytes(old_content)

        # 沙盒代码覆盖同名文件
        code = (
            "with open(OUTPUT_DIR + '/report.txt', 'w') as f:\n"
            "    f.write('new report data')\n"
            "print('written')"
        )
        result = await executor.execute(code, "覆盖文件")
        assert "written" in result.summary

        # 旧文件恢复，新文件被重命名
        assert Path(ws["output"], "report.txt").read_bytes() == old_content
        assert Path(ws["output"], "report (1).txt").exists()

    @pytest.mark.asyncio
    async def test_no_overwrite_no_dedup(self, executor, ws):
        """文件未被覆盖 → 备份被清理，无多余文件"""
        Path(ws["output"], "keep.txt").write_bytes(b"untouched")

        result = await executor.execute("print('no file write')", "无写入")

        assert Path(ws["output"], "keep.txt").read_bytes() == b"untouched"
        assert not Path(ws["output"], "keep.txt.dedup_bak").exists()
        assert not Path(ws["output"], "keep (1).txt").exists()


# ============================================================
# 4. 安全拦截
# ============================================================

class TestSecurityEndToEnd:
    """安全拦截 E2E"""

    @pytest.mark.asyncio
    async def test_import_os_system_blocked(self, executor):
        """import os 放行，但 os.system 不存在（scoped_os 无此属性）"""
        result = await executor.execute("import os\nprint(hasattr(os, 'system'))")
        assert "False" in result.summary

    @pytest.mark.asyncio
    async def test_import_subprocess_blocked(self, executor):
        result = await executor.execute("import subprocess")
        assert "验证失败" in result.summary

    @pytest.mark.asyncio
    async def test_eval_blocked(self, executor):
        result = await executor.execute("eval('1+1')")
        assert "验证失败" in result.summary

    @pytest.mark.asyncio
    async def test_dunder_escape_blocked(self, executor):
        result = await executor.execute("[].__class__.__bases__")
        assert "验证失败" in result.summary

    @pytest.mark.asyncio
    async def test_open_outside_workspace(self, executor):
        """打开 workspace 外的文件被 _scoped_open 拦截"""
        result = await executor.execute("open('/etc/hosts').read()")
        assert "文件访问被拒绝" in result.summary or "PermissionError" in result.summary

    @pytest.mark.asyncio
    async def test_path_traversal_blocked(self, executor):
        """路径穿越（../）被 realpath + 边界检查拦截"""
        result = await executor.execute("open('../../etc/passwd').read()")
        # 可能被白名单拦截(PermissionError)，也可能 realpath 后文件不存在(FileNotFoundError)
        assert any(s in result.summary for s in (
            "文件访问被拒绝", "PermissionError", "文件不存在", "FileNotFoundError",
        ))

    @pytest.mark.asyncio
    async def test_symlink_escape_blocked(self, executor, ws):
        """符号链接指向 workspace 外被拦截"""
        link = Path(ws["workspace"], "escape_link")
        try:
            link.symlink_to("/etc")
        except OSError:
            pytest.skip("Cannot create symlink")

        result = await executor.execute("open('escape_link/passwd').read()")
        assert "文件访问被拒绝" in result.summary or "PermissionError" in result.summary

    @pytest.mark.asyncio
    async def test_empty_code_rejected(self, executor):
        result = await executor.execute("")
        assert "验证失败" in result.summary

    @pytest.mark.asyncio
    async def test_syntax_error_rejected(self, executor):
        result = await executor.execute("def foo(")
        assert "验证失败" in result.summary

    @pytest.mark.asyncio
    async def test_path_hidden_in_output(self, executor, ws):
        """真实路径在输出中被替换为虚拟路径"""
        result = await executor.execute("print(WORKSPACE_DIR)")
        assert str(ws["workspace"]) not in result.summary
        assert "WORKSPACE_DIR" in result.summary

    @pytest.mark.asyncio
    async def test_output_dir_hidden(self, executor, ws):
        """OUTPUT_DIR 路径在输出中被替换为变量名"""
        result = await executor.execute("print(OUTPUT_DIR)")
        assert str(ws["output"]) not in result.summary
        assert "OUTPUT_DIR" in result.summary


# ============================================================
# confirm_delete E2E（stateless subprocess）
# ============================================================

class TestConfirmDeleteEndToEnd:
    """删除操作确认流程 E2E"""

    @pytest.mark.asyncio
    async def test_remove_without_confirm_blocked(self, executor, ws):
        """os.remove 无 confirm_delete → PermissionError"""
        Path(ws["workspace"], "temp.txt").write_text("delete me")
        result = await executor.execute(
            "import os\nos.remove('temp.txt')", "删除文件",
        )
        assert "删除操作需要用户确认" in result.summary
        # 文件未被删除
        assert Path(ws["workspace"], "temp.txt").exists()

    @pytest.mark.asyncio
    async def test_remove_with_confirm_allowed(self, executor, ws):
        """os.remove + confirm_delete → 删除成功"""
        Path(ws["workspace"], "temp.txt").write_text("delete me")
        result = await executor.execute(
            "import os\nos.remove('temp.txt')\nprint('deleted')",
            "删除文件",
            confirm_delete=["temp.txt"],
        )
        assert "deleted" in result.summary
        assert not Path(ws["workspace"], "temp.txt").exists()

    @pytest.mark.asyncio
    async def test_confirm_delete_path_mismatch_blocked(self, executor, ws):
        """confirm_delete 传的文件名和代码里的不匹配 → 仍然拒绝"""
        Path(ws["workspace"], "important.xlsx").write_text("data")
        result = await executor.execute(
            "import os\nos.remove('important.xlsx')", "删除",
            confirm_delete=["other.xlsx"],  # 不匹配
        )
        assert "删除操作需要用户确认" in result.summary
        assert Path(ws["workspace"], "important.xlsx").exists()

    @pytest.mark.asyncio
    async def test_rmdir_always_blocked(self, executor, ws):
        """os.rmdir 始终拒绝"""
        Path(ws["workspace"], "empty_dir").mkdir()
        result = await executor.execute(
            "import os\nos.rmdir('empty_dir')", "删除目录",
        )
        assert "删除目录被禁止" in result.summary

    @pytest.mark.asyncio
    async def test_os_listdir_works(self, executor, ws):
        """os.listdir 正常工作（E2E 验证 scoped_os 注入）"""
        Path(ws["workspace"], "data.csv").write_text("a,b\n1,2")
        result = await executor.execute(
            "import os\nprint(os.listdir('.'))", "列目录",
        )
        assert "data.csv" in result.summary

    @pytest.mark.asyncio
    async def test_os_walk_relative_paths(self, executor, ws):
        """os.walk 返回相对路径（不泄露绝对路径）"""
        sub = Path(ws["workspace"], "sub")
        sub.mkdir()
        (sub / "file.txt").write_text("hi")
        result = await executor.execute(
            "import os\nfor r,d,f in os.walk('.'):\n  for fn in f:\n    print(os.path.join(r,fn))",
            "遍历",
        )
        # 不含绝对路径
        assert ws["workspace"] not in result.summary
        # 含相对路径
        assert "./sub/file.txt" in result.summary or "sub/file.txt" in result.summary


# ============================================================
# PandasProxy 截断提示 E2E
# ============================================================

class TestPandasTruncationHint:
    """PandasProxy 截断到 2000 行时自动输出提示"""

    @pytest.mark.asyncio
    async def test_truncation_hint_shown(self, executor, ws):
        """读取大 CSV 时输出截断提示"""
        # 创建 3000 行 CSV
        csv_path = Path(ws["workspace"], "big.csv")
        lines = ["id,value"] + [f"{i},{i*10}" for i in range(3000)]
        csv_path.write_text("\n".join(lines))

        result = await executor.execute(
            "df = pd.read_csv('big.csv')\nprint(f'rows={len(df)}')",
            "读大文件",
        )
        assert "rows=2000" in result.summary
        assert "截断" in result.summary or "2000" in result.summary

    @pytest.mark.asyncio
    async def test_no_hint_for_small_file(self, executor, ws):
        """读取小文件不输出截断提示"""
        csv_path = Path(ws["workspace"], "small.csv")
        lines = ["id,value"] + [f"{i},{i*10}" for i in range(100)]
        csv_path.write_text("\n".join(lines))

        result = await executor.execute(
            "df = pd.read_csv('small.csv')\nprint(f'rows={len(df)}')",
            "读小文件",
        )
        assert "rows=100" in result.summary
        assert "截断" not in result.summary

    @pytest.mark.asyncio
    async def test_nrows_none_bypasses_limit(self, executor, ws):
        """nrows=None 跳过截断限制"""
        csv_path = Path(ws["workspace"], "big.csv")
        lines = ["id,value"] + [f"{i},{i*10}" for i in range(3000)]
        csv_path.write_text("\n".join(lines))

        result = await executor.execute(
            "df = pd.read_csv('big.csv', nrows=None)\nprint(f'rows={len(df)}')",
            "全读",
        )
        assert "rows=3000" in result.summary

    @pytest.mark.asyncio
    async def test_excel_auto_header_detection(self, executor, ws):
        """ERP 导出 Excel 自动检测表头行（表头不在第一行）"""
        import pandas as _pd
        # 模拟 ERP 导出：第 0 行是标题，第 1 行是真正表头
        data = [
            ["利润表-店铺利润表", None, None],
            ["店铺", "收入", "利润"],
            ["旗舰店", 5000, 2000],
            ["京东店", 3000, 1500],
        ]
        df_raw = _pd.DataFrame(data)
        xlsx_path = Path(ws["workspace"], "erp_report.xlsx")
        df_raw.to_excel(str(xlsx_path), index=False, header=False)

        result = await executor.execute(
            "df = pd.read_excel('erp_report.xlsx')\nprint(df.columns.tolist())",
            "读ERP表",
        )
        # 应该自动检测到 header=1，列名是 [店铺, 收入, 利润]
        assert "店铺" in result.summary
        assert "Unnamed" not in result.summary

    @pytest.mark.asyncio
    async def test_csv_gbk_auto_encoding(self, executor, ws):
        """GBK 编码 CSV 自动检测"""
        csv_path = Path(ws["workspace"], "gbk.csv")
        csv_path.write_bytes("店铺,金额\n旗舰店,5000\n京东,3000\n".encode("gbk"))

        result = await executor.execute(
            "df = pd.read_csv('gbk.csv')\nprint(df.columns.tolist())",
            "读GBK",
        )
        assert "店铺" in result.summary

    @pytest.mark.asyncio
    async def test_csv_tsv_auto_delimiter(self, executor, ws):
        """Tab 分隔 CSV 自动检测分隔符"""
        tsv_path = Path(ws["workspace"], "data.csv")
        tsv_path.write_text("店铺\t金额\t利润\n旗舰店\t5000\t2000\n")

        result = await executor.execute(
            "df = pd.read_csv('data.csv')\nprint(len(df.columns))",
            "读TSV",
        )
        assert "3" in result.summary

    @pytest.mark.asyncio
    async def test_excel_unnamed_columns_cleaned(self, executor, ws):
        """合并单元格产生的空 Unnamed 列自动清理"""
        import pandas as _pd
        # 模拟带空列的 Excel
        data = [["店铺", None, "金额", None, "利润"],
                ["旗舰店", None, 5000, None, 2000]]
        df_raw = _pd.DataFrame(data)
        xlsx_path = Path(ws["workspace"], "merged.xlsx")
        df_raw.to_excel(str(xlsx_path), index=False, header=False)

        result = await executor.execute(
            "df = pd.read_excel('merged.xlsx')\nprint(df.columns.tolist())\nprint(len(df.columns))",
            "读合并表",
        )
        assert "Unnamed" not in result.summary


# ============================================================
# 5. 超时和进程管理
# ============================================================

class TestTimeoutAndProcessManagement:
    """超时和进程生命周期 E2E"""

    @pytest.mark.asyncio
    async def test_infinite_loop_timeout(self, ws):
        """死循环被超时机制杀死"""
        executor = SandboxExecutor(
            timeout=2.0,
            workspace_dir=ws["workspace"],
        )
        result = await executor.execute("while True: pass", "死循环")
        assert "超时" in result.summary

    @pytest.mark.asyncio
    async def test_long_calculation_within_timeout(self, executor):
        """正常计算在超时内完成"""
        code = "total = sum(range(1000000))\nprint(total)"
        result = await executor.execute(code, "大计算")
        assert "499999500000" in result.summary

    @pytest.mark.asyncio
    async def test_error_does_not_hang(self, executor):
        """运行时错误不会导致挂起"""
        start = time.monotonic()
        result = await executor.execute("1/0", "除零")
        elapsed = time.monotonic() - start
        assert elapsed < 10  # 应快速返回，不等到超时
        assert "执行错误" in result.summary

    @pytest.mark.asyncio
    async def test_result_truncation(self, ws):
        """超长输出被截断"""
        executor = SandboxExecutor(
            timeout=30.0,
            max_result_chars=200,
            workspace_dir=ws["workspace"],
        )
        result = await executor.execute("print('x' * 1000)", "超长输出")
        assert "已截断" in result.summary
        assert len(result.summary) < 500  # 截断后不应太长


# ============================================================
# 6. 并发隔离
# ============================================================

class TestConcurrencyIsolation:
    """多用户并发执行隔离 E2E"""

    @pytest.mark.asyncio
    async def test_two_users_concurrent(self, tmp_path):
        """两个用户同时执行，workspace 互不干扰"""
        ws_a = tmp_path / "user_a"
        ws_b = tmp_path / "user_b"
        ws_a.mkdir()
        ws_b.mkdir()

        # User A 的 workspace 有文件
        (ws_a / "data.txt").write_text("user_a_data")
        # User B 的 workspace 有文件
        (ws_b / "data.txt").write_text("user_b_data")

        executor_a = SandboxExecutor(timeout=30.0, workspace_dir=str(ws_a))
        executor_b = SandboxExecutor(timeout=30.0, workspace_dir=str(ws_b))

        results = await asyncio.gather(
            executor_a.execute("print(open('data.txt').read())", "User A"),
            executor_b.execute("print(open('data.txt').read())", "User B"),
        )

        assert "user_a_data" in results[0].summary
        assert "user_b_data" in results[1].summary
        # 互不交叉
        assert "user_b_data" not in results[0].summary
        assert "user_a_data" not in results[1].summary

    @pytest.mark.asyncio
    async def test_globals_not_shared(self, tmp_path):
        """两次执行的全局变量不共享（独立进程）"""
        executor = SandboxExecutor(timeout=30.0, workspace_dir=str(tmp_path))

        await executor.execute("shared_var = 'leaked'", "设置变量")
        result = await executor.execute("print(shared_var)", "读取变量")

        assert "执行错误" in result.summary or "NameError" in result.summary


# ============================================================
# 7. 边缘情况
# ============================================================

class TestEdgeCases:
    """边缘情况 E2E"""

    @pytest.mark.asyncio
    async def test_only_comments(self, executor):
        """只有注释的代码"""
        result = await executor.execute("# this is a comment", "纯注释")
        assert "成功" in result.summary

    @pytest.mark.asyncio
    async def test_multiline_string(self, executor):
        """多行字符串"""
        code = "text = '''line1\nline2\nline3'''\nprint(len(text.split('\\n')))"
        result = await executor.execute(code, "多行字符串")
        assert "3" in result.summary

    @pytest.mark.asyncio
    async def test_nested_function_definition(self, executor):
        """嵌套函数定义"""
        code = (
            "def outer(x):\n"
            "    def inner(y):\n"
            "        return x + y\n"
            "    return inner(10)\n"
            "print(outer(5))"
        )
        result = await executor.execute(code, "嵌套函数")
        assert "15" in result.summary

    @pytest.mark.asyncio
    async def test_list_comprehension(self, executor):
        """列表推导式"""
        result = await executor.execute(
            "str([x**2 for x in range(5)])", "列表推导",
        )
        assert "[0, 1, 4, 9, 16]" in result.summary

    @pytest.mark.asyncio
    async def test_exception_with_chinese_message(self, executor):
        """中文异常消息"""
        result = await executor.execute(
            "raise ValueError('数据格式错误')", "中文异常",
        )
        assert "执行错误" in result.summary

    @pytest.mark.asyncio
    async def test_file_not_found(self, executor):
        """读不存在的文件"""
        result = await executor.execute(
            "open('nonexistent.txt').read()", "文件不存在",
        )
        assert "执行错误" in result.summary

    @pytest.mark.asyncio
    async def test_workspace_dir_not_in_result(self, executor, ws):
        """确保所有输出中不泄露真实路径"""
        code = (
            "print(WORKSPACE_DIR)\n"
            "print(STAGING_DIR)\n"
            "print(OUTPUT_DIR)"
        )
        result = await executor.execute(code, "路径检查")
        assert ws["workspace"] not in result.summary
        assert ws["staging"] not in result.summary
        assert ws["output"] not in result.summary

    @pytest.mark.asyncio
    async def test_async_await_rejected(self, executor):
        """子进程不支持 async/await"""
        result = await executor.execute("await some_func()", "async代码")
        assert "不支持" in result.summary or "验证失败" in result.summary

    @pytest.mark.asyncio
    async def test_import_allowed_module_in_code(self, executor):
        """用户代码中 import 白名单模块"""
        code = "import re\nprint(bool(re.match(r'\\d+', '123')))"
        result = await executor.execute(code, "import re")
        assert "True" in result.summary

    @pytest.mark.asyncio
    async def test_import_blocked_module_in_code(self, executor):
        """用户代码中 import 黑名单模块"""
        result = await executor.execute("import socket", "import blocked")
        assert "验证失败" in result.summary
