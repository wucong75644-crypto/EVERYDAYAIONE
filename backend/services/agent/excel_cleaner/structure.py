"""Excel 结构检测 (Layer 1)。

从 xlsx ZIP 内 XML 提取：
  - merged_ranges（合并单元格）
  - hidden_rows / hidden_cols（隐藏行列）
  - has_auto_filter

含列字母 ↔ 索引工具（_col_letter_to_index / _col_index_to_letter_local）和
xlsx ZIP 路径解析（_parse_sheet_tags / _resolve_sheet_xml_path）。

XML > 500MB 时降级到 openpyxl read_only 流式读 mergedCells（Bug-9 修复）。
"""
from __future__ import annotations

import re
from pathlib import Path
from zipfile import BadZipFile, ZipFile

from loguru import logger

from services.agent.excel_cleaner.report import ExcelStructure

# ── 安全阀 ──
_MAX_XML_SIZE = 500 * 1024 * 1024  # 500MB: 解压后 XML 超此大小走 fallback


# ── 正则 ──

_RE_MERGE = re.compile(r'<mergeCell\s+ref="([A-Z]+)(\d+):([A-Z]+)(\d+)"')
_RE_HIDDEN_ROW = re.compile(
    r'<row\s[^>]*?r="(\d+)"[^>]*?hidden="1"'
    r'|<row\s[^>]*?hidden="1"[^>]*?r="(\d+)"'
)
_RE_HIDDEN_COL = re.compile(
    r'<col\s[^>]*?min="(\d+)"[^>]*?max="(\d+)"[^>]*?hidden="1"'
    r'|<col\s[^>]*?hidden="1"[^>]*?min="(\d+)"[^>]*?max="(\d+)"'
)
_RE_AUTO_FILTER = re.compile(r'<autoFilter\b')
_RE_REL = re.compile(
    r'<Relationship\s[^>]*?Id="(rId\d+)"[^>]*?Target="([^"]*worksheet[^"]*)"'
    r'|<Relationship\s[^>]*?Target="([^"]*worksheet[^"]*)"[^>]*?Id="(rId\d+)"'
)


# ── 列字母工具 ──

def _col_letter_to_index(col_str: str) -> int:
    """Excel 列字母 → 1-indexed 数字（A=1, B=2, ..., Z=26, AA=27）。"""
    result = 0
    for ch in col_str.upper():
        result = result * 26 + (ord(ch) - ord("A") + 1)
    return result


def _col_index_to_letter_local(idx: int) -> str:
    """0-indexed 列索引 → Excel 列字母（0=A, 1=B, ..., 25=Z, 26=AA）。

    与 file_meta.dataclass._col_index_to_letter 实现一致，
    保留本地副本是因 file_meta 包 init 链经 formulas 模块回引 excel_cleaner，
    跨包 import 会触发循环依赖。
    """
    result = ""
    n = idx
    while True:
        result = chr(ord("A") + n % 26) + result
        n = n // 26 - 1
        if n < 0:
            break
    return result


# ── ZIP / XML 路径解析 ──

def _parse_sheet_tags(xml: str) -> list[tuple[str, str, str]]:
    """从 workbook.xml 提取 <sheet> 标签，不依赖属性顺序。

    Returns: [(name, sheetId, rId), ...]
    """
    results = []
    for m in re.finditer(r'<sheet\s([^>]*?)/?>', xml):
        attrs = m.group(1)
        name_m = re.search(r'name="([^"]*)"', attrs)
        sid_m = re.search(r'sheetId="(\d+)"', attrs)
        rid_m = re.search(r'r:id="(rId\d+)"', attrs)
        if name_m and rid_m:
            results.append((
                name_m.group(1),
                sid_m.group(1) if sid_m else "0",
                rid_m.group(1),
            ))
    return results


def _resolve_sheet_xml_path(
    zf: ZipFile, sheet_name: str | int,
) -> str | None:
    """解析 xlsx ZIP 内 sheet 名/索引 → XML 路径。"""
    try:
        wb_xml = zf.read("xl/workbook.xml").decode("utf-8", errors="replace")
    except KeyError:
        return None

    sheets = _parse_sheet_tags(wb_xml)
    if not sheets:
        return None

    # 确定目标 rId
    if isinstance(sheet_name, int):
        if sheet_name >= len(sheets):
            return None
        target_rid = sheets[sheet_name][2]
    else:
        target_rid = None
        name_lower = str(sheet_name).lower().strip()
        for name, _, rid in sheets:
            if name.lower().strip() == name_lower:
                target_rid = rid
                break
        if target_rid is None:
            return None

    # rId → XML 路径
    try:
        rels_xml = zf.read("xl/_rels/workbook.xml.rels").decode(
            "utf-8", errors="replace"
        )
    except KeyError:
        return None

    for m in _RE_REL.finditer(rels_xml):
        rid = m.group(1) or m.group(4)
        target = m.group(2) or m.group(3)
        if rid == target_rid:
            # Target 可能是 "/xl/worksheets/sheet1.xml" 或 "worksheets/sheet1.xml"
            target = target.lstrip("/")
            if not target.startswith("xl/"):
                target = f"xl/{target}"
            return target
    return None


# ── 主入口 ──

def _detect_structure(
    excel_path: str, sheet_name: str | int,
) -> ExcelStructure | None:
    """Layer 1: 从 xlsx ZIP 内 XML 提取结构元数据。失败时返回 None（降级）。

    Bug-9 修复：XML > _MAX_XML_SIZE 时不再直接返回 None 丢失全部信息，
    改为调 openpyxl read_only 流式解析至少拿到 mergedCells。
    """
    if not excel_path.lower().endswith((".xlsx", ".xlsm")):
        return None  # .xls 旧格式不是 ZIP，无法解析

    try:
        with ZipFile(excel_path, "r") as zf:
            xml_path = _resolve_sheet_xml_path(zf, sheet_name)
            if xml_path is None:
                return None

            # 安全阀：超大 XML 走 openpyxl 流式 fallback
            for info in zf.infolist():
                if info.filename == xml_path:
                    if info.file_size > _MAX_XML_SIZE:
                        logger.warning(
                            f"Excel XML too large ({info.file_size:,} bytes), "
                            f"fallback to openpyxl streaming | {Path(excel_path).name}"
                        )
                        return _detect_structure_streaming(excel_path, sheet_name)
                    break

            raw = zf.read(xml_path).decode("utf-8", errors="replace")
    except (BadZipFile, KeyError, OSError) as e:
        logger.debug(f"Excel structure detection failed: {e}")
        return None

    structure = ExcelStructure()

    # 合并区域
    for m in _RE_MERGE.finditer(raw):
        min_col = _col_letter_to_index(m.group(1))
        min_row = int(m.group(2))
        max_col = _col_letter_to_index(m.group(3))
        max_row = int(m.group(4))
        structure.merged_ranges.append((min_row, max_row, min_col, max_col))

    # 隐藏行
    for m in _RE_HIDDEN_ROW.finditer(raw):
        row_num = int(m.group(1) or m.group(2))
        structure.hidden_rows.add(row_num)

    # 隐藏列
    for m in _RE_HIDDEN_COL.finditer(raw):
        min_c = int(m.group(1) or m.group(3))
        max_c = int(m.group(2) or m.group(4))
        for c in range(min_c, max_c + 1):
            structure.hidden_cols.add(c)

    # 自动筛选
    structure.has_auto_filter = bool(_RE_AUTO_FILTER.search(raw))

    del raw  # 释放 XML 字符串内存
    return structure


def _detect_structure_streaming(
    excel_path: str, sheet_name: str | int,
) -> ExcelStructure | None:
    """Bug-9 fallback：超大 XML 用 openpyxl read_only 流式读 mergedCells。

    无法拿到隐藏行/列和 autofilter（read_only 模式不支持），
    但至少保证大文件场景下合并单元格信息不丢失。
    """
    try:
        import openpyxl
    except ImportError:
        logger.warning("openpyxl not installed, skip streaming fallback")
        return None

    try:
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=False)
    except Exception as e:
        logger.debug(f"openpyxl streaming load failed: {e}")
        return None

    try:
        if isinstance(sheet_name, int):
            sheet_names_list = wb.sheetnames
            if sheet_name >= len(sheet_names_list):
                return None
            ws = wb[sheet_names_list[sheet_name]]
        else:
            if sheet_name not in wb.sheetnames:
                return None
            ws = wb[sheet_name]

        structure = ExcelStructure()
        # read_only 模式下 merged_cells 通常可用（依赖 openpyxl 版本）
        merged = getattr(ws, "merged_cells", None)
        if merged is not None:
            for rng in list(merged.ranges):
                structure.merged_ranges.append(
                    (rng.min_row, rng.max_row, rng.min_col, rng.max_col)
                )
        return structure
    finally:
        wb.close()
