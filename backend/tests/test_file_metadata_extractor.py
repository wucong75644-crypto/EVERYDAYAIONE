"""
文件元信息提取 单元测试

测试 extract_spreadsheet_metadata / extract_metadata_for_files / format_workspace_files_prompt
所有测试在临时目录中运行，不依赖外部服务。
"""

import asyncio
import csv
import os
from pathlib import Path

import pytest

from services.file_metadata_extractor import (
    _decode_bytes,
    _format_cell,
    _fmt_size,
    _infer_column_meta,
    _is_date,
    _is_numeric,
    _to_float,
    extract_metadata_for_files,
    extract_spreadsheet_metadata,
    format_workspace_files_prompt,
)


# ============================================================
# Fixtures
# ============================================================


@pytest.fixture
def workspace(tmp_path):
    """临时 workspace 目录"""
    return str(tmp_path)


@pytest.fixture
def sample_xlsx(workspace):
    """创建测试用 xlsx 文件"""
    from openpyxl import Workbook

    path = os.path.join(workspace, "利润表.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # 表头
    headers = ["店铺名称", "日期", "销售额", "利润", "平台", "状态"]
    ws.append(headers)

    # 数据行
    data = [
        ["店铺A", "2026-04-20", 15000, 3200, "淘宝", "已完成"],
        ["店铺B", "2026-04-20", 22000, 4100, "拼多多", "已完成"],
        ["店铺C", "2026-04-21", 8500, 1800, "抖音", "退款中"],
        ["店铺A", "2026-04-21", 12000, 2500, "淘宝", "已完成"],
        ["店铺D", "2026-04-22", 5000, -200, "淘宝", "已取消"],
    ]
    for row in data:
        ws.append(row)

    wb.save(path)
    return path


@pytest.fixture
def sample_csv_utf8(workspace):
    """创建 UTF-8 编码的 CSV"""
    path = os.path.join(workspace, "数据.csv")
    with open(path, "w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["商品名", "价格", "数量", "类别"])
        writer.writerow(["苹果", "5.5", "100", "水果"])
        writer.writerow(["香蕉", "3.2", "200", "水果"])
        writer.writerow(["牛奶", "12.0", "50", "饮品"])
        writer.writerow(["面包", "8.0", "80", "烘焙"])
    return path


@pytest.fixture
def sample_csv_gbk(workspace):
    """创建 GBK 编码的 CSV"""
    path = os.path.join(workspace, "gbk数据.csv")
    with open(path, "w", encoding="gbk", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["商品", "金额"])
        writer.writerow(["苹果", "100"])
        writer.writerow(["香蕉", "200"])
    return path


@pytest.fixture
def sample_tsv(workspace):
    """创建 TSV 文件"""
    path = os.path.join(workspace, "data.tsv")
    with open(path, "w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f, delimiter="\t")
        writer.writerow(["name", "value"])
        writer.writerow(["alpha", "10"])
        writer.writerow(["beta", "20"])
    return path


@pytest.fixture
def empty_xlsx(workspace):
    """创建空 xlsx（只有表头）"""
    from openpyxl import Workbook

    path = os.path.join(workspace, "空表.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.append(["列A", "列B"])
    wb.save(path)
    return path


@pytest.fixture
def large_xlsx(workspace):
    """创建较大的 xlsx（>500 行）"""
    from openpyxl import Workbook

    path = os.path.join(workspace, "大表.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.append(["ID", "值", "类别"])
    for i in range(600):
        ws.append([i + 1, (i + 1) * 10.5, f"类别{i % 3}"])
    wb.save(path)
    return path


# ============================================================
# 工具函数测试
# ============================================================


class TestHelpers:
    def test_is_numeric(self):
        assert _is_numeric(42) is True
        assert _is_numeric(3.14) is True
        assert _is_numeric("100") is True
        assert _is_numeric("1,234") is True
        assert _is_numeric("-5.5") is True
        assert _is_numeric("abc") is False
        assert _is_numeric("") is False
        assert _is_numeric(None) is False

    def test_to_float(self):
        assert _to_float(42) == 42.0
        assert _to_float("1,234") == 1234.0
        assert _to_float("abc") is None
        assert _to_float(None) is None

    def test_is_date(self):
        assert _is_date("2026-04-20") is True
        assert _is_date("2026/04/20") is True
        assert _is_date("2026-04-20 14:30:00") is True
        assert _is_date("abc") is False
        assert _is_date("12") is False
        assert _is_date(None) is False

    def test_is_date_with_datetime_object(self):
        from datetime import datetime, date
        assert _is_date(datetime(2026, 4, 20)) is True
        assert _is_date(date(2026, 4, 20)) is True

    def test_format_cell(self):
        assert _format_cell(None) == ""
        assert _format_cell(42) == "42"
        assert _format_cell("hello") == "hello"
        assert _format_cell("a" * 50) == "a" * 27 + "..."


class TestInferColumnMeta:
    def test_numeric_column(self):
        values = [100, 200, 300, None, 500]
        meta = _infer_column_meta("金额", values, total_rows=1000)
        assert meta["name"] == "金额"
        assert meta["dtype"] == "数值"
        assert meta["min"] == 100.0
        assert meta["max"] == 500.0
        assert meta["non_null"] == 800  # 4/5 * 1000

    def test_date_column(self):
        values = ["2026-04-20", "2026-04-21", "2026-04-22"]
        meta = _infer_column_meta("日期", values, total_rows=100)
        assert meta["dtype"] == "日期"

    def test_bool_column(self):
        values = ["是", "否", "是", "是"]
        meta = _infer_column_meta("有效", values, total_rows=100)
        assert meta["dtype"] == "布尔"

    def test_category_column(self):
        values = ["淘宝"] * 10 + ["拼多多"] * 5 + ["抖音"] * 3
        meta = _infer_column_meta("平台", values, total_rows=1000)
        assert meta["dtype"] == "文本"
        assert "categories" in meta
        assert meta["_unique_count"] == 3
        # 第一个分类应该是出现最多的
        assert meta["categories"][0][0] == "淘宝"

    def test_text_column_many_unique(self):
        values = [f"item_{i}" for i in range(30)]
        meta = _infer_column_meta("名称", values, total_rows=100)
        assert meta["dtype"] == "文本"
        assert "categories" not in meta  # unique > 20 不视为分类

    def test_empty_column(self):
        values = [None, None, ""]
        meta = _infer_column_meta("空列", values, total_rows=100)
        assert meta["dtype"] == "文本"
        assert meta["non_null"] == 0
        assert meta["sample"] == []

    def test_sample_values(self):
        values = ["A", "B", "A", "C"]
        meta = _infer_column_meta("名称", values, total_rows=100)
        assert len(meta["sample"]) == 2
        assert meta["sample"][0] == "A"
        assert meta["sample"][1] == "B"

    def test_string_numeric_column(self):
        """字符串形式的数字应被识别为数值"""
        values = ["100", "200", "300.5", "1,234"]
        meta = _infer_column_meta("金额", values, total_rows=100)
        assert meta["dtype"] == "数值"


# ============================================================
# xlsx 提取测试
# ============================================================


class TestExtractXlsx:
    def test_basic_extraction(self, sample_xlsx):
        meta = extract_spreadsheet_metadata(sample_xlsx)
        assert meta is not None
        assert meta["row_count"] == 5  # 5 数据行
        assert meta["col_count"] == 6
        assert len(meta["columns"]) == 6
        assert meta["columns"][0]["name"] == "店铺名称"

    def test_column_types(self, sample_xlsx):
        meta = extract_spreadsheet_metadata(sample_xlsx)
        cols = {c["name"]: c for c in meta["columns"]}
        assert cols["销售额"]["dtype"] == "数值"
        assert cols["日期"]["dtype"] == "日期"
        assert cols["店铺名称"]["dtype"] == "文本"

    def test_numeric_range(self, sample_xlsx):
        meta = extract_spreadsheet_metadata(sample_xlsx)
        cols = {c["name"]: c for c in meta["columns"]}
        assert cols["销售额"]["min"] == 5000.0
        assert cols["销售额"]["max"] == 22000.0
        assert cols["利润"]["min"] == -200.0

    def test_category_detection(self, sample_xlsx):
        meta = extract_spreadsheet_metadata(sample_xlsx)
        cols = {c["name"]: c for c in meta["columns"]}
        assert "categories" in cols["平台"]
        assert "categories" in cols["状态"]

    def test_preview_rows(self, sample_xlsx):
        meta = extract_spreadsheet_metadata(sample_xlsx)
        assert len(meta["preview_rows"]) == 3  # 前 3 行
        assert meta["preview_rows"][0][0] == "店铺A"

    def test_empty_xlsx(self, empty_xlsx):
        meta = extract_spreadsheet_metadata(empty_xlsx)
        # 只有表头没有数据行
        assert meta is not None
        assert meta["row_count"] == 0 or meta["row_count"] == 1
        assert len(meta["columns"]) == 2

    def test_large_xlsx(self, large_xlsx):
        meta = extract_spreadsheet_metadata(large_xlsx)
        assert meta is not None
        assert meta["row_count"] == 600
        assert meta["col_count"] == 3
        # 类别列应检测到 3 个类别
        cols = {c["name"]: c for c in meta["columns"]}
        assert cols["类别"]["_unique_count"] == 3

    def test_nonexistent_file(self):
        result = extract_spreadsheet_metadata("/nonexistent/path.xlsx")
        assert result is None

    def test_corrupted_file(self, workspace):
        path = os.path.join(workspace, "bad.xlsx")
        with open(path, "wb") as f:
            f.write(b"this is not an xlsx file")
        result = extract_spreadsheet_metadata(path)
        assert result is None


# ============================================================
# CSV 提取测试
# ============================================================


class TestExtractCsv:
    def test_utf8_csv(self, sample_csv_utf8):
        meta = extract_spreadsheet_metadata(sample_csv_utf8)
        assert meta is not None
        assert meta["row_count"] == 4
        assert meta["col_count"] == 4
        assert meta["columns"][0]["name"] == "商品名"

    def test_gbk_csv(self, sample_csv_gbk):
        meta = extract_spreadsheet_metadata(sample_csv_gbk)
        assert meta is not None
        assert meta["row_count"] == 2
        assert meta["columns"][0]["name"] == "商品"

    def test_tsv(self, sample_tsv):
        meta = extract_spreadsheet_metadata(sample_tsv)
        assert meta is not None
        assert meta["col_count"] == 2
        assert meta["columns"][0]["name"] == "name"

    def test_csv_type_inference(self, sample_csv_utf8):
        meta = extract_spreadsheet_metadata(sample_csv_utf8)
        cols = {c["name"]: c for c in meta["columns"]}
        assert cols["价格"]["dtype"] == "数值"
        assert cols["数量"]["dtype"] == "数值"
        assert cols["商品名"]["dtype"] == "文本"

    def test_csv_categories(self, sample_csv_utf8):
        meta = extract_spreadsheet_metadata(sample_csv_utf8)
        cols = {c["name"]: c for c in meta["columns"]}
        assert "categories" in cols["类别"]

    def test_empty_csv(self, workspace):
        path = os.path.join(workspace, "empty.csv")
        with open(path, "w") as f:
            f.write("")
        result = extract_spreadsheet_metadata(path)
        assert result is None


# ============================================================
# 异步批量提取测试
# ============================================================


class TestExtractMetadataForFiles:
    @pytest.mark.asyncio
    async def test_single_file(self, workspace, sample_xlsx):
        files = [{"workspace_path": "利润表.xlsx", "name": "利润表.xlsx", "size": 1000}]
        result = await extract_metadata_for_files(files, workspace)
        assert "利润表.xlsx" in result
        assert result["利润表.xlsx"] is not None
        assert result["利润表.xlsx"]["row_count"] == 5

    @pytest.mark.asyncio
    async def test_multiple_files(self, workspace, sample_xlsx, sample_csv_utf8):
        files = [
            {"workspace_path": "利润表.xlsx", "name": "利润表.xlsx", "size": 1000},
            {"workspace_path": "数据.csv", "name": "数据.csv", "size": 500},
        ]
        result = await extract_metadata_for_files(files, workspace)
        assert len(result) == 2
        assert result["利润表.xlsx"] is not None
        assert result["数据.csv"] is not None

    @pytest.mark.asyncio
    async def test_nonexistent_file(self, workspace):
        files = [{"workspace_path": "不存在.xlsx", "name": "不存在.xlsx", "size": 0}]
        result = await extract_metadata_for_files(files, workspace)
        assert result.get("不存在.xlsx") is None

    @pytest.mark.asyncio
    async def test_text_file_also_extracted(self, workspace):
        """所有文件类型都提取元数据（txt 返回行数/字数）"""
        (Path(workspace) / "readme.txt").write_text("line1\nline2\nline3")
        files = [{"workspace_path": "readme.txt", "name": "readme.txt", "size": 100}]
        result = await extract_metadata_for_files(files, workspace)
        assert "readme.txt" in result
        meta = result["readme.txt"]
        assert meta is not None
        assert meta["type"] == "text"
        assert meta["lines"] == 3

    @pytest.mark.asyncio
    async def test_empty_list(self, workspace):
        result = await extract_metadata_for_files([], workspace)
        assert result == {}

    @pytest.mark.asyncio
    async def test_max_files_limit(self, workspace):
        """超过 5 个文件时只提取前 5 个"""
        files = [
            {"workspace_path": f"file_{i}.xlsx", "name": f"file_{i}.xlsx", "size": 100}
            for i in range(8)
        ]
        result = await extract_metadata_for_files(files, workspace)
        # 文件不存在所以都是 None，但应该只尝试了 5 个
        assert len(result) <= 5


# ============================================================
# 格式化输出测试
# ============================================================


class TestFormatPrompt:
    def _make_meta(self, row_count=100, col_count=3):
        return {
            "row_count": row_count,
            "col_count": col_count,
            "columns": [
                {"name": "名称", "dtype": "文本", "non_null": row_count,
                 "sample": ["A", "B"]},
                {"name": "金额", "dtype": "数值", "non_null": row_count,
                 "sample": ["100", "200"], "min": 50.0, "max": 5000.0},
                {"name": "状态", "dtype": "文本", "non_null": row_count,
                 "sample": ["完成", "取消"],
                 "categories": [("完成", 70), ("取消", 30)], "_unique_count": 2},
            ],
            "preview_rows": [["A", "100", "完成"], ["B", "200", "取消"]],
        }

    def test_standard_format(self):
        """≤2 个表格文件使用标准档"""
        files = [{"workspace_path": "data.xlsx", "name": "data.xlsx", "size": 1024}]
        meta_map = {"data.xlsx": self._make_meta()}
        result = format_workspace_files_prompt(files, meta_map)

        assert "📊 data.xlsx" in result
        assert "100 行 × 3 列" in result
        assert "路径: data.xlsx" in result
        assert "| 列名 |" in result
        assert "| 名称 |" in result
        assert "pd.read_excel('data.xlsx')" in result

    def test_compact_format(self):
        """3+ 个表格文件使用紧凑档"""
        files = [
            {"workspace_path": f"file_{i}.xlsx", "name": f"file_{i}.xlsx", "size": 1024}
            for i in range(3)
        ]
        meta_map = {f"file_{i}.xlsx": self._make_meta() for i in range(3)}
        result = format_workspace_files_prompt(files, meta_map)

        # 紧凑档没有表格
        assert "| 列名 |" not in result
        assert "列:" in result

    def test_failed_extraction_fallback(self):
        """提取失败时降级"""
        files = [{"workspace_path": "bad.xlsx", "name": "bad.xlsx", "size": 2048}]
        meta_map = {"bad.xlsx": None}
        result = format_workspace_files_prompt(files, meta_map)

        assert "📊 bad.xlsx" in result
        assert "pd.read_excel('bad.xlsx')" in result

    def test_text_file_format(self):
        """非表格文件显示 file_read 指引"""
        files = [{"workspace_path": "readme.txt", "name": "readme.txt",
                  "size": 512, "mime_type": "text/plain"}]
        result = format_workspace_files_prompt(files, {})

        assert "📄 readme.txt" in result
        assert "file_read" in result

    def test_mixed_files(self):
        """混合文件类型"""
        files = [
            {"workspace_path": "data.xlsx", "name": "data.xlsx", "size": 1024},
            {"workspace_path": "notes.txt", "name": "notes.txt",
             "size": 256, "mime_type": "text/plain"},
        ]
        meta_map = {"data.xlsx": self._make_meta()}
        result = format_workspace_files_prompt(files, meta_map)

        assert "📊 data.xlsx" in result
        assert "📄 notes.txt" in result

    def test_csv_read_command(self):
        """CSV 文件使用 pd.read_csv"""
        files = [{"workspace_path": "data.csv", "name": "data.csv", "size": 512}]
        meta_map = {"data.csv": self._make_meta()}
        result = format_workspace_files_prompt(files, meta_map)

        assert "pd.read_csv('data.csv')" in result

    def test_empty_files(self):
        """空文件列表"""
        result = format_workspace_files_prompt([], {})
        assert result == ""

    def test_numeric_range_in_standard(self):
        """标准档中数值列显示范围"""
        files = [{"workspace_path": "data.xlsx", "name": "data.xlsx", "size": 1024}]
        meta_map = {"data.xlsx": self._make_meta()}
        result = format_workspace_files_prompt(files, meta_map)

        assert "范围:" in result
        assert "50" in result
        assert "5,000" in result

    def test_category_in_standard(self):
        """标准档中分类列显示枚举"""
        files = [{"workspace_path": "data.xlsx", "name": "data.xlsx", "size": 1024}]
        meta_map = {"data.xlsx": self._make_meta()}
        result = format_workspace_files_prompt(files, meta_map)

        assert '"完成"' in result
        assert '"取消"' in result

    def test_xls_fallback(self):
        """.xls 文件提取失败应降级显示"""
        files = [{"workspace_path": "old.xls", "name": "old.xls", "size": 4096}]
        meta_map = {"old.xls": None}
        result = format_workspace_files_prompt(files, meta_map)

        assert "📊 old.xls" in result
        assert "pd.read_excel('old.xls')" in result

    def test_over_five_files_truncation(self):
        """>5 个表格文件时只展示前 5 个 + 提示"""
        files = [
            {"workspace_path": f"sheet_{i}.xlsx", "name": f"sheet_{i}.xlsx", "size": 1024}
            for i in range(7)
        ]
        meta_map = {f"sheet_{i}.xlsx": self._make_meta() for i in range(7)}
        result = format_workspace_files_prompt(files, meta_map)

        assert "另有 2 个表格文件" in result
        # 前 5 个应该存在
        assert "sheet_0.xlsx" in result
        assert "sheet_4.xlsx" in result

    def test_binary_file_code_execute(self):
        """非表格二进制文件（如 PDF）应指引 code_execute"""
        files = [{"workspace_path": "report.pdf", "name": "report.pdf",
                  "size": 2048, "mime_type": "application/pdf"}]
        result = format_workspace_files_prompt(files, {})

        assert "📄 report.pdf" in result
        assert "code_execute" in result


# ============================================================
# _fmt_size 测试
# ============================================================


class TestFmtSize:
    def test_zero(self):
        assert _fmt_size(0) == "未知大小"

    def test_none(self):
        assert _fmt_size(None) == "未知大小"

    def test_bytes(self):
        assert _fmt_size(512) == "512B"

    def test_kilobytes(self):
        result = _fmt_size(5120)
        assert result == "5.0KB"

    def test_megabytes(self):
        result = _fmt_size(2 * 1024 * 1024)
        assert result == "2.0MB"

    def test_fractional_kb(self):
        result = _fmt_size(1536)  # 1.5KB
        assert result == "1.5KB"


# ============================================================
# _decode_bytes 测试
# ============================================================


class TestDecodeBytes:
    def test_utf8(self):
        text = "你好世界"
        result = _decode_bytes(text.encode("utf-8"))
        assert result == text

    def test_gbk(self):
        text = "你好世界"
        result = _decode_bytes(text.encode("gbk"))
        assert result == text

    def test_latin1(self):
        data = bytes(range(128, 256))
        result = _decode_bytes(data)
        assert isinstance(result, str)
        assert len(result) == 128

    def test_invalid_bytes_fallback(self):
        """无法解码的字节应 fallback 为 replace 模式"""
        # 构造一个 utf-8/gbk/latin-1 都能解码的字节序列（latin-1 兜底所有 0-255）
        # 所以测试 replace 模式需要更极端的情况
        # 实际上 latin-1 能解码任何单字节，所以 _decode_bytes 不会走到 replace 分支
        # 直接测试 latin-1 成功路径
        data = b"\x80\x81\x82"
        result = _decode_bytes(data)
        assert isinstance(result, str)


# ============================================================
# .xls 降级测试
# ============================================================


class TestXlsDegradation:
    def test_xls_returns_none(self, workspace):
        """.xls 文件应返回 None（openpyxl 不支持）"""
        path = os.path.join(workspace, "old_format.xls")
        with open(path, "wb") as f:
            # 写入 BIFF 文件头标识（真实 .xls 格式）
            f.write(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1" + b"\x00" * 100)
        result = extract_spreadsheet_metadata(path)
        assert result is None


# ============================================================
# CSV 单次流式读取测试（review 修复后的新逻辑）
# ============================================================


class TestCsvStreamRead:
    def test_small_csv_row_count_accurate(self, workspace):
        """小 CSV 单次读取后行数应精确"""
        path = os.path.join(workspace, "small.csv")
        with open(path, "w", encoding="utf-8", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["id", "value"])
            for i in range(50):
                writer.writerow([i, i * 10])

        meta = extract_spreadsheet_metadata(path)
        assert meta is not None
        assert meta["row_count"] == 50  # 精确计数

    def test_csv_with_fewer_rows_than_sample(self, workspace):
        """行数少于采样数时应精确计数"""
        path = os.path.join(workspace, "tiny.csv")
        with open(path, "w", encoding="utf-8", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["a", "b"])
            writer.writerow(["1", "2"])

        meta = extract_spreadsheet_metadata(path)
        assert meta is not None
        assert meta["row_count"] == 1

    def test_csv_wide_table(self, workspace):
        """宽表（多列）应正确提取"""
        path = os.path.join(workspace, "wide.csv")
        with open(path, "w", encoding="utf-8", newline="") as f:
            writer = csv.writer(f)
            headers = [f"col_{i}" for i in range(50)]
            writer.writerow(headers)
            for _ in range(10):
                writer.writerow([f"val_{i}" for i in range(50)])

        meta = extract_spreadsheet_metadata(path)
        assert meta is not None
        assert meta["col_count"] == 50
        assert meta["row_count"] == 10


# ============================================================
# 宽表模式识别
# ============================================================


class TestWideTablePattern:
    """_detect_wide_table_pattern + _format_wide_table_fallback 测试"""

    def test_pattern_detected_store_x_month(self):
        """典型宽表：店铺名_日期 → 识别模式（需 >50 列才触发）"""
        from services.file_metadata_extractor import _detect_wide_table_pattern

        stores = [f"店铺{chr(65+i)}" for i in range(5)]  # 5 店铺
        months = [f"2024-{m:02d}" for m in range(1, 13)]  # 12 月
        # 索引列 + 60 个值列（5 × 12）= 61 列，超过 50 阈值
        columns = [{"name": "科目", "dtype": "文本"}]
        for s in stores:
            for m in months:
                columns.append({"name": f"{s}_{m}", "dtype": "数值", "min": 0, "max": 100000})

        result = _detect_wide_table_pattern(columns)
        assert result is not None
        desc = result["description"]
        assert "宽表" in desc
        assert "5 个前缀" in desc
        assert "12 个日期" in desc  # 后缀是日期
        assert "店铺A" in desc
        assert result["value_col_count"] == 60
        assert len(result["index_cols"]) == 1

    def test_pattern_detected_with_dash_separator(self):
        """分隔符为 - 时也能识别（需 >50 列）"""
        from services.file_metadata_extractor import _detect_wide_table_pattern

        regions = [f"区域{i}" for i in range(13)]  # 13 区域
        quarters = ["Q1", "Q2", "Q3", "Q4"]  # 4 季度
        # 1 索引 + 52 值列 = 53 列
        columns = [{"name": "ID", "dtype": "文本"}]
        for region in regions:
            for q in quarters:
                columns.append({"name": f"{region}-{q}", "dtype": "数值"})

        result = _detect_wide_table_pattern(columns)
        assert result is not None
        assert "13 个前缀" in result["description"]
        assert result["value_col_count"] == 52

    def test_no_pattern_without_separator(self):
        """列名无分隔符 → 返回 None"""
        from services.file_metadata_extractor import _detect_wide_table_pattern

        columns = [{"name": f"column{i}", "dtype": "文本"} for i in range(60)]
        result = _detect_wide_table_pattern(columns)
        assert result is None

    def test_no_pattern_below_threshold(self):
        """列数 ≤ 50 → 返回 None"""
        from services.file_metadata_extractor import _detect_wide_table_pattern

        columns = [{"name": f"a_{i}", "dtype": "文本"} for i in range(50)]
        result = _detect_wide_table_pattern(columns)
        assert result is None

    def test_no_pattern_low_separator_hit_rate(self):
        """分隔符命中率 < 60% → 返回 None"""
        from services.file_metadata_extractor import _detect_wide_table_pattern

        # 40 个有分隔符，30 个没有 → 命中率 57% < 60%
        columns = [{"name": f"a_{i}", "dtype": "文本"} for i in range(40)]
        columns += [{"name": f"plain{i}", "dtype": "文本"} for i in range(30)]
        result = _detect_wide_table_pattern(columns)
        assert result is None

    def test_value_range_in_description(self):
        """值列有 min/max 时描述包含范围"""
        from services.file_metadata_extractor import _detect_wide_table_pattern

        columns = [{"name": "idx", "dtype": "文本"}]
        for i in range(60):
            columns.append({
                "name": f"store_{i}", "dtype": "数值",
                "min": -5000, "max": 200000,
            })

        result = _detect_wide_table_pattern(columns)
        assert result is not None
        assert "值范围" in result["description"]
        assert "-5,000" in result["description"]

    def test_fallback_shows_head_and_tail(self):
        """模式识别失败时退化为前15+后5列名"""
        from services.file_metadata_extractor import _format_wide_table_fallback

        columns = [{"name": f"column{i}"} for i in range(100)]
        result = _format_wide_table_fallback(columns)
        assert "column0" in result
        assert "column14" in result  # 第 15 列
        assert "column99" in result  # 最后一列
        assert "省略" in result

    def test_fallback_short_list_no_ellipsis(self):
        """列数 ≤ head + tail 时只显示 head，不显示省略"""
        from services.file_metadata_extractor import _format_wide_table_fallback

        columns = [{"name": f"c{i}"} for i in range(20)]  # 20 = 15 + 5，不满足 > 条件
        result = _format_wide_table_fallback(columns)
        assert "c0" in result
        assert "c14" in result  # 前 15 列的最后一个
        assert "省略" not in result

    def test_format_standard_uses_pattern_for_wide_table(self):
        """_format_standard 对宽表用模式描述而非 8 列表格"""
        from services.file_metadata_extractor import _format_standard

        columns = [{"name": "科目", "dtype": "文本", "non_null": 100}]
        for s in ["店A", "店B", "店C"]:
            for m in range(1, 21):
                columns.append({
                    "name": f"{s}_{m}月", "dtype": "数值",
                    "non_null": 100,
                })
        meta = {
            "row_count": 50,
            "col_count": 61,
            "columns": columns,
            "preview_rows": [],
        }
        result = _format_standard("test.xlsx", "1MB", meta)
        assert "宽表" in result
        assert "前缀" in result
        # 不应出现 8 列表格格式
        assert "显示前8列" not in result

    def test_format_standard_normal_table_unchanged(self):
        """列数 ≤ 50 时仍用原来的表格格式"""
        from services.file_metadata_extractor import _format_standard

        columns = [
            {"name": f"col{i}", "dtype": "文本", "non_null": 10, "sample": [f"v{i}"]}
            for i in range(10)
        ]
        meta = {
            "row_count": 100,
            "col_count": 10,
            "columns": columns,
            "preview_rows": [],
        }
        result = _format_standard("data.csv", "500KB", meta)
        assert "| 列名 |" in result  # 表格格式
        assert "宽表" not in result
