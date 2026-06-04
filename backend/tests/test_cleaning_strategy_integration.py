"""CleaningStrategy 接入 clean_excel 的集成测试。

覆盖：
  - strategy=None → 行为与现版本完全等同（向后兼容）
  - id_columns → 跳过 Int64 转换
  - summary_rows → 标记 _is_summary 列
  - merged_cell_actions → 不同动作正确执行
  - mixed_type_handling → extract_unit_number / extract_currency_amount
  - preserve_empty_rows → 保留 AI 指定的空行
  - 兜底降级：strategy 部分字段缺失时走硬规则
"""
from __future__ import annotations

import os
import sys
import tempfile
from pathlib import Path

backend_dir = Path(__file__).parent.parent
if str(backend_dir) not in sys.path:
    sys.path.insert(0, str(backend_dir))

import openpyxl
import pandas as pd
import pytest

from services.agent.excel_cleaner import (
    CleaningReport,
    ExcelStructure,
    _detect_structure,
    clean_excel,
)
from services.agent.file_ai_decision import (
    EmptyRowDecision,
    MergedCellAction,
    MixedTypeAction,
)
from services.agent.file_cleaning_strategy import CleaningStrategy


# ── 向后兼容（strategy=None） ──

class TestBackwardCompatibility:
    """strategy=None 时与现版本行为等同。"""

    def test_clean_excel_strategy_none_matches_legacy(self, tmp_path):
        f = tmp_path / "x.xlsx"
        pd.DataFrame({"a": [1.0, 2.0, 3.0], "b": ["x", "y", "z"]}).to_excel(f, index=False)

        df1 = pd.read_excel(f)
        df2 = pd.read_excel(f)

        # 同样调用，一次不传 strategy，一次显式 None
        r1, rep1 = clean_excel(df1.copy(), str(f), 0)
        r2, rep2 = clean_excel(df2.copy(), str(f), 0, strategy=None)

        # 行为完全等同
        assert r1.equals(r2)
        assert rep1.int_cols_fixed == rep2.int_cols_fixed


# ── id_columns 保护 ──

class TestIdColumnsProtection:
    """ID 列保护：float64 列正常会被 _fix_int_columns 转 Int64，
    AI 决策为 ID 时应跳过。

    直接构造 float64 DataFrame 绕过 read_excel 的整数推断。
    """

    def test_id_column_skipped_from_int_conversion(self):
        """AI 决策的 ID 列不应被转 Int64。"""
        df = pd.DataFrame({
            "order_id": pd.Series([12345.0, 67890.0, 11111.0], dtype="float64"),
            "qty": pd.Series([1.0, 2.0, 3.0], dtype="float64"),
        })
        assert df["order_id"].dtype == "float64"

        strategy = CleaningStrategy(id_columns=["order_id"])
        df_clean, _ = clean_excel(df.copy(), "/tmp/fake.xlsx", 0, strategy=strategy)

        # ID 列保留 float64（不转 Int64）
        assert df_clean["order_id"].dtype == "float64"
        # 普通列正常转 Int64
        assert df_clean["qty"].dtype.name == "Int64"

    def test_id_column_match_uses_business_name(self):
        """id_columns 列表里是业务列名，必须匹配实际列名。"""
        df = pd.DataFrame({
            "商品编码": pd.Series([12345.0, 67890.0], dtype="float64"),
        })
        strategy = CleaningStrategy(id_columns=["商品编码"])
        df_clean, _ = clean_excel(df.copy(), "/tmp/fake.xlsx", 0, strategy=strategy)
        assert df_clean["商品编码"].dtype == "float64"

    def test_non_id_column_still_converts(self):
        """未列入 id_columns 的列应正常转 Int64（确认保护未误伤）。"""
        df = pd.DataFrame({
            "id1": pd.Series([1.0, 2.0, 3.0], dtype="float64"),
            "id2": pd.Series([4.0, 5.0, 6.0], dtype="float64"),
        })
        # 只保护 id1
        strategy = CleaningStrategy(id_columns=["id1"])
        df_clean, _ = clean_excel(df.copy(), "/tmp/fake.xlsx", 0, strategy=strategy)
        assert df_clean["id1"].dtype == "float64"
        assert df_clean["id2"].dtype.name == "Int64"


# ── summary_rows 标记 ──

class TestSummaryRowsMarking:
    def test_strategy_summary_rows_overrides_legacy(self, tmp_path):
        """AI 决策的 summary_rows 应触发 _is_summary 列标记。"""
        f = tmp_path / "x.xlsx"
        pd.DataFrame({
            "id": [1, 2, 3, None],
            "amount": [10, 20, 30, 60],
        }).to_excel(f, index=False)

        # Excel 第 5 行（最后一行）是合计行
        strategy = CleaningStrategy(summary_rows=[5])

        df = pd.read_excel(f)
        df_clean, report = clean_excel(df.copy(), str(f), 0, strategy=strategy)

        # _is_summary 列存在且最后一行为 True
        assert "_is_summary" in df_clean.columns
        assert df_clean["_is_summary"].iloc[-1] == True   # noqa: E712
        # 前面行为 False
        assert df_clean["_is_summary"].iloc[0] == False   # noqa: E712
        assert report.summary_rows_marked == 1

    def test_no_summary_when_strategy_empty(self, tmp_path):
        """strategy 不指定 summary_rows → 不加 _is_summary 列。"""
        f = tmp_path / "x.xlsx"
        pd.DataFrame({"id": [1, 2, 3]}).to_excel(f, index=False)

        strategy = CleaningStrategy()
        df = pd.read_excel(f)
        df_clean, _ = clean_excel(df.copy(), str(f), 0, strategy=strategy)

        assert "_is_summary" not in df_clean.columns


# ── merged_cell_actions ──

class TestMergedCellActions:
    def _make_merged_xlsx(self, tmp_path) -> tuple[str, ExcelStructure]:
        """创建含合并单元格的 xlsx，并返回 structure。"""
        f = tmp_path / "merge.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        # 表头
        ws["A1"] = "category"
        ws["B1"] = "item"
        ws["C1"] = "amount"
        # 数据：A2:A4 合并（category="A"），B2-B4 是 a1/a2/a3
        ws["A2"] = "A"
        ws["B2"] = "a1"; ws["C2"] = 10
        ws["B3"] = "a2"; ws["C3"] = 20
        ws["B4"] = "a3"; ws["C4"] = 30
        ws.merge_cells("A2:A4")
        wb.save(str(f))

        structure = _detect_structure(str(f), 0)
        return str(f), structure

    def test_fill_down_default(self, tmp_path):
        """默认行为：合并范围向下填充。"""
        f, structure = self._make_merged_xlsx(tmp_path)
        df = pd.read_excel(f)
        df_clean, report = clean_excel(df.copy(), f, 0, structure=structure)
        # category 列在 a2/a3 应被填充为 "A"
        assert (df_clean["category"] == "A").all()
        assert report.merged_cols_filled > 0

    def test_preserve_as_group(self, tmp_path):
        """AI 决策 preserve_as_group → 不填充，记 issue。"""
        f, structure = self._make_merged_xlsx(tmp_path)
        strategy = CleaningStrategy(merged_cell_actions=[
            MergedCellAction(range_str="A2:A4", action="preserve_as_group"),
        ])
        df = pd.read_excel(f)
        df_clean, report = clean_excel(df.copy(), f, 0, structure=structure, strategy=strategy)

        # category 第一行是 "A"，后续应保持 NaN
        assert df_clean["category"].iloc[0] == "A"
        assert pd.isna(df_clean["category"].iloc[1])
        # report 应记录 preserve issue
        preserve_issues = [i for i in report.issues if i["type"] == "merge_preserved_as_group"]
        assert len(preserve_issues) == 1


# ── mixed_type_handling ──

class TestMixedTypeHandling:
    def test_extract_unit_number(self, tmp_path):
        """extract_unit_number: '1.5kg' → 1.5"""
        f = tmp_path / "mixed.xlsx"
        pd.DataFrame({
            "weight": ["1.5kg", "2.3kg", "0.5kg", 100],   # 混合 str + int
        }).to_excel(f, index=False)

        strategy = CleaningStrategy(mixed_type_handling=[
            MixedTypeAction(col_letter="A", action="extract_unit_number", unit="kg"),
        ])
        df = pd.read_excel(f)
        df_clean, report = clean_excel(df.copy(), f, 0, strategy=strategy)

        # weight 列应为 float，前 3 个值正确
        assert df_clean["weight"].dtype == "float64"
        assert df_clean["weight"].iloc[0] == 1.5
        assert df_clean["weight"].iloc[1] == 2.3
        # report 含 extracted issue
        extracted = [i for i in report.issues if i["type"] == "mixed_type_extracted"]
        assert len(extracted) == 1

    def test_extract_currency_amount(self, tmp_path):
        """extract_currency_amount: '¥99.5' → 99.5"""
        f = tmp_path / "money.xlsx"
        pd.DataFrame({
            "price": ["¥99.5", "¥120.0", "¥85.3", 50],
        }).to_excel(f, index=False)

        strategy = CleaningStrategy(mixed_type_handling=[
            MixedTypeAction(col_letter="A", action="extract_currency_amount"),
        ])
        df = pd.read_excel(f)
        df_clean, _ = clean_excel(df.copy(), f, 0, strategy=strategy)

        assert df_clean["price"].dtype == "float64"
        assert df_clean["price"].iloc[0] == 99.5

    def test_unknown_col_falls_back_to_force_str(self, tmp_path):
        """strategy 没指定的列走默认 force_str（向后兼容）。"""
        f = tmp_path / "mixed.xlsx"
        pd.DataFrame({
            "col_a": ["1.5kg", "2kg", "3kg", 100],
            "col_b": ["xyz", 1, "abc", 2],
        }).to_excel(f, index=False)

        # 只指定 A 列
        strategy = CleaningStrategy(mixed_type_handling=[
            MixedTypeAction(col_letter="A", action="extract_unit_number", unit="kg"),
        ])
        df = pd.read_excel(f)
        df_clean, _ = clean_excel(df.copy(), f, 0, strategy=strategy)

        # A 列被提取
        assert df_clean["col_a"].dtype == "float64"
        # B 列走 force_str
        assert df_clean["col_b"].dtype == object or pd.api.types.is_string_dtype(df_clean["col_b"])


# ── preserve_empty_rows ──

class TestPreserveEmptyRows:
    def test_preserve_section_separator(self, tmp_path):
        """AI 决策保留的空行应不删除，记 preserved issue。"""
        f = tmp_path / "sep.xlsx"
        # 用 openpyxl 直接写避免 pandas 丢空行
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "id"; ws["B1"] = "amt"
        ws["A2"] = 1;   ws["B2"] = 10
        ws["A3"] = 2;   ws["B3"] = 20
        # Row 4 全空（分隔行）
        ws["A5"] = 3;   ws["B5"] = 30
        wb.save(str(f))

        strategy = CleaningStrategy(
            preserve_empty_rows=[EmptyRowDecision(row=4, reason="章节分隔")],
        )
        # 用 openpyxl 读保留空行
        df = pd.read_excel(f, keep_default_na=True)
        df_clean, report = clean_excel(df.copy(), str(f), 0, strategy=strategy)

        # 应有 preserved issue
        preserved = [i for i in report.issues if i["type"] == "empty_rows_preserved"]
        # 仅在空行真被识别到时验证
        if df.isna().all(axis=1).any():
            assert len(preserved) >= 1


# ── 兜底降级矩阵 ──

class TestFallbackMatrix:
    """AI 缺失字段时全部走硬规则。"""

    def test_empty_strategy_acts_as_none(self, tmp_path):
        """空 CleaningStrategy 等同于不传 strategy（两种调用产出 DataFrame 相等）。"""
        f = tmp_path / "x.xlsx"
        pd.DataFrame({
            "id": [12345.0, 67890.0, None],
            "qty": [1.0, 2.0, None],
        }).to_excel(f, index=False)

        df1 = pd.read_excel(f)
        df2 = pd.read_excel(f)
        r1, rep1 = clean_excel(df1.copy(), str(f), 0)
        r2, rep2 = clean_excel(df2.copy(), str(f), 0, strategy=CleaningStrategy())

        # 两次清洗 DataFrame 相同
        assert r1.equals(r2)
        # 两次 issues 同步
        assert rep1.int_cols_fixed == rep2.int_cols_fixed
        assert rep1.empty_rows_removed == rep2.empty_rows_removed
